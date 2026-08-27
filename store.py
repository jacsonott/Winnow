"""SQLite storage layer for Winnow.

Design notes
------------
* One case = one SQLite file. Each imported CSV becomes its own `src_<id>` table
  with an explicit `rid INTEGER PRIMARY KEY` so row identity is stable forever.
* Source tables are never mutated. Tags, notes and layouts live in sidecar tables
  keyed by (source_id, rid), so re-importing the same file keeps your work.
* Filtering/sorting is materialised once into a temp-attached `v.view_N` table of
  (pos, rid). The grid then pages with `WHERE pos BETWEEN ? AND ?`, which stays
  O(window) no matter how deep you scroll. Naive LIMIT/OFFSET does not.
"""

from __future__ import annotations

import contextlib
import csv
import datetime
import errno
import fnmatch
import getpass
import hashlib
import io
import itertools
import json
import os
import re
import shutil
import socket
import sqlite3
import tempfile
import threading
import time
from typing import Any, Iterable, Iterator, Sequence

from openpyxl import Workbook

try:  # POSIX only — see sweep_orphan_views for why Windows needs no substitute
    import fcntl
except ImportError:  # pragma: no cover - Windows
    fcntl = None  # type: ignore[assignment]

import structparse  # noqa: F401 — registers the JSON/XML extraction ops into timeparse.OPERATIONS
import timeparse

BATCH = 20_000
SAMPLE_ROWS = 500
TRIGRAM_MIN_LEN = 3  # the trigram FTS5 tokenizer can't index anything shorter — see _fts_like_pattern
SARGABLE_OPS = {"equals", "in"}  # ops a plain single-column B-tree index actually accelerates (see _ensure_column_index_building)
# SQLite gained LIKE/GLOB pushdown onto fts5 trigram tables in 3.45.0 —
# the query form build_fts's detail=none index depends on. On an older
# runtime the same LIKE still returns correct rows (SQLite just evaluates
# it by scanning the fts table's docs — fallback speed), so nothing breaks;
# building an index that can't accelerate anything is just wasted disk,
# hence _ensure_fts_building gates on this.
TRIGRAM_LIKE_MIN_SQLITE = (3, 45, 0)
# search_all_sources stops counting a source's matches here and reports
# `capped` instead of an exact number. The modal it feeds only ranks
# "which tables hit, and roughly how hard" before you open one — an exact
# 480,113 is no more useful than "1,000+", and getting it costs a full
# count over every matching row on a source whose index isn't built yet.
SEARCH_ALL_COUNT_CAP = 1000
# How many terms a "paste a list" sweep will break down per source before it
# reports the union count alone. The breakdown costs one extra capped count
# per term on each source that matched at all (see
# _iter_search_all_sources), which is what makes an entry-per-term list
# possible; a list long enough to make that N+1 shape hurt is one where the
# per-term rows would be unreadable anyway. Well clear of a realistic IOC
# paste — it exists to bound an accident (a whole wordlist pasted in), not
# to ration normal use.
SEARCH_ALL_TERM_BREAKDOWN_MAX = 250

# The extensions a directory import (scan_import_directory) recognizes by
# default — every format ingest_csv/ingest_json already know how to read.
# Deliberately mirrors app.js's file-picker <input accept> list rather than
# inventing a separate notion of "supported" — one format list, two places
# it has to be spelled out (a browser <input> can't read a Python constant).
DEFAULT_IMPORT_EXTENSIONS = {".csv", ".tsv", ".txt", ".psv", ".json", ".jsonl", ".ndjson"}
# The SQLite set is separate because these files can't be bulk-imported
# blind — which tables to pull out is a per-file choice (see
# preview_sqlite_tables) — so directory import ignores them while the
# server-disk file browser still lists them.
SQLITE_IMPORT_EXTENSIONS = {".db", ".sqlite", ".sqlite3", ".db-wal"}
# scan_import_directory stops walking once matched+excluded together hit
# this many entries — same "cap and say so" reasoning as
# SEARCH_ALL_COUNT_CAP, guarding against an analyst accidentally pointing
# the scan at a triage source volume instead of its (much smaller) output
# folder.
MAX_SCAN_RESULTS = 5000

# TA first: triage reaches for "this is the adversary" far more often than
# "this is fine", so the most-used tag sits on hotkey 1 under the left hand.
# Order here is display order everywhere tags render (ribbon, menus, undo).
DEFAULT_TAGS = [
    ("TA", "#c0392b", "1"),
    ("Suspicious", "#d68a2e", "2"),
    ("Benign", "#5d8a66", "3"),
]

META_SCHEMA = """
CREATE TABLE IF NOT EXISTS sources (
    id          INTEGER PRIMARY KEY,
    name        TEXT NOT NULL,
    path        TEXT,
    table_name  TEXT NOT NULL,
    row_count   INTEGER NOT NULL DEFAULT 0,
    columns     TEXT NOT NULL,          -- json: [{name, type}]
    file_hash   TEXT,
    imported_at TEXT,
    has_fts     INTEGER NOT NULL DEFAULT 0,
    nickname    TEXT                    -- analyst display name; `name` stays the imported file's identity
);
CREATE TABLE IF NOT EXISTS tag_defs (
    id     INTEGER PRIMARY KEY,
    name   TEXT NOT NULL,
    color  TEXT NOT NULL,
    hotkey TEXT
);
CREATE TABLE IF NOT EXISTS row_tags (
    source_id INTEGER NOT NULL,
    rid       INTEGER NOT NULL,
    tag_id    INTEGER NOT NULL,
    PRIMARY KEY (source_id, rid, tag_id)
) WITHOUT ROWID;
CREATE INDEX IF NOT EXISTS ix_row_tags_tag ON row_tags(source_id, tag_id);
CREATE TABLE IF NOT EXISTS row_notes (
    source_id INTEGER NOT NULL,
    rid       INTEGER NOT NULL,
    note      TEXT NOT NULL,
    PRIMARY KEY (source_id, rid)
) WITHOUT ROWID;
CREATE TABLE IF NOT EXISTS layouts (
    source_id INTEGER PRIMARY KEY,
    payload   TEXT NOT NULL
);
CREATE TABLE IF NOT EXISTS saved_views (
    id        INTEGER PRIMARY KEY,
    source_id INTEGER NOT NULL,
    name      TEXT NOT NULL,
    payload   TEXT NOT NULL,
    saved_at  TEXT
);
-- Legacy — presets are saved filters now (workspace-level, cross-case; see
-- Store.pop_legacy_presets and CLAUDE.md). Table kept only so a case file
-- from before that change has somewhere for its old presets to still be
-- read from and migrated out of on next open; nothing writes here anymore.
CREATE TABLE IF NOT EXISTS filter_presets (
    id         INTEGER PRIMARY KEY,
    name       TEXT NOT NULL,
    col_sig    TEXT NOT NULL,          -- sha256 of sorted, lowercased column names
    col_names  TEXT NOT NULL,          -- json array, original casing, for display/diffing
    payload    TEXT NOT NULL,          -- json: {filter_tree, sort, search, search_mode, search_terms}
    created_at TEXT
);
CREATE INDEX IF NOT EXISTS ix_filter_presets_sig ON filter_presets(col_sig);
CREATE TABLE IF NOT EXISTS merges (
    id         INTEGER PRIMARY KEY,
    name       TEXT NOT NULL,
    source_ids TEXT NOT NULL,          -- json array of real source ids, creation order
    created_at TEXT
);
CREATE TABLE IF NOT EXISTS open_tabs (
    source_id INTEGER PRIMARY KEY      -- signed, same convention as elsewhere: negative = merge id
);
-- The SQL pane's named sub-tabs. In the case file rather than workspace/ or
-- localStorage because a worked-out query is analysis *about this evidence*
-- ("the join that pulls 4624s against the RDP source"), not a UI preference
-- like tl.sidebar and not cross-case reusable like a saved filter — it
-- should travel with the case when it's handed to another analyst. Still
-- only ever SELECTs the analyst typed, so this stays within invariant #1:
-- no source table is touched.
CREATE TABLE IF NOT EXISTS sql_tabs (
    id   INTEGER PRIMARY KEY,
    name TEXT NOT NULL,
    sql  TEXT NOT NULL DEFAULT '',
    pos  INTEGER NOT NULL DEFAULT 0     -- left-to-right strip order; ties break by id
);
-- Derived (computed) column definitions. The VALUES live in a per-source
-- sidecar table drv_<source_id> (rid INTEGER PRIMARY KEY, one TEXT column
-- per derived column) created lazily by add_derived_column — the
-- row_tags/row_notes sidecar pattern, so src_<id> is never mutated
-- (invariant #1). sources.columns stays base-only (it is the imported
-- file's identity — merge signatures and workspace header-set keys hang
-- off it); derived entries are merged into src["columns"] at read time.
CREATE TABLE IF NOT EXISTS derived_columns (
    id             INTEGER PRIMARY KEY,
    source_id      INTEGER NOT NULL,
    name           TEXT NOT NULL,      -- analyst-chosen; collision-checked vs base+derived
    input_column   TEXT NOT NULL,      -- a base column name
    op_id          TEXT NOT NULL,      -- timeparse.OPERATIONS key
    params         TEXT NOT NULL DEFAULT '{}',
    status         TEXT NOT NULL DEFAULT 'building',  -- building | ready | partial | error
    parse_failures INTEGER,            -- non-empty inputs that produced NULL; set at backfill end
    created_at     TEXT
);
-- No index on source_id on purpose: this table holds one row per derived
-- column in the whole case (tens at the very most), so a scan is one page
-- either way, and at the 64KB page size _tune sets, an index would cost
-- more file than it could ever save. The lookup IS on a hot path
-- (_source_lite_on, once per page fetch) — it's just a hot path over a
-- table that fits in a single page.
-- Per-case settings (first key: ts_format, the case-level datetime display
-- default). In the case file, not workspace/, because "how this case's
-- timestamps read" should travel with the case when it's handed to
-- another analyst — same reasoning as sql_tabs.
CREATE TABLE IF NOT EXISTS case_settings (
    key   TEXT PRIMARY KEY,
    value TEXT NOT NULL
);
"""

IDENT_OK = re.compile(r"^[A-Za-z_][A-Za-z0-9_]*$")
NUM_RE = re.compile(r"^-?\d+(\.\d+)?$")
# Anything that looks like a leading ISO-ish or US date; used only to pick a
# default sort column and to hint the UI, never to rewrite the stored value.
# Month-name timestamps ("JUN 23 2026 00:11:00", "23 Jun 2026", "June 23,
# 2026 5:11 PM") — the third recognized family, name-first and day-first
# both, since exporters disagree. The alternation is strict (a weekday or a
# random word never matches) and _TS_MONTHS is the number lookup the
# parsers share; it's a superset of the alternation only by "sept".
_TS_MONTH_NAME = (r"(?:jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|june?|july?"
                  r"|aug(?:ust)?|sep(?:t(?:ember)?)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)")
_TS_MONTHS = {
    "jan": 1, "feb": 2, "mar": 3, "apr": 4, "may": 5, "jun": 6,
    "jul": 7, "aug": 8, "sep": 9, "sept": 9, "oct": 10, "nov": 11, "dec": 12,
    "january": 1, "february": 2, "march": 3, "april": 4, "june": 6,
    "july": 7, "august": 8, "september": 9, "october": 10,
    "november": 11, "december": 12,
}
DATE_RE = re.compile(
    r"^\d{4}-\d{2}-\d{2}[ T]|^\d{1,2}/\d{1,2}/\d{4}"
    r"|^" + _TS_MONTH_NAME + r"\s+\d{1,2}\s*,?\s*\d{4}"
    r"|^\d{1,2}\s+" + _TS_MONTH_NAME + r"\s*,?\s*\d{4}",
    re.IGNORECASE)


def q(ident: str) -> str:
    """Quote an identifier for interpolation into SQL."""
    return '"' + ident.replace('"', '""') + '"'


def _blob_expr(cols: list[str]) -> str:
    """Every column concatenated into one search target — the same
    expression backs both sides of substring search: the src_<id>_doc view
    the trigram index is built over (build_fts), and every LIKE-scan
    fallback (Contains mode when no index is ready yet or never got one,
    Advanced mode's whole-query LIKE fallback, and a single Advanced-mode
    term _fts_like_pattern can't route through the index). Keeping them
    the same expression is what makes the indexed and fallback paths
    return identical results."""
    return " || ' ' || ".join(f"COALESCE({q(c)},'')" for c in cols)


def _fts_like_pattern(term: str) -> str | None:
    """The bare-LIKE pattern that routes a Contains/Advanced term through
    the trigram index (`doc LIKE ?` on the fts table), or None when this
    term has to take the blob-LIKE fallback scan instead. Two reasons a
    term can't use the index, both verified empirically:

    - Under TRIGRAM_MIN_LEN (3 chars) the trigram tokenizer has nothing to
      look up, so the "index-assisted" form degrades to scanning anyway.
    - SQLite only pushes a *bare* `LIKE ?` down onto a trigram table —
      adding an ESCAPE clause turns the plan back into a full scan. So the
      pattern here is deliberately unescaped, which is only safe when the
      term contains no LIKE wildcards: a `%`/`_` in the term would match
      as a wildcard instead of literally (a superset of right answers —
      silently wrong). Those terms go to the fallback, which escapes them.
      A backslash is fine unescaped — LIKE has no escape character at all
      unless an ESCAPE clause says so — which matters here because
      Windows paths make `\\` the single most-searched special character
      in this tool."""
    if len(term) < TRIGRAM_MIN_LEN or "%" in term or "_" in term:
        return None
    return f"%{term}%"


def _numeric_expr(sql_expr: str) -> str:
    """A REAL-valued SQL expression for a 'number'-typed column that is NULL
    for anything that doesn't actually look numeric (same pattern as
    infer_type/NUM_RE), instead of using bare CAST(... AS REAL).

    Columns are stored as TEXT (see CLAUDE.md — evidence fidelity beats sort
    elegance). SQLite's own CAST(text AS REAL) silently returns 0.0 for
    non-numeric text, so a stray "N/A"/blank/typo in an otherwise-numeric
    column is indistinguishable from a genuine 0 in a numeric sort or a
    `>`/`<` filter. Gating the cast behind the same regex ingest-time type
    inference uses means bad values come out NULL instead — SQLite sorts
    NULL to one edge rather than scattering it through real data, and a NULL
    comparison is never true, so garbage is visibly excluded rather than
    silently masquerading as zero."""
    return f"(CASE WHEN {sql_expr} REGEXP '{NUM_RE.pattern}' THEN CAST({sql_expr} AS REAL) ELSE NULL END)"


def _regexp(pattern: str, value: Any) -> bool:
    if value is None:
        return False
    try:
        return re.search(pattern, str(value), re.IGNORECASE) is not None
    except re.error:
        return False


# Same two shapes DATE_RE recognizes at ingest, capturing time-of-day too
# (all optional groups — a bare date still matches). Kept in sync with
# app.js's TS_ISO_RE/TS_US_RE (parseTimestamp) by hand — there's no shared
# module between the two, just the same two format families.
_TS_ISO_RE = re.compile(r"^(\d{4})-(\d{2})-(\d{2})(?:[ T](\d{2}):(\d{2})(?::(\d{2}))?)?")
_TS_US_RE = re.compile(r"^(\d{1,2})/(\d{1,2})/(\d{4})(?:[ ,]+(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM|am|pm)?)?")
_TS_MONTH_RE = re.compile(
    r"^(?:(" + _TS_MONTH_NAME + r")\s+(\d{1,2})|(\d{1,2})\s+(" + _TS_MONTH_NAME + r"))"
    r"\s*,?\s*(\d{4})"
    r"(?:[ ,]+(\d{1,2}):(\d{2})(?::(\d{2}))?\s*(AM|PM|am|pm)?)?",
    re.IGNORECASE)


def _month_parts(m):
    """(y, mo, d, hh, mi, ss) from a _TS_MONTH_RE match — the AM/PM rules
    identical to the US branch's."""
    mo = _TS_MONTHS[(m.group(1) or m.group(4)).lower()]
    d = int(m.group(2) or m.group(3))
    hh = int(m.group(6) or 0)
    ampm = m.group(9)
    if ampm:
        if ampm.lower() == "pm" and hh < 12:
            hh += 12
        if ampm.lower() == "am" and hh == 12:
            hh = 0
    return int(m.group(5)), mo, d, hh, int(m.group(7) or 0), int(m.group(8) or 0)


def _day_bucket(raw: Any) -> str | None:
    """Registered as the SQL function DAY_BUCKET(x) — collapses a
    'datetime'-typed value to its calendar day ("YYYY-MM-DD"), so grouping
    by a timestamp column buckets by day instead of by exact timestamp
    (which, at second/millisecond precision, would put nearly every row in
    its own group of one). None for anything that doesn't match either
    recognized shape, same "leave it alone rather than guess" rule
    parseTimestamp follows client-side — those rows bucket together under
    a NULL group instead of being silently mis-grouped."""
    if raw is None:
        return None
    s = str(raw).strip()
    m = _TS_ISO_RE.match(s)
    if m:
        return f"{m.group(1)}-{m.group(2)}-{m.group(3)}"
    m = _TS_US_RE.match(s)
    if m:
        return f"{int(m.group(3)):04d}-{int(m.group(1)):02d}-{int(m.group(2)):02d}"
    m = _TS_MONTH_RE.match(s)
    if m:
        y, mo, d, _, _, _ = _month_parts(m)
        return f"{y:04d}-{mo:02d}-{d:02d}"
    return None


def _ts_normalize(raw: Any) -> str | None:
    """Registered as the SQL function TS_NORMALIZE(x) — a zero-padded,
    lexicographically-sortable "YYYY-MM-DD HH:MM:SS" for a value in either
    shape _TS_ISO_RE/_TS_US_RE recognize, missing time-of-day treated as
    midnight. Powers the timeframe filter's range check (TS_NORMALIZE(col)
    BETWEEN ? AND ?, both bounds run through this same function) — a bare
    text/numeric comparison on the raw stored value would sort the US
    "M/D/YYYY" shape wrong (lexicographic order isn't chronological order
    for it) and mismatch on missing leading zeros either way. None for
    anything unparseable, same "leave it alone" rule as DAY_BUCKET/
    parseTimestamp — an unparseable value can't be judged in-range or not,
    so it's excluded rather than guessed at."""
    if raw is None:
        return None
    s = str(raw).strip()
    m = _TS_ISO_RE.match(s)
    if m:
        y, mo, d, h, mi, sec = m.groups()
        return f"{y}-{mo}-{d} {h or '00'}:{mi or '00'}:{sec or '00'}"
    m = _TS_US_RE.match(s)
    if m:
        mo, d, y, h, mi, sec, ampm = m.groups()
        hh = int(h) if h else 0
        if ampm:
            if ampm.lower() == "pm" and hh < 12:
                hh += 12
            if ampm.lower() == "am" and hh == 12:
                hh = 0
        return f"{int(y):04d}-{int(mo):02d}-{int(d):02d} {hh:02d}:{int(mi or 0):02d}:{int(sec or 0):02d}"
    m = _TS_MONTH_RE.match(s)
    if m:
        y, mo, d, hh, mi, sec = _month_parts(m)
        return f"{y:04d}-{mo:02d}-{d:02d} {hh:02d}:{mi:02d}:{sec:02d}"
    return None


# Microseconds between 1601-01-01 (the Windows/WebKit epoch) and 1970-01-01
# (the Unix epoch) — Chromium stores every *_time/*_utc column (History's
# urls.last_visit_time, visits.visit_time, Cookies' creation_utc/expires_utc,
# downloads' start_time/end_time, ...) as microseconds since 1601-01-01.
WEBKIT_EPOCH_OFFSET_US = 11_644_473_600_000_000


def _webkit_to_iso(value: Any) -> str | None:
    """Chrome/WebKit timestamp -> ISO 8601 UTC string, or None if `value`
    doesn't actually look like one in that format. Used two ways: to
    convert a column's values on SQLite-table ingest (opt-in, see
    ingest_sqlite_table), and — the same function, not just the same idea
    — to screen candidate columns for that option's default-checked state
    in preview_sqlite_tables (a column counts as "likely a WebKit
    timestamp" if most of its sampled values round-trip through here to a
    plausible date)."""
    try:
        v = int(value)
    except (TypeError, ValueError):
        return None
    unix_us = v - WEBKIT_EPOCH_OFFSET_US
    try:
        dt = datetime.datetime.fromtimestamp(unix_us / 1_000_000, tz=datetime.timezone.utc)
    except (OverflowError, OSError, ValueError):
        return None
    if dt.year < 1990 or dt.year > 2100:  # in range arithmetically but not a plausible date
        return None
    return dt.strftime("%Y-%m-%dT%H:%M:%S.%f")[:-3] + "Z"



# The pseudo-column that means "group these rows by the tags on them" rather
# than by anything in the file (see group_summary). It's in the reserved set
# below so sanitize_columns renames a real header of this name at ingest,
# which is what makes the sentinel unambiguous with a column an analyst can
# actually see.
TAG_GROUP_COLUMN = "__tag__"

# Column names that can't be used as-is: "rid" collides with every src_<id>
# table's own primary-key column, and "rank"/"rowid" are FTS5-reserved \u2014
# CREATE VIRTUAL TABLE ... USING fts5(...) rejects a column literally named
# either of those (case-insensitive) with "reserved fts5 column name", which
# previously surfaced as build_fts() raising well after the source table
# and its rows were already committed \u2014 a CSV with a plausible header like
# "Rank" (a real column name in e.g. some log formats) failed the *entire*
# import with a confusing error, even though the data itself was fine.
RESERVED_COLUMN_NAMES = {"rid", "rank", "rowid", TAG_GROUP_COLUMN}


def sanitize_columns(raw: list[str]) -> list[str]:
    """Make header names unique and safe to quote. Blank headers become col_N.

    Dedups against every name actually emitted so far, not just first-seen
    original names \u2014 a header that already contains the disambiguation
    pattern (e.g. ["Name", "Name", "Name_1"], plausible for CSVs that have
    already been through another tool's own dedup pass) used to collide:
    the second "Name" became "Name_1", which then collided outright with the
    third column's own literal "Name_1", and CREATE TABLE failed with
    "duplicate column name" \u2014 aborting an otherwise perfectly valid import."""
    out: list[str] = []
    seen: dict[str, int] = {}
    used: set[str] = set()
    for i, name in enumerate(raw):
        clean = (name or "").strip().lstrip("\ufeff") or f"col_{i + 1}"
        if clean.lower() in RESERVED_COLUMN_NAMES:
            clean = f"{clean}_1"
        base = clean
        while clean.lower() in used:
            seen[base.lower()] = seen.get(base.lower(), 0) + 1
            clean = f"{base}_{seen[base.lower()]}"
        used.add(clean.lower())
        out.append(clean)
    return out


def column_signature(cols: list[str]) -> str:
    """Order/case-independent fingerprint of a column-name set — used to
    check two sources have "the same" columns for merge eligibility."""
    return hashlib.sha256("\x1f".join(sorted(c.strip().lower() for c in cols)).encode()).hexdigest()


def infer_type(values: Iterable[str]) -> str:
    vals = [v for v in values if v not in (None, "")]
    if not vals:
        return "text"
    if all(NUM_RE.match(v) for v in vals):
        return "number"
    if sum(1 for v in vals if DATE_RE.match(v)) >= max(1, int(len(vals) * 0.8)):
        return "datetime"
    return "text"


def _json_leaf_text(value: Any) -> str:
    """A JSON value at a point flattening stops (see _flatten_json) ->
    TEXT, same all-values-are-TEXT convention as CSV ingest. An object or
    array is JSON-stringified whole rather than dropped or half-flattened —
    evidence fidelity over tidiness, same reasoning CLAUDE.md gives for not
    typing columns at ingest."""
    if value is None:
        return ""
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (dict, list)):
        return json.dumps(value, ensure_ascii=False)
    return str(value)


def _flatten_json(obj: dict, max_depth: int | None) -> dict[str, str]:
    """Flattens nested OBJECTS into dot-notation keys — {"user": {"name":
    "a"}} -> {"user.name": "a"} — but never arrays: a list value is always
    JSON-stringified as one column's text, at whatever depth it's found,
    per the ingest UI's own "flatten" explanation (an array's length can
    vary record to record, so index-expanding it would make the column set
    itself vary; a nested object's key set is comparatively stable).

    max_depth=0 flattens nothing — one column per top-level key, "none"
    mode, matching CSV's one-column-per-header shape exactly. max_depth=None
    flattens without limit ("full"). Any other int is how many levels of
    *object* nesting to unfold before stringifying whatever's left ("depth
    N") — e.g. max_depth=1 unfolds `user.name` but leaves `user.address`
    (a further-nested object) as one JSON-text column."""
    out: dict[str, str] = {}

    def walk(value: Any, key_path: str, depth: int) -> None:
        if isinstance(value, dict) and (max_depth is None or depth < max_depth):
            if not value:
                out[key_path] = "{}"
                return
            for k, v in value.items():
                walk(v, f"{key_path}.{k}" if key_path else str(k), depth + 1)
        else:
            out[key_path] = _json_leaf_text(value)

    for k, v in obj.items():
        walk(v, str(k), 0)
    return out


def _iter_json_records(path: str) -> Iterable[dict]:
    """Yields one dict per record. `.jsonl`/`.ndjson` (one JSON value per
    line — the genuinely streamable shape, never loaded fully into memory)
    vs a single `.json` document (a JSON array of records, a single record
    object, or — tolerated rather than rejected — a bare scalar/array of
    scalars, wrapped as {"value": ...} so it still lands as one column
    instead of erroring out). A whole `.json` document has to be parsed in
    one shot (no generic streaming parser in the standard library), so
    unlike CSV/JSONL ingest, memory use scales with file size for that
    shape — acceptable at the scale this tool already targets (CLAUDE.md's
    "Known limits" already documents ingest as not built for huge files
    without extra tooling)."""
    ext = os.path.splitext(path)[1].lower()
    if ext in (".jsonl", ".ndjson"):
        with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                obj = json.loads(line)
                yield obj if isinstance(obj, dict) else {"value": obj}
        return
    with open(path, "r", encoding="utf-8-sig", errors="replace") as f:
        data = json.load(f)
    if isinstance(data, list):
        for obj in data:
            yield obj if isinstance(obj, dict) else {"value": obj}
    elif isinstance(data, dict):
        yield data
    else:
        yield {"value": data}


class IngestCancelled(Exception):
    """An ingest job was cancelled — raised out of the ingest_* method
    after the partial source has been dropped. Callers that don't pass a
    `cancel` callable never see it."""


class OpCancelled(Exception):
    """A registered cancellable operation (view/timeline build, group
    summary) was interrupted via cancel_op. server.py maps it to HTTP 499 —
    not an error, not the analyst's fault, keep what you had."""


# ------------------------------------------------- the views scratch database

# Naming for the per-Store views database (invariant #3). The prefix is the
# only thing identifying one of these files as ours once the process that
# made it is gone, so sweep_orphan_views matches on it — don't change it
# without leaving a sweep for the old shape behind.
VIEWS_PREFIX = "winnow-views-"
VIEWS_SUFFIX = ".db"
VIEWS_DB_SUFFIXES = ("", "-wal", "-shm")  # the three files a WAL database is


def _preferred_views_dir() -> str | None:
    """Where a *new* views database goes: /dev/shm when it's usable (tmpfs,
    Linux-only — see Store.__init__ for why that's worth 20-100%), else None
    for tempfile's own default (the platform tempdir: /tmp, or on Windows
    %TMP%/%TEMP%, which for an account with neither set is C:\\Windows\\Temp)."""
    if os.path.isdir("/dev/shm") and os.access("/dev/shm", os.W_OK):
        return "/dev/shm"
    return None


def _views_dirs() -> list[str]:
    """Every directory a views database could be sitting in — the preferred
    one *and* the platform tempdir, not just today's choice. A machine can
    have accumulated files in both (a /dev/shm that wasn't writable on some
    earlier run, a container where it appeared later), and the whole point of
    the sweep is to find files an earlier process left behind."""
    dirs = []
    preferred = _preferred_views_dir()
    if preferred:
        dirs.append(preferred)
    with contextlib.suppress(Exception):
        tmp = tempfile.gettempdir()
        if tmp not in dirs:
            dirs.append(tmp)
    return dirs


def _views_file_is_orphaned(path: str) -> bool:
    """Whether no live Store still owns `path`. Every Store holds an
    exclusive flock on its own views file for its entire life
    (Store.__init__), so a file this can lock has no owner left.

    Fail-safe in both directions, which is what makes it usable against a
    file another process might be mid-query on: an unreadable candidate, or
    one whose lock is held (or whose filesystem doesn't implement flock at
    all — the same OSError, and the same answer), is reported as still in
    use and therefore never deleted. On Windows there is no flock and none
    is needed: the OS refuses to unlink a file any process has open, so the
    delete in sweep_orphan_views is itself the liveness test there."""
    if fcntl is None:  # pragma: no cover - Windows
        return True
    try:
        fd = os.open(path, os.O_RDWR)
    except OSError:
        return False
    try:
        fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError:
        return False
    finally:
        os.close(fd)  # releases the probe's own lock immediately
    return True


def sweep_orphan_views() -> dict:
    """Delete views databases left behind by Winnow processes that are no
    longer running, and report `{"removed": n, "bytes_freed": n}`.

    Store.close() removes its own — but only a *clean* exit ever reaches it.
    A SIGKILL, a power loss, an OOM kill, or (before server.py grew its
    lifespan shutdown hook) simply closing the terminal did not, and every
    one of those stranded a `winnow-views-<random>.db` plus its `-wal`/`-shm`
    siblings that nothing would ever clean up again. On Linux they sit in
    /dev/shm and cost RAM until reboot; on macOS/Windows they sit in the
    platform tempdir — `C:\\Windows\\Temp` for an account with no TMP/TEMP
    set — and accumulate across reboots, one set per killed session, at
    whatever size that session's biggest materialised view reached.

    Deleting another process's file is only safe because views are
    re-derivable scratch that dies with the process anyway (invariant #3)
    *and* because _views_file_is_orphaned proves the owner is gone first —
    a second, live Winnow on the same machine keeps its own file. Called
    once at server startup; never fatal (any error just leaves the file)."""
    removed, bytes_freed = 0, 0
    for d in _views_dirs():
        try:
            entries = os.listdir(d)
        except OSError:
            continue  # tempdir unreadable or gone — nothing to sweep, not an error
        for name in entries:
            # -wal/-shm are removed with the .db they belong to, below.
            if not (name.startswith(VIEWS_PREFIX) and name.endswith(VIEWS_SUFFIX)):
                continue
            base = os.path.join(d, name)
            if not _views_file_is_orphaned(base):
                continue
            for suffix in VIEWS_DB_SUFFIXES:
                path = base + suffix
                try:
                    size = os.path.getsize(path)
                except OSError:
                    continue  # no -wal/-shm for this one; normal
                try:
                    os.remove(path)
                except OSError:
                    break  # can't remove the base => don't touch its siblings
                removed += 1
                bytes_freed += size
        for name in entries:
            # A -wal/-shm whose .db is already gone. Only a dead owner can
            # leave that shape (close() removes the .db first, and a live
            # Store always has its .db), so there's nothing to probe.
            if not name.startswith(VIEWS_PREFIX) or not name.endswith(("-wal", "-shm")):
                continue
            path = os.path.join(d, name)
            if os.path.exists(path[:-4]):  # both suffixes are 4 chars
                continue
            try:
                size = os.path.getsize(path)
                os.remove(path)
            except OSError:
                continue
            removed += 1
            bytes_freed += size
    return {"removed": removed, "bytes_freed": bytes_freed}


# ----------------------------------------------------- the case "in use" lock

# A case file open in one Winnow is not safe to open in a second one: SQLite's
# own WAL locking keeps the *file* consistent, but nothing in this app
# invalidates a second process's caches, its frontend's row counts, or its
# idea of which tabs are open — and a long write (compact() is minutes by its
# own docstring) will simply fail the other process's writes once the 5s busy
# timeout runs out. On a network share it is worse than that: WAL needs a
# shared-memory -shm mapping and does not work over SMB/NFS at all, which is
# exactly the setup two analysts would use to collide.
#
# So each Store drops an advisory marker next to its case file and a second
# Winnow probes it before opening. Advisory is the operative word — this
# refuses nothing on its own. probe_case_lock() reports, server.py decides,
# and the analyst always has a way through.
CASE_LOCK_SUFFIX = ".winnow-lock"
CASE_LOCK_HEARTBEAT_SEC = 30
# Five missed beats. Generous on purpose, because the two wrong answers are
# not symmetric: calling a live lock stale silently permits the collision
# this whole mechanism exists to catch, while calling a dead lock live costs
# one extra click on the prompt's "Open anyway".
CASE_LOCK_STALE_AFTER_SEC = 150


def case_lock_path(case_path: str) -> str:
    return os.path.abspath(case_path) + CASE_LOCK_SUFFIX


def _read_lock_record(path: str) -> dict:
    """The marker's contents, or {} if there aren't any usable ones.

    A torn read is expected and benign: _CaseLock._write rewrites the file in
    place (it must — a write-to-temp-and-rename would move the flock onto an
    unlinked inode), so a probe can catch a half-written record. {} then means
    "no heartbeat signal", and probe_case_lock falls back to the flock."""
    try:
        with open(path, "r", encoding="utf-8") as f:
            rec = json.load(f)
        return rec if isinstance(rec, dict) else {}
    except (OSError, ValueError):
        return {}


def _flock_state(path: str) -> str:
    """'held' | 'free' | 'unknown' — what flock says about an existing marker.

    'unknown' covers every way a filesystem can decline to answer (no fcntl
    at all on Windows, ENOLCK/EOPNOTSUPP on a share, an unreadable file).
    That's a third state rather than an error because the heartbeat is the
    other half of the answer and is checked independently."""
    if fcntl is None:  # pragma: no cover - Windows
        return "unknown"
    try:
        fd = os.open(path, os.O_RDWR)
    except OSError:
        return "unknown"
    try:
        fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
    except OSError as e:
        return "held" if e.errno in (errno.EACCES, errno.EAGAIN, errno.EWOULDBLOCK) else "unknown"
    finally:
        os.close(fd)  # releases the probe's own lock immediately
    return "free"


def _heartbeat_state(rec: dict) -> tuple[str, float | None]:
    """'live' | 'stale' | 'unknown', plus the beat's age in seconds.

    Wall-clock comparison across two machines, which is the point — it is the
    only signal that survives a filesystem without working flock. Clock skew
    is therefore real: a holder whose clock is far behind reads as stale. The
    skew would have to exceed CASE_LOCK_STALE_AFTER_SEC to matter, and the
    consequence is a missed prompt, not a wrong write."""
    hb = rec.get("heartbeat_at")
    if not isinstance(hb, (int, float)) or isinstance(hb, bool):
        return "unknown", None
    age = max(0.0, time.time() - float(hb))
    return ("live" if age <= CASE_LOCK_STALE_AFTER_SEC else "stale"), age


def probe_case_lock(case_path: str) -> dict | None:
    """Who, if anyone, already has this case open. None means free.

    Two independent signals, and either one alone is enough to report a
    conflict: the flock (exact, instant, local filesystems) and the heartbeat
    (works anywhere a file can be read, which is what covers the share). A
    marker whose flock is free *and* whose heartbeat has gone stale is the
    residue of a killed process and reports free — the next Store overwrites
    it in place.

    Deliberately biased toward reporting a conflict, the opposite of
    _views_file_is_orphaned's bias: there, a wrong answer deletes a live
    process's file, so it errs toward "in use"; here a wrong answer only
    raises a prompt the analyst can click through."""
    path = case_lock_path(case_path)
    if not os.path.exists(path):
        return None
    rec = _read_lock_record(path)
    flock_state = _flock_state(path)
    hb_state, age = _heartbeat_state(rec)
    if flock_state == "held":
        evidence = "flock"
    elif hb_state == "live":
        evidence = "heartbeat"
    elif flock_state == "unknown" and hb_state == "unknown":
        # The marker exists but says nothing and can't be locked — most
        # likely unreadable. Report it rather than assume it's junk.
        evidence = "unreadable"
    else:
        return None
    return {
        "host": rec.get("host"),
        "user": rec.get("user"),
        "pid": rec.get("pid"),
        "started_at": rec.get("started_at"),
        "heartbeat_age_sec": None if age is None else round(age, 1),
        "evidence": evidence,
        "lock_path": path,
    }


def describe_case_lock(holder: dict) -> str:
    """One line naming the holder, for the CLI refusal and the API detail."""
    who = holder.get("user") or "an unknown user"
    where = holder.get("host") or "an unknown host"
    bits = f"{who}@{where}"
    if holder.get("pid"):
        bits += f" (pid {holder['pid']})"
    if holder.get("started_at"):
        bits += f", open since {holder['started_at']}"
    if holder.get("evidence") == "unreadable":
        return f"A lock file exists at {holder['lock_path']} but can't be read"
    age = holder.get("heartbeat_age_sec")
    if age is not None:
        bits += f", last seen {int(age)}s ago"
    return bits


class _CaseLock:
    """The marker one Store holds for its case file, for its whole life.

    Modelled on the views file's flock (invariant #3) and shares its two
    liveness guarantees: server.py's _lifespan shutdown hook reaches
    Store.close() on every clean exit, and a hard kill leaves something the
    next process can recognise as dead — here the stale heartbeat, since a
    case marker can't be swept blindly the way a scratch views file can.

    Best-effort by construction. Every failure path leaves `held` False and
    lets the Store open anyway: an unwritable directory (evidence on
    read-only media), a filesystem with no flock, a marker someone else
    already holds. Refusing to open is server.py's decision, made from
    probe_case_lock *before* a Store is ever constructed — a Store that has
    got this far is one the analyst has already been asked about."""

    def __init__(self, case_path: str):
        self.path = case_lock_path(case_path)
        self.case_path = os.path.abspath(case_path)
        self.fd: int | None = None
        self.held = False
        self._started_at = time.strftime("%Y-%m-%dT%H:%M:%S")
        self._stop = threading.Event()
        self._thread: threading.Thread | None = None

    def acquire(self) -> bool:
        try:
            fd = os.open(self.path, os.O_RDWR | os.O_CREAT, 0o644)
        except OSError:
            return False
        if fcntl is not None:
            try:
                fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
            except OSError:
                # Held by someone else, or a filesystem without flock. Either
                # way this process doesn't own the file: don't write to it and
                # don't delete it on the way out.
                os.close(fd)
                return False
        self.fd = fd
        self.held = True
        try:
            self._write()
        except OSError:
            pass  # the flock still stands; only the heartbeat is missing
        self._thread = threading.Thread(target=self._beat, daemon=True, name="winnow-case-lock")
        self._thread.start()
        return True

    def _write(self) -> None:
        payload = json.dumps({
            "host": socket.gethostname(),
            "user": _current_user(),
            "pid": os.getpid(),
            "case_path": self.case_path,
            "started_at": self._started_at,
            "heartbeat_at": time.time(),
        }).encode("utf-8")
        # In place on the locked fd. A rename would put the flock on an
        # unlinked inode and silently un-hold the lock.
        os.lseek(self.fd, 0, os.SEEK_SET)
        os.ftruncate(self.fd, 0)
        os.write(self.fd, payload)
        with contextlib.suppress(OSError):
            os.fsync(self.fd)  # the point is that another *machine* can read it

    def _beat(self) -> None:
        while not self._stop.wait(CASE_LOCK_HEARTBEAT_SEC):
            try:
                self._write()
            except OSError:
                return  # marker gone or disk full — stop; the lock is advisory

    def release(self) -> None:
        self._stop.set()
        if self._thread is not None and self._thread.is_alive():
            self._thread.join(5)
        self._thread = None
        if self.fd is None:
            return
        fd, self.fd, self.held = self.fd, None, False
        with contextlib.suppress(OSError):
            os.close(fd)  # drops the flock
        # Only unlink a marker that still names this process. On a
        # filesystem without flock a second Winnow the analyst opened anyway
        # will have overwritten the record with its own pid — deleting the
        # file then would leave that live process looking free to a third.
        rec = _read_lock_record(self.path)
        if rec.get("pid") == os.getpid() and rec.get("host") == socket.gethostname():
            with contextlib.suppress(OSError):
                os.remove(self.path)


def _current_user() -> str:
    try:
        return getpass.getuser()
    except Exception:  # pragma: no cover - no pwd entry / no env vars
        return "unknown"


class Store:
    def __init__(self, path: str, default_tags: list[tuple] | None = None):
        self.path = path
        self._default_tags = default_tags or DEFAULT_TAGS
        self.lock = threading.RLock()
        self.db = sqlite3.connect(path, check_same_thread=False)
        self.db.row_factory = sqlite3.Row
        self.db.create_function("REGEXP", 2, _regexp, deterministic=True)
        self.db.create_function("DAY_BUCKET", 1, _day_bucket, deterministic=True)
        self.db.create_function("TS_NORMALIZE", 1, _ts_normalize, deterministic=True)
        self._tune(self.db)
        # Materialised views live in their own on-disk database, named and
        # WAL-journalled (not the anonymous `ATTACH DATABASE ''` this used
        # to be) so that the read-only connections in _reader() can attach
        # it too — the anonymous temp attach only ever existed on the one
        # connection that ran it, which is what forced every read through
        # self.lock (old invariant #4). Still deleted on close(), same
        # lifecycle guarantee the anonymous attach gave for free.
        #
        # A *named* attach can't ride PRAGMA temp_store=MEMORY the way the
        # old anonymous one did — that pragma only covers SQLite's own TEMP
        # schema, not a file attached by path — so on plain disk this was
        # measured 20-100% slower on every views/paging/timeline benchmark
        # (bench --vs-ref). Two mitigations, both needed: /dev/shm (tmpfs,
        # Linux-only) keeps the file RAM-backed while still being a normal
        # path a second connection can ATTACH, and v.synchronous=OFF drops
        # the remaining fsyncs — safe here and only here, because views are
        # re-derivable scratch state that dies with the process anyway (a
        # crash costs a rebuild the frontend already does on any 409), never
        # evidence like the case file, where _ingest_synchronous_off
        # documents why OFF is off-limits. On macOS/Windows there's no
        # tmpfs path, so the tempdir file leans on synchronous=OFF plus the
        # OS page cache alone.
        fd, self._views_path = tempfile.mkstemp(
            suffix=VIEWS_SUFFIX, prefix=VIEWS_PREFIX, dir=_preferred_views_dir()
        )
        # The mkstemp fd is kept open and flocked for this Store's whole
        # life rather than closed straight away: it's the liveness signal
        # sweep_orphan_views() probes, so the startup janitor in a *second*
        # Winnow process can tell this file apart from one a killed process
        # abandoned. If the lock can't be taken (a filesystem without flock;
        # Windows, where fcntl doesn't exist at all) the fd is just closed —
        # the sweep's own probe fails the same way and errs toward keeping
        # the file, so the degraded mode is "not swept", never "swept while
        # in use". See _views_file_is_orphaned.
        self._views_fd: int | None = None
        if fcntl is not None:
            try:
                fcntl.flock(fd, fcntl.LOCK_EX | fcntl.LOCK_NB)
                self._views_fd = fd
            except OSError:
                os.close(fd)
        else:  # pragma: no cover - Windows
            os.close(fd)
        self.db.execute(f"ATTACH DATABASE '{self._views_path}' AS v")
        # WAL on the views db is what lets _reader() connections read a
        # view concurrently with the writer materialising the next one.
        self.db.execute("PRAGMA v.journal_mode=WAL")
        self.db.execute("PRAGMA v.synchronous=OFF")
        # "This case is open" marker for a *second* Winnow to probe — see
        # _CaseLock. Taken after the database is open so a failed open can't
        # strand one, and never fatal: `held` False just means nothing will
        # warn the next process.
        self._case_lock = _CaseLock(self.path)
        self._case_lock.acquire()
        # Public because server.py has to distinguish "STORE points at this
        # case" from "STORE points at this case and can still answer" — the
        # module global outlives close() on both the case-switch path and
        # the legacy-preset migration path.
        self.closed = False
        # Read-only connection pool for the pure-read paths (paging,
        # grouping, exports, search counts — everything that goes through
        # _reader()). Guarded by its own small lock, never self.lock: the
        # entire point is that checking a reader out must not contend with
        # the writer. Connections are created lazily on first miss and
        # returned up to the cap; a burst beyond the cap opens transient
        # connections that close on return instead of pooling.
        self._reader_pool: list[sqlite3.Connection] = []
        self._reader_lock = threading.Lock()
        self._readers_closed = False
        with self.db:
            # Detect *before* CREATE TABLE IF NOT EXISTS runs, so a case file
            # that predates open_tabs gets every existing source/merge
            # backfilled as open (matching today's "everything's a tab"
            # behavior) exactly once. A case that already has the table —
            # even with zero rows, e.g. the analyst closed every tab on
            # purpose — is left alone; an empty table isn't ambiguous with
            # "never migrated" once we've checked this early.
            open_tabs_existed = self.db.execute(
                "SELECT 1 FROM sqlite_master WHERE type='table' AND name='open_tabs'"
            ).fetchone() is not None
            self.db.executescript(META_SCHEMA)
            # CREATE TABLE IF NOT EXISTS can't add a column to a case file
            # from before nicknames existed — patch those in place.
            if not any(r[1] == "nickname" for r in self.db.execute("PRAGMA table_info(sources)")):
                self.db.execute("ALTER TABLE sources ADD COLUMN nickname TEXT")
            if not open_tabs_existed:
                ids = [r[0] for r in self.db.execute("SELECT id FROM sources")]
                ids += [-r[0] for r in self.db.execute("SELECT id FROM merges")]
                if ids:
                    self.db.executemany(
                        "INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", [(i,) for i in ids]
                    )
        self._seed_tags()
        self._view_seq = 0
        self._views: dict[str, dict] = {}
        self._undo_seq = 0
        # Tag-change undo history. Entries are newest-last; each owns a
        # v.undo_<n> delta table listing the rows the op *actually*
        # changed. Scratch, not evidence — it lives in the views database
        # (invariant #3) and dies with the process, same as a view.
        self._undo: list[dict] = []
        self._maxlen_cache: dict[int, tuple[int, dict[str, int]]] = {}
        self._fts_threads: dict[int, threading.Thread] = {}
        self._index_threads: dict[tuple[int, str], threading.Thread] = {}
        # Guards the two thread registries above — its own lock, not
        # self.lock, for the same reason as _search_job_lock: the ensure-*
        # helpers are called from read paths (group_summary, column_values,
        # the search sweep), and bookkeeping that waited on the connection
        # lock would make a "pure read" stall behind a long build_view
        # after all. Never held while acquiring any other lock.
        self._threads_lock = threading.Lock()
        self._fts_janitor: threading.Thread | None = None
        # Search-all job state. Its own lock, not self.lock: the worker
        # mutates the job record between per-source counts, and making that
        # wait on the connection lock would reintroduce exactly the
        # contention the per-source lock scoping exists to avoid.
        self._search_job_lock = threading.Lock()
        self._search_job: dict | None = None
        self._search_job_thread: threading.Thread | None = None
        self._search_job_seq = 0

        # Ingest jobs — same fire-and-forget-with-a-registry shape as the
        # search-all job above, but plural: a directory import legitimately
        # starts many at once. The semaphore caps how many actually ingest
        # concurrently; the rest sit 'queued'. Each worker's writes still go
        # through self.lock per BATCH like any other ingest, so the cap is
        # about not stacking N ingest threads' parse work on the CPU, not
        # about connection safety.
        self._ingest_jobs: dict[int, dict] = {}
        self._ingest_jobs_lock = threading.Lock()
        self._ingest_job_seq = 0
        self._ingest_sem = threading.Semaphore(self.MAX_CONCURRENT_INGESTS)

        # Cancellable-op registry (see _interruptible/cancel_op): token →
        # the connection running that op, so a cancel can interrupt() it.
        # Its own lock for the same reason _search_job_lock exists.
        self._op_lock = threading.Lock()
        self._op_conns: dict[str, sqlite3.Connection] = {}
        self._op_cancelled: set[str] = set()
        self._downgrade_legacy_fts()

    def _downgrade_legacy_fts(self) -> None:
        """Sources whose fts_<id> predates the current index shape still
        have has_fts=1 pointing at a table the query code can't use — the
        original word-tokenized table (can't do substring search at all),
        or the first-generation trigram table (multi-column, detail=full:
        the new `doc LIKE ?` query form would be a SQL error against it,
        and it's ~6x the size the index needs to be — the reason the shape
        changed). Either way: treat it as not-ready (the lazy background
        rebuild in _ensure_fts_building kicks in on next use) rather than
        erroring or serving wrong results under a stale flag. No schema
        migration needed: the shape is read straight back out of
        sqlite_master's own CREATE VIRTUAL TABLE text ('detail=none' only
        appears in current-shape DDL), so this is a one-time, idempotent
        correction.

        The stale tables themselves get dropped on a background thread,
        one per transaction — for a fat-trigram case file that's most of
        the file's bulk (measured 892MB per 285MB source), and freeing it
        shouldn't block Store construction. Freed pages go to SQLite's
        freelist for reuse by this case file; the file itself only shrinks
        under a VACUUM, which nothing here runs automatically."""
        stale: list[int] = []
        with self.lock, self.db:
            rows = self.db.execute("SELECT id FROM sources WHERE has_fts=1").fetchall()
            for (source_id,) in rows:
                row = self.db.execute(
                    "SELECT sql FROM sqlite_master WHERE type='table' AND name=?",
                    (f"fts_{source_id}",),
                ).fetchone()
                if not row or "detail=none" not in row[0]:
                    self.db.execute("UPDATE sources SET has_fts=0 WHERE id=?", (source_id,))
                    if row:
                        stale.append(source_id)
        if stale:
            t = threading.Thread(target=self._drop_stale_fts_worker, args=(stale,), daemon=True)
            self._fts_janitor = t
            t.start()

    def _drop_stale_fts_worker(self, source_ids: list[int]) -> None:
        for source_id in source_ids:
            name = f"fts_{source_id}"
            try:
                with self.lock, self.db:
                    # Re-check under the lock: a search may have triggered a
                    # rebuild since startup, and this name would now be the
                    # fresh index — only drop it if it's still the old shape.
                    row = self.db.execute(
                        "SELECT sql FROM sqlite_master WHERE type='table' AND name=?", (name,)
                    ).fetchone()
                    if row and "detail=none" not in row[0]:
                        self.db.execute(f"DROP TABLE IF EXISTS {q(name)}")
            except Exception:
                pass  # best-effort space reclamation — a survivor is dropped by its own rebuild
            time.sleep(0.02)

    def wait_for_fts_maintenance(self, timeout: float | None = None) -> None:
        """Blocks until the startup stale-index janitor finishes (or
        `timeout` elapses). Used by tests."""
        t = self._fts_janitor
        if t:
            t.join(timeout)

    def _ensure_fts_building(self, source_id: int) -> None:
        """Fire-and-forget: kicks off a background trigram-index build for
        this source if it doesn't already have one ready and isn't already
        building one. Callers (Contains/Advanced-mode search) call this and
        then immediately fall back to the LIKE-based path for the query
        they're already compiling — the next search against this source,
        once the build finishes, gets the fast indexed path instead.

        No-ops entirely on a pre-3.45 SQLite: the index is only ever
        queried through the trigram LIKE pushdown that version introduced,
        so on an older runtime it would cost disk and build time while
        every search still ran at fallback speed."""
        if sqlite3.sqlite_version_info < TRIGRAM_LIKE_MIN_SQLITE:
            return
        with self._threads_lock:
            existing = self._fts_threads.get(source_id)
            if existing and existing.is_alive():
                return
        # A _reader() for the has_fts check, not _source_lite/self.lock:
        # this helper is called from read paths mid-scan, and every call
        # site runs outside any writer transaction (ingest fires it after
        # its final commit), so committed state is exactly what to check.
        with self._reader() as ro:
            if self._source_lite_on(ro, source_id).get("has_fts"):
                return
        with self._threads_lock:
            existing = self._fts_threads.get(source_id)
            if existing and existing.is_alive():
                return
            t = threading.Thread(target=self._build_fts_worker, args=(source_id,), daemon=True)
            self._fts_threads[source_id] = t
        t.start()

    def _build_fts_worker(self, source_id: int) -> None:
        try:
            self.build_fts(source_id)
        except Exception:
            pass  # best-effort background upgrade — the next search attempt retries

    def _is_fts_building(self, source_id: int) -> bool:
        with self._threads_lock:
            t = self._fts_threads.get(source_id)
        return bool(t and t.is_alive())

    def wait_for_fts(self, source_id: int, timeout: float | None = None) -> bool:
        """Blocks until a background index build for this source finishes
        (or `timeout` elapses). Used by tests, and available for any caller
        that genuinely needs up-to-date has_fts before proceeding. Returns
        whether the index is ready by the time it returns."""
        with self._threads_lock:
            t = self._fts_threads.get(source_id)
        if t:
            t.join(timeout)
        return bool(self.get_source(source_id).get("has_fts"))

    @staticmethod
    def _column_index_name(table_name: str, column: str, purpose: str = "filter") -> str:
        # Column names are arbitrary user data (CSV headers) — hash rather
        # than interpolate one into the index name, so this never has to
        # worry about identifier-safety beyond what q() already guarantees
        # for the CREATE INDEX statement itself. purpose="filter" hashes
        # just the column, unchanged from before this had a purpose
        # argument, so index names already on disk in an existing case file
        # stay recognized; a distinct purpose (currently only "sort") salts
        # the hash so it never collides with the filter index's name — see
        # _ensure_sort_index_building for why they have to be two different
        # physical indexes.
        key = column if purpose == "filter" else f"{column}\x00{purpose}"
        h = hashlib.md5(key.encode()).hexdigest()[:12]
        return f"idx_{table_name}_{h}"

    def _column_index_exists(self, table_name: str, column: str, purpose: str = "filter") -> bool:
        name = self._column_index_name(table_name, column, purpose)
        # A _reader(), same reasoning as _ensure_fts_building's has_fts
        # check: called from read paths, always outside a writer
        # transaction, and an index only "exists" once its CREATE commits.
        with self._reader() as ro:
            row = ro.execute(
                "SELECT 1 FROM sqlite_master WHERE type='index' AND name=?", (name,)
            ).fetchone()
        return row is not None

    def _ensure_column_index_building(self, source_id: int, column: str, table_name: str | None = None) -> None:
        """Fire-and-forget: kicks off a background plain B-tree index build
        for (source_id, column) if one doesn't already exist and isn't
        already building — same pattern as _ensure_fts_building. Only
        called for SARGABLE_OPS ('equals'/'in'): those are the filter shapes
        a plain single-column index actually accelerates. 'contains'/
        'starts' are LIKE patterns (trigram FTS already covers substring
        search); numeric comparisons go through _numeric_expr, a functional
        expression a plain index on the raw column wouldn't match anyway.
        Callers keep using today's scan for the filter they're already
        compiling — the next application of this same filter, once the
        build finishes, gets the indexed path instead. Verified on a
        42 GB/11-member merged case: an EventId-equals filter went from
        6-8s per application (disk-bound full scan of every member, never
        staying cached — the working set is bigger than available RAM) to
        ~50ms once indexed.

        `table_name` lets a caller that already resolved it skip the
        get_source lookup. That isn't micro-optimisation: get_source runs a
        COUNT(DISTINCT rid) over row_tags, ~12ms on a heavily tagged source,
        and this is called on every filter compile and every group-by — once
        per member, so a merge multiplies it."""
        self._ensure_purpose_index_building(source_id, column, "filter", table_name)

    def _ensure_sort_index_building(self, source_id: int, column: str, table_name: str | None = None) -> None:
        """Same fire-and-forget background build as
        _ensure_column_index_building, but for an ORDER BY column instead of
        a sargable filter — triggered from build_view for every non-numeric
        sort column, once per member for a merge.

        This can't reuse the filter index: _compile_order emits
        `col COLLATE NOCASE` for a non-numeric column, and SQLite will only
        use an index to serve a comparison whose collating sequence matches
        the index's own declared collation — verified with EXPLAIN QUERY
        PLAN both ways: a plain (BINARY) index does not serve a NOCASE
        ORDER BY (the temp B-tree stays), and a NOCASE index does not serve
        a bare `=` filter (which compares BINARY, the column's default
        collation) either. So a column that's both filtered-on and
        sorted-on ends up with two small physical indexes, not one — still
        far cheaper than unconditionally indexing every column both ways.

        Numeric columns are skipped entirely: a numeric sort goes through
        _numeric_expr, a functional expression no plain index (of either
        collation) can match — same exclusion the filter index already
        documents, enforced by build_view filtering sort_cols to
        colnames[col] != 'number' before calling this."""
        self._ensure_purpose_index_building(source_id, column, "sort", table_name)

    def _ensure_purpose_index_building(self, source_id: int, column: str, purpose: str,
                                        table_name: str | None = None) -> None:
        key = (source_id, column, purpose)
        with self._threads_lock:
            existing = self._index_threads.get(key)
            if existing and existing.is_alive():
                return
        table_name = table_name or self._source_lite(source_id)["table_name"]
        if self._column_index_exists(table_name, column, purpose):
            return
        with self._threads_lock:
            existing = self._index_threads.get(key)
            if existing and existing.is_alive():
                return
            t = threading.Thread(target=self._build_column_index_worker,
                                 args=(column, table_name, purpose), daemon=True)
            self._index_threads[key] = t
        t.start()

    def _build_column_index_worker(self, column: str, table_name: str, purpose: str = "filter") -> None:
        try:
            name = self._column_index_name(table_name, column, purpose)
            collate = " COLLATE NOCASE" if purpose == "sort" else ""
            with self.lock, self.db:
                self.db.execute(f"CREATE INDEX IF NOT EXISTS {q(name)} ON {q(table_name)}({q(column)}{collate})")
        except Exception:
            pass  # best-effort background upgrade — the next filter/sort application retries

    def wait_for_column_index(self, source_id: int, column: str, timeout: float | None = None,
                               purpose: str = "filter") -> bool:
        """Blocks until a background index build for (source_id, column,
        purpose) finishes (or `timeout` elapses). Used by tests. Returns
        whether the index exists by the time it returns."""
        with self._threads_lock:
            t = self._index_threads.get((source_id, column, purpose))
        if t:
            t.join(timeout)
        src = self.get_source(source_id)
        return self._column_index_exists(src["table_name"], column, purpose)

    def list_column_indexes(self, source_id: int) -> list[dict]:
        """Which of this source's columns currently have one of the
        auto-created indexes (filter or sort — the Tables modal doesn't
        distinguish, both are "just an index" from the analyst's side), and
        whether one is being built right now.

        These get created behind the analyst's back (a sargable filter, a
        value dropdown, a group-by, or now a sort all trigger one) and never
        expire, so on a case where the same headers get imported and
        filtered/sorted over and over they accumulate silently and take
        real disk. This is what lets the Tables modal show them and offer a
        drop.

        The index *name* carries only an md5 of the column (see
        _column_index_name — a column name is arbitrary CSV-header text, so
        it's deliberately never interpolated into an identifier), which means
        there's no way to read a column back out of a name. Going the other
        direction works fine: hash each of the source's known columns (both
        purposes) and look for that name, which is what this does."""
        if source_id < 0:
            return []  # a merge has no table of its own; its members carry the indexes
        src = self.get_source(source_id)
        table = src["table_name"]
        with self.lock:
            existing = {
                r[0] for r in self.db.execute(
                    "SELECT name FROM sqlite_master WHERE type='index' AND tbl_name=?", (table,)
                )
            }
        with self._threads_lock:
            building = {col for (sid, col, _purpose), t in self._index_threads.items()
                        if sid == source_id and t.is_alive()}
        out = []
        for c in src["columns"]:
            # A derived column's indexes live on the drv_<id> sidecar, where
            # its values are — see _index_table_for.
            ctable = self._index_table_for(src, c["name"], table)
            name = self._column_index_name(ctable, c["name"])
            sort_name = self._column_index_name(ctable, c["name"], "sort")
            if name in existing or sort_name in existing:
                out.append({"column": c["name"], "index_name": name if name in existing else sort_name, "building": False})
            elif c["name"] in building:
                out.append({"column": c["name"], "index_name": name, "building": True})
        return out

    def drop_column_index(self, source_id: int, column: str) -> None:
        """Drops every auto-created index (filter and/or sort — see
        _ensure_sort_index_building for why a column can have both) on this
        column. The pages go to the case file's freelist, not back to the
        OS — `compact()` is what returns them. Nothing breaks without the
        index: every query it accelerates still returns the same rows by
        scanning, and the next filter/sort/group-by on this column just
        builds it again."""
        src = self.get_source(source_id)
        names = {c["name"] for c in src["columns"]}
        if column not in names:
            raise KeyError(column)
        with self.lock, self.db:
            for purpose in ("filter", "sort"):
                name = self._column_index_name(self._index_table_for(src, column), column, purpose)
                self.db.execute(f"DROP INDEX IF EXISTS {q(name)}")

    # Slack VACUUM needs beyond a full second copy of the file, for the
    # journal/WAL it writes alongside. Deliberately generous — running out
    # of disk part-way through is the one failure mode worth refusing to
    # risk, and the analyst can always free space and retry.
    VACUUM_HEADROOM = 256 * 1024 * 1024

    def compact(self) -> dict:
        """VACUUM the case file, returning it to the size its live data
        actually needs.

        Nothing in normal operation shrinks a case file. Dropping a source,
        dropping a stale FTS index (the startup janitor frees whole
        hundreds of MB doing this), dropping a column index — all of it
        goes to SQLite's freelist for reuse by this same file, which is
        the right default (reuse is free, rewriting the file is not). But
        after a big cleanup that can be a large, permanently-parked chunk
        of disk, and VACUUM is the only thing that gives it back. It stays
        an explicit, confirmed action rather than anything automatic: it
        rewrites the entire file, so on a multi-GB case it's minutes of
        held lock, and it needs a full second copy of the file free on
        disk while it runs — checked up front here, because running out
        of space mid-VACUUM is a far worse outcome than declining.

        temp_store is forced back to FILE for the duration. _tune sets it
        to MEMORY, which is right for the sorts and temp b-trees ordinary
        queries build — but VACUUM's scratch copy of the whole database
        obeys the same pragma, and materialising a 42 GB case in RAM is
        not a thing to find out about the hard way.

        The WAL is checkpointed and truncated on both sides of the VACUUM.
        Afterwards so the reclaimed bytes don't just move into a -wal file;
        *before* so the two sizes are comparable at all — in WAL mode
        recently written pages live in the -wal until a checkpoint, so a
        case whose data has never been checkpointed reports a main-file
        size of one page and would show a nonsense negative reclaim."""
        with self.lock:
            self.db.commit()
            self.db.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            before = os.path.getsize(self.path)
            directory = os.path.dirname(os.path.abspath(self.path)) or "."
            free = shutil.disk_usage(directory).free
            needed = before + self.VACUUM_HEADROOM
            if free < needed:
                raise ValueError(
                    f"Not enough free disk space to compact: VACUUM rewrites the whole file, "
                    f"so it needs about {needed // (1024 * 1024)} MB free and "
                    f"{free // (1024 * 1024)} MB is available"
                )
            self.db.execute("PRAGMA temp_store=FILE")
            try:
                self.db.execute("VACUUM")
                self.db.execute("PRAGMA wal_checkpoint(TRUNCATE)")
            finally:
                self.db.execute("PRAGMA temp_store=MEMORY")
            after = os.path.getsize(self.path)
        return {"before_bytes": before, "after_bytes": after, "reclaimed_bytes": max(0, before - after)}

    @staticmethod
    def _tune(conn: sqlite3.Connection) -> None:
        # page_size only takes effect on a database with no tables yet, and
        # only if set before the first write of any kind — including the
        # journal_mode=WAL switch below, which itself allocates page 1 and
        # silently locks in whatever page size was already in force
        # (verified: page_size after journal_mode=WAL is a no-op). On an
        # existing case file this pragma is a no-op regardless of order —
        # changing an established file's page size needs a VACUUM.
        conn.execute("PRAGMA page_size=65536")
        conn.execute("PRAGMA journal_mode=WAL")
        conn.execute("PRAGMA synchronous=NORMAL")
        conn.execute("PRAGMA temp_store=MEMORY")
        conn.execute("PRAGMA cache_size=-262144")  # ~256 MB page cache
        # 1 GB. mmap is file-backed and shared with the OS page cache, so
        # this costs address space (fine on 64-bit), not RAM — and it lets
        # the big sequential scans (view materialisation, LIKE fallback,
        # FTS builds) skip a read() syscall + memcpy per page on the
        # multi-GB cases this tool targets. SQLite silently clamps it to
        # its compile-time max (typically ~2 GB) — asking for more than an
        # old build allows degrades gracefully rather than erroring.
        conn.execute("PRAGMA mmap_size=1073741824")
        # Let the sorter use worker threads: a filtered/sorted view build
        # over millions of rows is one big external sort, and this is the
        # one knob that parallelises it (default is 0 = single-threaded).
        conn.execute("PRAGMA threads=4")

    def _seed_tags(self) -> None:
        with self.lock, self.db:
            n = self.db.execute("SELECT count(*) FROM tag_defs").fetchone()[0]
            if not n:
                self.db.executemany(
                    "INSERT INTO tag_defs(name, color, hotkey) VALUES (?,?,?)",
                    self._default_tags,
                )

    def close(self) -> None:
        with self._reader_lock:
            self._readers_closed = True
            idle, self._reader_pool = self._reader_pool, []
        for conn in idle:
            with contextlib.suppress(sqlite3.Error):
                conn.close()
        # Stop any running ingest jobs before the connection goes away — a
        # worker mid-batch would otherwise die on a closed database. Cancel
        # is cooperative (checked per BATCH), so the join is bounded by one
        # batch commit plus the partial-source drop.
        with self._ingest_jobs_lock:
            jobs = list(self._ingest_jobs.values())
            for j in jobs:
                j["cancelled"] = True
        for j in jobs:
            t = j.get("thread")
            if t and t is not threading.current_thread() and t.is_alive():
                t.join(15)
        with self.lock:
            self.db.close()
        # Release the flock (see __init__) before unlinking, so that if the
        # remove below is the one that fails, what's left behind is an
        # unlocked file the next process's sweep_orphan_views can collect.
        if self._views_fd is not None:
            with contextlib.suppress(OSError):
                os.close(self._views_fd)
            self._views_fd = None
        # A reader still checked out by an in-flight request holds the
        # unlinked file open — fine on POSIX (the pages stay readable until
        # its close), and on Windows the remove just fails and is skipped,
        # leaving a stray temp file rather than an error. Either way it's no
        # longer permanent: the sweep at the next server start takes it.
        for suffix in VIEWS_DB_SUFFIXES:
            try:
                os.remove(self._views_path + suffix)
            except OSError:
                pass
        self._case_lock.release()
        self.closed = True

    READER_POOL_CAP = 6  # matches app.js's PAGE_FETCH_CONCURRENCY — the burstiest client

    def _open_reader(self) -> sqlite3.Connection:
        """One read-only connection for _reader()'s pool. Never takes
        self.lock: WAL mode (_tune, and the `PRAGMA v.journal_mode=WAL` set
        on the writer at attach time) is what makes concurrent, unlocked
        reads against both the case file and the views database safe — a
        reader sees the latest *committed* data and is never blocked by, or
        blocks, a writer holding self.lock for a long
        build_view/ingest/search sweep. Same registered functions as the
        writer (a group_virtual WHERE can contain DAY_BUCKET(...); a raw
        filter fragment echoed into an export can contain REGEXP), so a
        query compiled against one connection never fails on the other."""
        ro = sqlite3.connect(f"file:{self.path}?mode=ro", uri=True, check_same_thread=False)
        ro.row_factory = sqlite3.Row
        ro.create_function("REGEXP", 2, _regexp, deterministic=True)
        ro.create_function("DAY_BUCKET", 1, _day_bucket, deterministic=True)
        ro.create_function("TS_NORMALIZE", 1, _ts_normalize, deterministic=True)
        # Same mmap reasoning as _tune (address space, not RAM — and shared
        # with the writer's mapping via the OS page cache). busy_timeout
        # covers the rare WAL edge (recovery, checkpoint restart) where
        # even a reader can see SQLITE_BUSY briefly.
        ro.execute("PRAGMA mmap_size=1073741824")
        ro.execute("PRAGMA busy_timeout=5000")
        ro.execute(f"ATTACH DATABASE 'file:{self._views_path}?mode=ro' AS v")
        return ro

    @contextlib.contextmanager
    def _reader(self) -> Iterator[sqlite3.Connection]:
        """Checkout/return for the read-only pool. Every pure-read path
        (fetch_rows, tag_positions, group_summary, exports, search counts,
        ...) runs its queries on one of these instead of self.db, which is
        what makes a slow build_view or ingest invisible to paging.

        Two rules for using it:

        - Never inside an open writer transaction. A reader sees committed
          data only; code that has just INSERTed under `with self.db:` and
          needs to read it back must stay on self.db (e.g. ingest's
          get_source call). Every current _reader() caller is a top-level
          read-only Store method — keep it that way.
        - The connection is checked out by one thread at a time (that's
          what the pool provides; a shared sqlite3 connection would
          serialize its users at the C level and quietly reintroduce the
          contention this removes). Holding one across a generator's yields
          is fine — exports do — because the pool just opens another for
          whoever else asks.

        On any exception the connection is closed, not returned: the only
        realistic errors here are a view/source table dropped mid-read or
        the store closing, and a connection whose last statement died is
        cheaper to replace than to prove clean. The pool refills lazily."""
        with self._reader_lock:
            conn = self._reader_pool.pop() if self._reader_pool else None
        if conn is None:
            conn = self._open_reader()
        try:
            yield conn
        except BaseException:
            with contextlib.suppress(sqlite3.Error):
                conn.close()
            raise
        with self._reader_lock:
            if not self._readers_closed and len(self._reader_pool) < self.READER_POOL_CAP:
                self._reader_pool.append(conn)
                conn = None
        if conn is not None:
            conn.close()

    @staticmethod
    @contextlib.contextmanager
    def _dropped_view_is_expired() -> Iterator[None]:
        """Maps "no such table" during an unlocked view read to the same
        KeyError contract a missing handle raises, so server.py's existing
        KeyError → 409 mapping turns it into the "view expired" the
        frontend already rebuilds on. This race is new with _reader():
        under the old single-connection design the lock serialized a page
        fetch against the eviction that drops its view table; now a reader
        holding a handle can lose the table to _evict_root_views
        mid-request. (A *source* table dropped mid-read lands here too —
        same answer: the world changed under the request, rebuild.)"""
        try:
            yield
        except sqlite3.OperationalError as e:
            if "no such table" in str(e).lower():
                raise KeyError("View expired — rebuild it") from None
            raise

    # ------------------------------------------------------------------ ingest

    # ------------------------------------------------------- cancellable ops

    def cancel_op(self, token: str) -> bool:
        """Cancels the in-flight operation registered under `token` (a
        client-generated one-shot id riding on the request that started
        it — see _interruptible). Returns whether there was anything to
        interrupt. Marking the token cancelled *first* closes the start
        race: a cancel that lands before the operation registers still
        takes effect the moment it tries to."""
        with self._op_lock:
            self._op_cancelled.add(token)
            while len(self._op_cancelled) > 512:  # one-shot ids; keep it bounded
                self._op_cancelled.pop()
            conn = self._op_conns.get(token)
            if conn is not None:
                conn.interrupt()
                return True
        return False

    @contextlib.contextmanager
    def _interruptible(self, token: str | None, conn: sqlite3.Connection):
        """Registers `conn` as interruptible under `token` for the block,
        turning cancel_op(token) into OpCancelled out of this block.

        The discipline that makes interrupt() safe to expose at all: a
        writer caller enters this block while *holding* self.lock and
        unregisters before releasing it, so the only statements that can be
        running on the writer connection while its token is registered are
        the block's own. interrupt() aborts running statements only; a
        cancel arriving in the gap between two statements is a no-op (the
        flag clears when the statement count hits zero), which surfaces as
        "the cancel didn't land — click again", never as a mis-aimed kill
        of someone else's write. Reader-pool callers pass their checked-out
        reader (and stack this innermost): an interrupted reader raises out
        of _reader(), which already closes — not repools — a connection
        that died mid-statement."""
        if not token:
            yield
            return
        with self._op_lock:
            if token in self._op_cancelled:
                raise OpCancelled("Cancelled")
            self._op_conns[token] = conn
        try:
            yield
        except sqlite3.OperationalError as e:
            with self._op_lock:
                was_cancelled = token in self._op_cancelled
            if was_cancelled and "interrupt" in str(e).lower():
                raise OpCancelled("Cancelled") from e
            raise
        finally:
            with self._op_lock:
                self._op_conns.pop(token, None)

    @contextlib.contextmanager
    def _ingest_synchronous_off(self):
        """synchronous=OFF for the duration of one import's batch-commit
        loop, restored to NORMAL (the steady-state setting from _tune)
        whether the import finished, raised, or was interrupted. WAL stays
        on — only the fsync-on-commit behavior is relaxed.

        This is the shared connection (invariant #4), so the relaxed
        setting is in force for any other write that happens to land in the
        same window, not just this import's own — the risk that bounds is
        the same one ingest already accepts: a crash mid-import costs the
        in-flight import (and, in this narrow window, an unlucky concurrent
        write), and every batch before the crash is already durably
        committed to the WAL up to the last synchronous=NORMAL checkpoint.
        Never combine with journal_mode=OFF (that would risk case-file
        corruption, not just losing the in-flight write) — this case file
        is the durable artifact and holds tags/notes/sessions that aren't
        re-derivable from the source, unlike a throwaway temp database
        where that risk would be easy to accept."""
        with self.lock, self.db:
            self.db.execute("PRAGMA synchronous=OFF")
        try:
            yield
        finally:
            with self.lock, self.db:
                self.db.execute("PRAGMA synchronous=NORMAL")

    def ingest_csv(
        self,
        path: str,
        name: str | None = None,
        delimiter: str | None = None,
        build_fts: bool = True,
        has_header: bool = True,
        column_types: list[str] | None = None,
        progress=None,
        cancel=None,
    ) -> dict:
        """Stream a delimited file into its own table. Returns the source record.

        Commits in BATCH-sized chunks instead of one giant transaction, and
        only holds self.lock for the duration of each chunk's insert — not
        for the whole file. self.lock guards every Store method (it's the
        one thing serializing all access to the shared connection), so on a
        huge import (minutes, at the scale this tool targets) the old
        one-transaction-for-the-whole-file approach froze every other
        request — fetching rows, tagging, opening another already-loaded
        source — for the entire duration. It also meant one bad row near the
        end (a malformed line, a full disk) rolled back everything already
        inserted, discarding all of it.

        sources.row_count is updated after every chunk, so if something does
        go wrong partway through, whatever was already committed stays
        (and stays visible/queryable with an accurate count) instead of the
        whole import vanishing — the caller still sees the error and knows
        the source is incomplete, but doesn't lose everything that
        succeeded before it.
        """
        name = name or os.path.basename(path)
        size = os.path.getsize(path)
        file_hash = self._quick_hash(path)

        # 1 MB read buffer: the default 8 KB means a 20 GB file is ~2.5M
        # read() syscalls before the csv layer even sees a byte.
        fh = open(path, "r", encoding="utf-8-sig", errors="replace", newline="", buffering=1 << 20)
        head = fh.read(64 * 1024)
        fh.seek(0)
        if delimiter is None:
            delimiter = self._sniff(head)
        reader = csv.reader(fh, delimiter=delimiter)

        try:
            first_row = next(reader)
        except StopIteration:
            fh.close()
            raise ValueError("File is empty")
        if has_header:
            cols = sanitize_columns(first_row)
            ncols = len(cols)
            leftover_row = None
        else:
            ncols = len(first_row)
            cols = sanitize_columns([None] * ncols)
            leftover_row = first_row  # not a header — it's the first data row

        with self.lock, self.db:
            cur = self.db.execute(
                "INSERT INTO sources(name, path, table_name, columns, file_hash, imported_at)"
                " VALUES (?,?,?,?,?,?)",
                (name, os.path.abspath(path), "", "[]", file_hash,
                 time.strftime("%Y-%m-%dT%H:%M:%S")),
            )
            source_id = cur.lastrowid
            self.db.execute("INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", (source_id,))
            table = f"src_{source_id}"
            coldefs = ", ".join(f"{q(c)} TEXT" for c in cols)
            self.db.execute(f"CREATE TABLE {q(table)} (rid INTEGER PRIMARY KEY, {coldefs})")
            self.db.execute("UPDATE sources SET table_name=? WHERE id=?", (table, source_id))

        placeholders = ",".join("?" * ncols)
        insert = f"INSERT INTO {q(table)} ({','.join(q(c) for c in cols)}) VALUES ({placeholders})"

        sample: list[list[str]] = []
        batch: list[tuple] = []
        total = 0
        ragged = 0
        t0 = time.time()
        error: Exception | None = None

        # This per-row append loop is deliberate: a chunked
        # list(islice(reader, BATCH)) rewrite was measured *slower* (~7% on
        # a 200k-row file) — islice's per-item indirection costs more than
        # the tuple() copy it saves, and executemany binds tuples slightly
        # faster than lists. Don't "optimise" it back.
        rows_iter = ([leftover_row] if leftover_row is not None else [])
        with self._ingest_synchronous_off():
            try:
                for row in itertools.chain(rows_iter, reader):
                    if len(row) != ncols:
                        # Ragged row: pad or trim rather than dropping evidence.
                        ragged += 1
                        row = (row + [""] * ncols)[:ncols]
                    batch.append(tuple(row))
                    if len(sample) < SAMPLE_ROWS:
                        sample.append(row)
                    if len(batch) >= BATCH:
                        if cancel is not None and cancel():
                            raise IngestCancelled(f"Import of {name} cancelled")
                        total = self._commit_ingest_batch(insert, batch, source_id, total)
                        batch.clear()
                        if progress:
                            # NOT fh.tell(): a text file being iterated (the
                            # csv reader drives the file iterator) raises
                            # "telling position disabled by next() call" on
                            # tell(). The underlying BufferedReader's byte
                            # position is legal to read and ahead of the
                            # decoded position by at most the 1 MB buffer.
                            progress(total, fh.buffer.tell(), size)
                if batch:
                    if cancel is not None and cancel():
                        raise IngestCancelled(f"Import of {name} cancelled")
                    total = self._commit_ingest_batch(insert, batch, source_id, total)
                    batch.clear()
                if progress:
                    progress(total, size, size)  # exact final tick — the loop's ticks stop at the last full batch
            except IngestCancelled:
                # An explicit cancel discards the partial import — unlike a
                # mid-file *error*, which keeps what already committed (see
                # the docstring). The analyst asked for this source not to
                # exist, and a keep-what-committed half-table would look
                # exactly like a complete import in every table list.
                self.drop_source(source_id)
                raise
            except Exception as e:
                error = e
            finally:
                fh.close()

        # Best-effort even on failure: whatever's in `sample` (up to
        # SAMPLE_ROWS, collected during parsing regardless of whether the
        # batch it was in ever committed) is enough to type the columns, so a
        # partial import is still typed/browsable rather than left with the
        # placeholder columns=[] forever.
        types = [infer_type([r[i] for r in sample]) for i in range(ncols)] if sample else ["text"] * ncols
        if column_types:
            types = [(column_types[i] if i < len(column_types) and column_types[i] in ("text", "number", "datetime") else t)
                     for i, t in enumerate(types)]
        colmeta = [{"name": c, "type": t} for c, t in zip(cols, types)]
        with self.lock, self.db:
            self.db.execute("UPDATE sources SET columns=? WHERE id=?", (json.dumps(colmeta), source_id))

        if error is not None:
            raise error

        if build_fts and total:
            # Backgrounded, not built inline: the trigram index takes real
            # time to build (see build_fts) and ingest_csv should return as
            # soon as rows are typed/committed so the analyst can start
            # browsing immediately — has_fts is 0 in the returned record
            # until the background build finishes.
            self._ensure_fts_building(source_id)

        elapsed = time.time() - t0
        rec = self.get_source(source_id)
        rec["elapsed_sec"] = round(elapsed, 2)
        rec["rows_per_sec"] = int(total / elapsed) if elapsed > 0 else 0
        rec["ragged_rows"] = ragged
        return rec

    def _commit_ingest_batch(self, insert_sql: str, batch: list[tuple], source_id: int, total: int) -> int:
        """Insert+commit one ingest batch in its own short transaction,
        updating sources.row_count as we go. Only holds self.lock for this
        one batch — see ingest_csv."""
        with self.lock, self.db:
            self.db.executemany(insert_sql, batch)
            total += len(batch)
            self.db.execute("UPDATE sources SET row_count=? WHERE id=?", (total, source_id))
        return total

    def preview_csv_text(
        self, text: str, delimiter: str | None = None, has_header: bool = True, max_rows: int = 50
    ) -> dict:
        """Read-only sample of a delimited file's first rows, for the import
        preview UI. Never touches the database — no source row, no table."""
        if delimiter is None:
            delimiter = self._sniff(text[:8192])
        reader = csv.reader(io.StringIO(text), delimiter=delimiter)
        try:
            first_row = next(reader)
        except StopIteration:
            raise ValueError("File is empty")
        if has_header:
            cols = sanitize_columns(first_row)
            ncols = len(cols)
            leftover_row = None
        else:
            ncols = len(first_row)
            cols = sanitize_columns([None] * ncols)
            leftover_row = first_row

        sample: list[list[str]] = []
        for row in itertools.chain([leftover_row] if leftover_row is not None else [], reader):
            if len(sample) >= max_rows:
                break
            if len(row) != ncols:
                row = (row + [""] * ncols)[:ncols]
            sample.append(row)

        types = [infer_type([r[i] for r in sample]) for i in range(ncols)] if sample else ["text"] * ncols
        return {"delimiter": delimiter, "columns": cols, "sample_rows": sample, "inferred_types": types}

    def preview_sqlite_tables(self, path: str) -> dict:
        """Read-only look at every table in an uploaded/pointed-to SQLite
        file — Chromium's History/Cookies/Web Data/... or any other
        .db/.sqlite — for the import UI's table picker: names, row counts,
        columns, and which columns look like a WebKit/Chrome timestamp
        (see _webkit_to_iso) so that option can default to checked without
        the analyst needing to already know which columns are epoch-encoded.
        Never writes to the case db — a completely separate connection to
        a completely separate file."""
        ro = sqlite3.connect(f"file:{path}?mode=ro", uri=True, check_same_thread=False)
        ro.row_factory = sqlite3.Row
        try:
            table_names = [
                r["name"] for r in ro.execute(
                    "SELECT name FROM sqlite_master WHERE type='table' AND name NOT LIKE 'sqlite_%' ORDER BY name"
                )
            ]
            tables = []
            for t in table_names:
                cols = ro.execute(f"PRAGMA table_info({q(t)})").fetchall()
                row_count = ro.execute(f"SELECT COUNT(*) FROM {q(t)}").fetchone()[0]
                likely_ts = []
                for c in cols:
                    if not re.search(r"time|utc|date", c["name"], re.IGNORECASE):
                        continue
                    sample = [
                        r[0] for r in ro.execute(
                            f"SELECT {q(c['name'])} FROM {q(t)} WHERE {q(c['name'])} IS NOT NULL LIMIT 20"
                        )
                    ]
                    if sample and sum(1 for v in sample if _webkit_to_iso(v)) >= max(1, int(len(sample) * 0.8)):
                        likely_ts.append(c["name"])
                tables.append({
                    "name": t,
                    "row_count": row_count,
                    "columns": [{"name": c["name"], "type": c["type"] or "TEXT"} for c in cols],
                    "likely_timestamp_columns": likely_ts,
                })
            return {"tables": tables}
        finally:
            ro.close()

    def ingest_sqlite_table(
        self,
        path: str,
        table_name: str,
        name: str | None = None,
        build_fts: bool = True,
        timestamp_columns: list[str] | None = None,
        progress=None,
        cancel=None,
    ) -> dict:
        """Imports one table from an external SQLite file as a new source —
        same TEXT-column, batched-commit convention as ingest_csv (see its
        docstring: self.lock held per BATCH-sized chunk, not for the whole
        import, so a large table doesn't freeze every other request for
        the whole read). `timestamp_columns` — a subset of the *source*
        table's own column names, typically pre-populated from
        preview_sqlite_tables' heuristic and confirmed/edited by the
        analyst — get converted from WebKit/Chrome epoch microseconds to
        an ISO datetime string as they're read, rather than importing the
        opaque integer. That's an explicit, analyst-confirmed conversion,
        not an automatic rewrite of the evidence: the source .db file
        itself is opened read-only and never touched."""
        name = name or f"{os.path.splitext(os.path.basename(path))[0]}.{table_name}"
        timestamp_columns = set(timestamp_columns or [])
        file_hash = self._quick_hash(path)

        ro = sqlite3.connect(f"file:{path}?mode=ro", uri=True, check_same_thread=False)
        ro.row_factory = sqlite3.Row
        try:
            src_cols_info = ro.execute(f"PRAGMA table_info({q(table_name)})").fetchall()
            if not src_cols_info:
                raise ValueError(f"No such table: {table_name}")
            src_colnames = [c["name"] for c in src_cols_info]
            cols = sanitize_columns(src_colnames)
            ncols = len(cols)
            ts_idx = {i for i, sc in enumerate(src_colnames) if sc in timestamp_columns}
            total_rows = ro.execute(f"SELECT COUNT(*) FROM {q(table_name)}").fetchone()[0]

            with self.lock, self.db:
                cur = self.db.execute(
                    "INSERT INTO sources(name, path, table_name, columns, file_hash, imported_at)"
                    " VALUES (?,?,?,?,?,?)",
                    (name, os.path.abspath(path), "", "[]", file_hash, time.strftime("%Y-%m-%dT%H:%M:%S")),
                )
                source_id = cur.lastrowid
                self.db.execute("INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", (source_id,))
                dest_table = f"src_{source_id}"
                coldefs = ", ".join(f"{q(c)} TEXT" for c in cols)
                self.db.execute(f"CREATE TABLE {q(dest_table)} (rid INTEGER PRIMARY KEY, {coldefs})")
                self.db.execute("UPDATE sources SET table_name=? WHERE id=?", (dest_table, source_id))

            placeholders = ",".join("?" * ncols)
            insert = f"INSERT INTO {q(dest_table)} ({','.join(q(c) for c in cols)}) VALUES ({placeholders})"

            def to_text(i, v):
                if v is None:
                    return ""
                if i in ts_idx:
                    iso = _webkit_to_iso(v)
                    if iso:
                        return iso
                if isinstance(v, (bytes, bytearray)):
                    return f"<{len(v)} bytes>"
                return str(v)

            select_cols = ",".join(q(c) for c in src_colnames)
            src_cursor = ro.execute(f"SELECT {select_cols} FROM {q(table_name)}")

            sample: list[list[str]] = []
            total = 0
            t0 = time.time()
            error: Exception | None = None
            with self._ingest_synchronous_off():
                try:
                    while True:
                        rows = src_cursor.fetchmany(BATCH)
                        if not rows:
                            break
                        if cancel is not None and cancel():
                            raise IngestCancelled(f"Import of {name} cancelled")
                        batch = [tuple(to_text(i, v) for i, v in enumerate(r)) for r in rows]
                        if len(sample) < SAMPLE_ROWS:
                            sample.extend(batch[: SAMPLE_ROWS - len(sample)])
                        total = self._commit_ingest_batch(insert, batch, source_id, total)
                        if progress:
                            progress(total, total, total_rows)
                    if progress:
                        progress(total, total, total_rows)
                except IngestCancelled:
                    # Same cancel-discards-the-partial contract as ingest_csv.
                    self.drop_source(source_id)
                    raise
                except Exception as e:
                    error = e
        finally:
            ro.close()

        types = [infer_type([r[i] for r in sample]) for i in range(ncols)] if sample else ["text"] * ncols
        colmeta = [{"name": c, "type": t} for c, t in zip(cols, types)]
        with self.lock, self.db:
            self.db.execute("UPDATE sources SET columns=? WHERE id=?", (json.dumps(colmeta), source_id))

        if error is not None:
            raise error

        if build_fts and total:
            self._ensure_fts_building(source_id)

        elapsed = time.time() - t0
        rec = self.get_source(source_id)
        rec["elapsed_sec"] = round(elapsed, 2)
        rec["rows_per_sec"] = int(total / elapsed) if elapsed > 0 else 0
        return rec

    @staticmethod
    def _json_flatten_depth(flatten_mode: str, flatten_depth: int) -> int | None:
        if flatten_mode == "full":
            return None
        if flatten_mode == "depth":
            return max(0, flatten_depth)
        return 0  # "none"

    def preview_json_file(
        self, path: str, flatten_mode: str = "none", flatten_depth: int = 0, max_rows: int = 50
    ) -> dict:
        """Read-only sample for the import preview UI — same shape as
        preview_csv_text (columns/sample_rows/inferred_types), plus
        record_count so the UI can show how many rows the real ingest will
        produce. Reads the whole file (see _iter_json_records — a .json
        document can't be safely byte-truncated the way CSV's sniff-first-
        chunk can) but writes nothing; column discovery here is the same
        full pass ingest_json's own first pass does, recomputed on every
        flatten-mode change in the UI since the two are independent calls."""
        max_depth = self._json_flatten_depth(flatten_mode, flatten_depth)
        seen_cols: list[str] = []
        seen_set: set[str] = set()
        sample_flat: list[dict[str, str]] = []
        record_count = 0
        for rec in _iter_json_records(path):
            flat = _flatten_json(rec, max_depth)
            for k in flat:
                if k not in seen_set:
                    seen_set.add(k)
                    seen_cols.append(k)
            if len(sample_flat) < max_rows:
                sample_flat.append(flat)
            record_count += 1
        if record_count == 0:
            raise ValueError("File has no records")
        cols = sanitize_columns(seen_cols)
        key_by_col = dict(zip(cols, seen_cols))
        sample_rows = [[flat.get(key_by_col[c], "") for c in cols] for flat in sample_flat]
        types = [infer_type([r[i] for r in sample_rows]) for i in range(len(cols))] if sample_rows else ["text"] * len(cols)
        return {"columns": cols, "sample_rows": sample_rows, "inferred_types": types, "record_count": record_count}

    def ingest_json(
        self,
        path: str,
        name: str | None = None,
        flatten_mode: str = "none",
        flatten_depth: int = 0,
        build_fts: bool = True,
        progress=None,
        cancel=None,
    ) -> dict:
        """Streams a .json/.jsonl file into its own table — same TEXT-
        column, batched-commit convention as ingest_csv (self.lock held per
        BATCH-sized chunk, not the whole import).

        Two full passes rather than one: JSON has no fixed header row the
        way a CSV's first line is one, so which columns exist at all can
        only be known by looking at every record first (same spirit as
        CSV's ragged-row tolerance, just discovered up front here instead
        of per-row) — pass 1 flattens every record only far enough to
        collect the union of column keys, pass 2 re-reads and inserts
        against that now-fixed column set, filling in "" for any key a
        given record doesn't have (same convention as a short CSV row
        getting padded)."""
        name = name or os.path.basename(path)
        file_hash = self._quick_hash(path)
        max_depth = self._json_flatten_depth(flatten_mode, flatten_depth)

        seen_cols: list[str] = []
        seen_set: set[str] = set()
        record_count = 0
        for rec in _iter_json_records(path):
            for k in _flatten_json(rec, max_depth):
                if k not in seen_set:
                    seen_set.add(k)
                    seen_cols.append(k)
            record_count += 1
            # Pass 1 runs before the source row exists, so a cancel here has
            # nothing to clean up — it just stops the scan.
            if cancel is not None and record_count % 50_000 == 0 and cancel():
                raise IngestCancelled(f"Import of {name or os.path.basename(path)} cancelled")
        if record_count == 0:
            raise ValueError("File has no records")

        cols = sanitize_columns(seen_cols)
        ncols = len(cols)
        key_by_col = dict(zip(cols, seen_cols))

        with self.lock, self.db:
            cur = self.db.execute(
                "INSERT INTO sources(name, path, table_name, columns, file_hash, imported_at)"
                " VALUES (?,?,?,?,?,?)",
                (name, os.path.abspath(path), "", "[]", file_hash, time.strftime("%Y-%m-%dT%H:%M:%S")),
            )
            source_id = cur.lastrowid
            self.db.execute("INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", (source_id,))
            table = f"src_{source_id}"
            coldefs = ", ".join(f"{q(c)} TEXT" for c in cols)
            self.db.execute(f"CREATE TABLE {q(table)} (rid INTEGER PRIMARY KEY, {coldefs})")
            self.db.execute("UPDATE sources SET table_name=? WHERE id=?", (table, source_id))

        placeholders = ",".join("?" * ncols)
        insert = f"INSERT INTO {q(table)} ({','.join(q(c) for c in cols)}) VALUES ({placeholders})"

        sample: list[list[str]] = []
        batch: list[tuple] = []
        total = 0
        t0 = time.time()
        error: Exception | None = None
        with self._ingest_synchronous_off():
            try:
                for rec in _iter_json_records(path):
                    flat = _flatten_json(rec, max_depth)
                    row = [flat.get(key_by_col[c], "") for c in cols]
                    batch.append(tuple(row))
                    if len(sample) < SAMPLE_ROWS:
                        sample.append(row)
                    if len(batch) >= BATCH:
                        if cancel is not None and cancel():
                            raise IngestCancelled(f"Import of {name} cancelled")
                        total = self._commit_ingest_batch(insert, batch, source_id, total)
                        batch.clear()
                        if progress:
                            progress(total, total, record_count)
                if batch:
                    if cancel is not None and cancel():
                        raise IngestCancelled(f"Import of {name} cancelled")
                    total = self._commit_ingest_batch(insert, batch, source_id, total)
                    batch.clear()
                if progress:
                    progress(total, total, record_count)
            except IngestCancelled:
                # Same cancel-discards-the-partial contract as ingest_csv.
                self.drop_source(source_id)
                raise
            except Exception as e:
                error = e

        types = [infer_type([r[i] for r in sample]) for i in range(ncols)] if sample else ["text"] * ncols
        colmeta = [{"name": c, "type": t} for c, t in zip(cols, types)]
        with self.lock, self.db:
            self.db.execute("UPDATE sources SET columns=? WHERE id=?", (json.dumps(colmeta), source_id))

        if error is not None:
            raise error

        if build_fts and total:
            self._ensure_fts_building(source_id)

        elapsed = time.time() - t0
        rec = self.get_source(source_id)
        rec["elapsed_sec"] = round(elapsed, 2)
        rec["rows_per_sec"] = int(total / elapsed) if elapsed > 0 else 0
        return rec

    def ingest_rows(
        self,
        columns: list[str],
        rows: Iterable,
        *,
        name: str,
        path: str | None = None,
        build_fts: bool = True,
        column_types: list[str] | None = None,
        progress=None,
    ) -> dict:
        """Generic ingest for rows produced in-process — the path plugin
        ingest formats feed (see plugin_api.py), and the natural seam for
        any future non-file producer. Same conventions as ingest_csv, on
        purpose: all-TEXT columns through sanitize_columns, contiguous rid
        from 1 (invariant #2's root_virtual carve-out depends on it),
        self.lock held per BATCH-sized chunk rather than for the whole
        iterable, ragged rows padded/trimmed and counted rather than
        dropped, types inferred from a SAMPLE_ROWS sample (overridable via
        column_types — a parser that *knows* a column is a datetime
        shouldn't be at the mercy of a 500-row sample), and the FTS build
        kicked off in the background.

        `rows` is any iterable of sequences aligned to `columns` — a
        generator keeps memory flat on multi-GB inputs. Cells may be
        str/int/None; everything is stringified (None -> "") because source
        tables are TEXT and evidence fidelity beats type elegance
        (CLAUDE.md). A mid-iteration parser exception is re-raised after
        the column metadata is written, same partial-import behavior as a
        malformed line late in a CSV: whatever committed stays, with an
        accurate row_count, and the caller sees the error."""
        if not columns:
            raise ValueError("No columns")
        cols = sanitize_columns(list(columns))
        ncols = len(cols)
        file_hash = self._quick_hash(path) if path and os.path.isfile(path) else None

        with self.lock, self.db:
            cur = self.db.execute(
                "INSERT INTO sources(name, path, table_name, columns, file_hash, imported_at)"
                " VALUES (?,?,?,?,?,?)",
                (name, os.path.abspath(path) if path else None, "", "[]", file_hash,
                 time.strftime("%Y-%m-%dT%H:%M:%S")),
            )
            source_id = cur.lastrowid
            self.db.execute("INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", (source_id,))
            table = f"src_{source_id}"
            coldefs = ", ".join(f"{q(c)} TEXT" for c in cols)
            self.db.execute(f"CREATE TABLE {q(table)} (rid INTEGER PRIMARY KEY, {coldefs})")
            self.db.execute("UPDATE sources SET table_name=? WHERE id=?", (table, source_id))

        placeholders = ",".join("?" * ncols)
        insert = f"INSERT INTO {q(table)} ({','.join(q(c) for c in cols)}) VALUES ({placeholders})"

        sample: list[list[str]] = []
        batch: list[tuple] = []
        total = 0
        ragged = 0
        t0 = time.time()
        error: Exception | None = None
        with self._ingest_synchronous_off():
            try:
                for raw in rows:
                    row = ["" if v is None else v if isinstance(v, str) else str(v) for v in raw]
                    if len(row) != ncols:
                        ragged += 1
                        row = (row + [""] * ncols)[:ncols]
                    batch.append(tuple(row))
                    if len(sample) < SAMPLE_ROWS:
                        sample.append(row)
                    if len(batch) >= BATCH:
                        total = self._commit_ingest_batch(insert, batch, source_id, total)
                        batch.clear()
                        if progress:
                            progress(total, 0, 0)
                if batch:
                    total = self._commit_ingest_batch(insert, batch, source_id, total)
                    batch.clear()
            except Exception as e:
                error = e

        types = [infer_type([r[i] for r in sample]) for i in range(ncols)] if sample else ["text"] * ncols
        if column_types:
            types = [(column_types[i] if i < len(column_types) and column_types[i] in ("text", "number", "datetime") else t)
                     for i, t in enumerate(types)]
        colmeta = [{"name": c, "type": t} for c, t in zip(cols, types)]
        with self.lock, self.db:
            self.db.execute("UPDATE sources SET columns=? WHERE id=?", (json.dumps(colmeta), source_id))

        if error is not None:
            raise error

        if build_fts and total:
            self._ensure_fts_building(source_id)

        elapsed = time.time() - t0
        rec = self.get_source(source_id)
        rec["elapsed_sec"] = round(elapsed, 2)
        rec["rows_per_sec"] = int(total / elapsed) if elapsed > 0 else 0
        rec["ragged_rows"] = ragged
        return rec

    @staticmethod
    def _import_pattern_matches(pattern: str, filename: str, rel_path: str) -> bool:
        """A pattern containing '/' matches against the file's path relative
        to the scan root (posix-separated, so a Windows-collected triage's
        backslashes never matter); one without matches the bare filename
        anywhere in the tree. Lets `*_Amcache_UnassociatedFileEntries.csv`
        work as a simple filename glob while `RegistryHives/*` can still
        exclude a whole subfolder. Case-insensitive — KAPE/EZTools output
        casing isn't consistent enough to make analysts type it exactly."""
        pat = pattern.strip().lower()
        if not pat:
            return False
        target = rel_path.lower() if "/" in pat else filename.lower()
        return fnmatch.fnmatch(target, pat)

    def scan_import_directory(
        self, root: str, *, recursive: bool = True,
        extensions: list[str] | None = None,
        include_patterns: list[str] | None = None,
        exclude_patterns: list[str] | None = None,
        filename_patterns: list[str] | None = None,
    ) -> dict:
        """Preview for directory import (server.py's /api/ingest/dir/scan):
        walks `root`, buckets every file it finds into "would import" or
        "excluded, and why" — no ingestion happens here. Pure filesystem +
        fnmatch, no self.lock held beyond one cheap read of already-imported
        paths, so this is safe to call live as the analyst edits patterns.

        A file is excluded for exactly one of: its extension isn't
        recognized ("extension"), include_patterns is non-empty and nothing
        in it matched ("no include pattern matched"), a pattern in
        exclude_patterns matched ("excluded by pattern: <pattern>"), or it
        couldn't be stat'd ("unreadable" — a broken symlink, a permission
        error mid-walk). Surfacing the reason is the point: tuning a
        profile's patterns is a visible feedback loop instead of guessing
        whether "*Amcache*" actually matched anything.

        `already_imported` flags a matched file whose absolute path already
        equals some source's stored path — same abspath-equality convention
        workspace.CaseRegistry.find_by_path already uses, no extra
        normalization added. This never blocks re-import (no hard skip
        here); it's the frontend's default-uncheck signal so re-running the
        same import doesn't silently duplicate every table.

        `filename_patterns` (from plugin-registered ingest formats — see
        plugin_api.py) is a second way past the extension gate, not a
        second include filter: a file whose extension isn't recognized
        still matches when its bare name fnmatches one of these. It exists
        because the files plugins are built for are exactly the ones an
        extension list can never name — "$MFT", "$J" — and it deliberately
        reuses the bare-filename half of _import_pattern_matches' rules so
        "matches" means one thing in this scan. include/exclude patterns
        still apply to these files afterward, unchanged."""
        root_abs = os.path.abspath(root)
        exts = {
            (e if e.startswith(".") else "." + e).lower()
            for e in (extensions or DEFAULT_IMPORT_EXTENSIONS)
        }
        includes = [p for p in (include_patterns or []) if p.strip()]
        excludes = [p for p in (exclude_patterns or []) if p.strip()]
        fname_pats = [p for p in (filename_patterns or []) if p.strip()]

        with self.lock:
            existing_paths = {
                os.path.abspath(r[0]) for r in self.db.execute("SELECT path FROM sources WHERE path IS NOT NULL")
            }

        walker = os.walk(root_abs) if recursive else itertools.islice(os.walk(root_abs), 1)
        matched: list[dict] = []
        excluded: list[dict] = []
        truncated = False
        for dirpath, _dirnames, filenames in walker:
            for fname in sorted(filenames):
                if len(matched) + len(excluded) >= MAX_SCAN_RESULTS:
                    truncated = True
                    break
                fpath = os.path.join(dirpath, fname)
                rel = os.path.relpath(fpath, root_abs).replace(os.sep, "/")
                try:
                    size = os.path.getsize(fpath)
                except OSError:
                    excluded.append({"path": fpath, "rel_path": rel, "reason": "unreadable"})
                    continue
                entry = {"path": fpath, "rel_path": rel, "size_bytes": size}
                ext = os.path.splitext(fname)[1].lower()
                by_filename_pattern = any(fnmatch.fnmatch(fname.lower(), p.lower()) for p in fname_pats)
                if ext not in exts and not by_filename_pattern:
                    excluded.append({**entry, "reason": "extension"})
                    continue
                if includes and not any(self._import_pattern_matches(p, fname, rel) for p in includes):
                    excluded.append({**entry, "reason": "no include pattern matched"})
                    continue
                hit = next((p for p in excludes if self._import_pattern_matches(p, fname, rel)), None)
                if hit:
                    excluded.append({**entry, "reason": f"excluded by pattern: {hit}"})
                    continue
                # kind routes the frontend's per-file import call: the two
                # built-in parsers by their own extensions, everything else
                # ("plugin") resolved client-side against the loaded plugin
                # formats — the scan doesn't know which format claimed an
                # extension/pattern, and doesn't need to.
                if ext in (".json", ".jsonl", ".ndjson"):
                    kind = "json"
                elif ext in DEFAULT_IMPORT_EXTENSIONS:
                    kind = "csv"
                else:
                    kind = "plugin"
                matched.append({
                    **entry,
                    "kind": kind,
                    "already_imported": fpath in existing_paths,
                })
            if truncated:
                break

        matched.sort(key=lambda e: e["rel_path"])
        excluded.sort(key=lambda e: e["rel_path"])
        return {"root": root_abs, "matched": matched, "excluded": excluded, "truncated": truncated}

    # ------------------------------------------------------------ ingest jobs

    MAX_CONCURRENT_INGESTS = 2  # parse work is CPU-bound; more just thrash
    INGEST_JOB_KEEP = 20        # finished jobs kept for the UI panel

    def start_ingest_job(self, kind: str, path: str, *, name: str | None = None,
                         options: dict | None = None, delete_after: bool = False) -> dict:
        """Runs an ingest on a background daemon thread and returns its job
        record immediately — the exact reason start_search_all_job exists,
        applied to the operation that takes the longest of anything in the
        app. A 50 GB import used to be one multi-minute POST the analyst
        sat behind with no progress, no way to tell "working" from
        "crashed", and a browser tab they couldn't use meanwhile.

        `kind` is 'csv' | 'json' | 'sqlite'. For 'sqlite',
        options['tables'] is [{table, name?, timestamp_columns?}, ...] —
        one uploaded file, one job, N sources, so the file is spooled and
        read once rather than re-uploaded per table. `delete_after` removes
        `path` when the job ends (the upload route's tempfile — which the
        old synchronous upload endpoints leaked on disk, tens of GB at
        this tool's file sizes).

        Progress lands in the job record from the per-BATCH `progress`
        callback the ingest paths already had (rows plus a unit/units_total
        pair: bytes for CSV — a percentage without pre-scanning the file —
        records for JSON, rows for SQLite). Cancellation is cooperative per
        BATCH; a cancelled ingest *drops* its partial source (see
        ingest_csv).

        'derive' is the odd one out: no file at all (`path` is ''), it
        backfills one derived column's values from a column already in the
        case. It rides this machinery rather than growing a second job
        system because it wants exactly what this one provides — a progress
        bar over a multi-million-row pass, per-BATCH cancellation, the jobs
        panel, and close()'s cancel-and-join."""
        if kind not in ("csv", "json", "sqlite", "derive"):
            raise ValueError(f"Unknown ingest kind: {kind}")
        try:
            size = os.path.getsize(path)
        except OSError:
            size = 0
        with self._ingest_jobs_lock:
            self._ingest_job_seq += 1
            job = {
                "job_id": self._ingest_job_seq,
                "kind": kind,
                "name": name or os.path.basename(path),
                "path": path,
                "status": "queued",
                "rows_done": 0,
                "units_done": 0,
                "units_total": size if kind == "csv" else ((options or {}).get("units_total") or 0),
                "unit": "bytes" if kind == "csv" else ("records" if kind == "json" else "rows"),
                "tables_done": 0,
                "tables_total": len((options or {}).get("tables") or []) if kind == "sqlite" else 0,
                "current_table": None,
                "source_ids": [],
                "result": None,
                "error": None,
                "cancelled": False,
                "started_at": time.time(),
                "finished_at": None,
                "options": options or {},
                "delete_after": delete_after,
                "thread": None,
            }
            self._ingest_jobs[job["job_id"]] = job
            self._prune_ingest_jobs_locked()
        t = threading.Thread(target=self._ingest_job_worker, args=(job,), daemon=True)
        with self._ingest_jobs_lock:
            job["thread"] = t
        t.start()
        return self._ingest_job_snapshot(job)

    def _ingest_job_worker(self, job: dict) -> None:
        acquired = False
        try:
            self._ingest_sem.acquire()
            acquired = True
            with self._ingest_jobs_lock:
                if job["cancelled"]:
                    job["status"] = "cancelled"
                    return
                job["status"] = "running"

            def cancel() -> bool:
                return job["cancelled"]  # plain bool read; set under the jobs lock

            def progress(rows: int, units: int, units_total: int) -> None:
                with self._ingest_jobs_lock:
                    job["rows_done"] = rows
                    job["units_done"] = units
                    job["units_total"] = units_total

            opts = job["options"]
            if job["kind"] == "csv":
                results = [self.ingest_csv(
                    job["path"], name=job["name"],
                    delimiter=opts.get("delimiter"), build_fts=opts.get("build_fts", True),
                    has_header=opts.get("has_header", True), column_types=opts.get("column_types"),
                    progress=progress, cancel=cancel,
                )]
            elif job["kind"] == "derive":
                # options carries def_ids (a flatten adding several columns
                # in one pass) or def_id (the single-column create and
                # re-derive paths) — one job either way.
                res = self.backfill_derived_columns(
                    opts.get("def_ids") or [opts["def_id"]], progress=progress, cancel=cancel,
                    drop_on_cancel=opts.get("drop_on_cancel", False),
                )
                with self._ingest_jobs_lock:
                    job["status"] = "done"
                    job["rows_done"] = res["rows"]
                    job["source_ids"] = [res["source_id"]]
                    job["result"] = [dict(c, rows=res["rows"]) for c in res["columns"]]
                return
            elif job["kind"] == "json":
                results = [self.ingest_json(
                    job["path"], name=job["name"],
                    flatten_mode=opts.get("flatten_mode", "none"),
                    flatten_depth=opts.get("flatten_depth", 0),
                    build_fts=opts.get("build_fts", True),
                    progress=progress, cancel=cancel,
                )]
            else:  # sqlite: one job, N tables out of the one spooled file
                results = []
                for i, t in enumerate(opts.get("tables") or []):
                    if job["cancelled"]:
                        raise IngestCancelled("cancelled between tables")
                    with self._ingest_jobs_lock:
                        job["current_table"] = t["table"]
                        job["tables_done"] = i
                    results.append(self.ingest_sqlite_table(
                        job["path"], t["table"], name=t.get("name"),
                        build_fts=opts.get("build_fts", True),
                        timestamp_columns=t.get("timestamp_columns"),
                        progress=progress, cancel=cancel,
                    ))
                    with self._ingest_jobs_lock:
                        job["tables_done"] = i + 1
            with self._ingest_jobs_lock:
                job["status"] = "done"
                job["rows_done"] = sum(r.get("row_count") or 0 for r in results)
                job["source_ids"] = [r["id"] for r in results]
                job["result"] = [
                    {k: r.get(k) for k in ("id", "name", "row_count", "elapsed_sec", "rows_per_sec", "ragged_rows")}
                    for r in results
                ]
        except IngestCancelled:
            with self._ingest_jobs_lock:
                job["status"] = "cancelled"
        except Exception as e:  # noqa: BLE001 — surfaced to the UI as job.error
            with self._ingest_jobs_lock:
                job["status"] = "error"
                job["error"] = str(e)
        finally:
            with self._ingest_jobs_lock:
                job["finished_at"] = time.time()
            if acquired:
                self._ingest_sem.release()
            if job["delete_after"]:
                with contextlib.suppress(OSError):
                    os.remove(job["path"])

    def _ingest_job_snapshot(self, job: dict) -> dict:
        keys = ("job_id", "kind", "name", "status", "rows_done", "units_done",
                "units_total", "unit", "tables_done", "tables_total", "current_table",
                "source_ids", "result", "error", "cancelled", "started_at", "finished_at")
        with self._ingest_jobs_lock:
            return {k: job[k] for k in keys}

    def _prune_ingest_jobs_locked(self) -> None:
        done = [j for j in self._ingest_jobs.values()
                if j["status"] in ("done", "error", "cancelled")]
        excess = len(done) - self.INGEST_JOB_KEEP
        if excess > 0:
            done.sort(key=lambda j: j["finished_at"] or 0)
            for j in done[:excess]:
                self._ingest_jobs.pop(j["job_id"], None)

    def list_ingest_jobs(self) -> list[dict]:
        with self._ingest_jobs_lock:
            ids = sorted(self._ingest_jobs, reverse=True)
            jobs = [self._ingest_jobs[i] for i in ids]
        return [self._ingest_job_snapshot(j) for j in jobs]

    def cancel_ingest_job(self, job_id: int) -> bool:
        with self._ingest_jobs_lock:
            job = self._ingest_jobs.get(job_id)
            if job is None or job["status"] in ("done", "error", "cancelled"):
                return False
            job["cancelled"] = True
            return True

    def wait_for_ingest_job(self, job_id: int | None = None, timeout: float | None = None) -> dict | None:
        """Blocks until the given (or every) ingest job finishes. Tests
        only; nothing in the request path waits on a job."""
        with self._ingest_jobs_lock:
            jobs = [j for j in self._ingest_jobs.values()
                    if job_id is None or j["job_id"] == job_id]
        snap = None
        for j in jobs:
            t = j.get("thread")
            if t and t.is_alive():
                t.join(timeout)
            snap = self._ingest_job_snapshot(j)
        return snap

    def build_fts(self, source_id: int) -> None:
        """(Re)builds the trigram substring index for a source: a single
        `doc` column (every column concatenated, via a `src_<id>_doc` view
        over `_blob_expr`) indexed with `tokenize='trigram', detail=none,
        columnsize=0`, queried with a bare `doc LIKE ?` (never MATCH — see
        below). Trigram indexing 3-character sequences is what makes
        Contains-mode search both correct *and* fast for a fragment buried
        mid-token (e.g. "jacso" inside `C:\\users\\jacso\\desktop\\file.txt`)
        — a real index lookup instead of a full-table LIKE scan.

        Why this exact shape — each of these was measured, not assumed
        (332K-row/285MB EvtxECmd source):

        - `detail=none, columnsize=0` drops the per-occurrence position
          lists (which nothing here uses — no ranking, no highlighting),
          cutting the index from 892MB to 143MB (–84%). The cost is that
          FTS5 must verify each candidate against the real text, so query
          time scales with *result count*: identical for rare/medium terms
          (the IOC-hunting case), ~0.8s worst-case for a term matching 94%
          of all rows — a degenerate result set anyway.
        - Under detail=none a multi-trigram MATCH is a *phrase* query and
          errors outright ("phrase queries are not supported"); the query
          form is SQLite 3.45+'s LIKE pushdown instead. That pushdown is
          per-column — a table-level `fts LIKE ?` runs against the hidden
          table-name column (NULL outside MATCH) and silently returns 0
          rows — which is why the index is one concatenated `doc` column
          rather than mirroring the source's columns.
        - content= points at the src_<id>_doc *view*, so candidate
          verification recomputes the concat on demand — no second stored
          copy of the data.

        Batches its own commits (DROP/CREATE once, then BATCH-sized INSERT
        chunks keyed by rid range) instead of one multi-second transaction,
        so a build never freezes every other request for its duration —
        same reasoning, and the same BATCH size, as ingest_csv's own
        chunked commits. Called synchronously here; _ensure_fts_building is
        what runs this on a background thread so ingest/search never block
        waiting for it.

        Each chunk also yields the thread briefly (time.sleep) right after
        releasing the lock, not just released-and-immediately-reacquired:
        measured directly, a tight release/reacquire loop with no yield
        starves a foreground request waiting on the same lock for several
        chunks in a row before the OS scheduler gives it a turn — one
        request stalled >13s despite every individual chunk holding the
        lock for under a second. A ~20ms yield dropped that worst case to
        ~1.4s (one chunk's worth, as intended) for a ~5% longer total
        build.

        Deliberately *not* followed by an FTS5 'optimize' (or an incremental
        'merge' loop), despite that being the usual advice after a chunked
        build. Measured on a 400K-row index built in BATCH-sized chunks:
        optimize does consolidate segments (the structure record shrank
        51->24 bytes) but query time was unchanged on every shape tried —
        rare term 4.6ms -> 4.0ms, a term matching every row 205ms -> 208ms,
        a miss 0.1ms -> 0.0ms, all inside run-to-run noise — while the case
        file grew ~7% (48.8MB -> 52.2MB), since the pre-merge pages are
        freed to the freelist and nothing here VACUUMs. Two reasons it
        doesn't pay: FTS5's own automerge already consolidates as the chunks
        go in, and under detail=none query cost is dominated by verifying
        candidates against the content view, not by walking segments."""
        src = self._source_lite(source_id)
        # Base columns only. A derived column is computed from a column
        # that's already in the doc blob, so indexing it would only make
        # the same evidence match twice — and keeping the doc view over the
        # source table alone means the index needs no rebuild when a
        # derived column is added or removed.
        table, cols = src["table_name"], [c["name"] for c in self._base_cols(src)]
        fts = f"fts_{source_id}"
        doc_view = f"{table}_doc"
        with self.lock, self.db:
            self.db.execute(f"DROP TABLE IF EXISTS {q(fts)}")
            self.db.execute(f"DROP VIEW IF EXISTS {q(doc_view)}")
            self.db.execute(
                f"CREATE VIEW {q(doc_view)} AS SELECT rid, {_blob_expr(cols)} AS doc FROM {q(table)}"
            )
            self.db.execute(
                f"CREATE VIRTUAL TABLE {q(fts)} USING fts5(doc, "
                f"content={q(doc_view)}, content_rowid='rid', tokenize='trigram', "
                f"detail=none, columnsize=0)"
            )
            max_rid = self.db.execute(f"SELECT COALESCE(MAX(rid), 0) FROM {q(table)}").fetchone()[0]
        for start in range(1, max_rid + 1, BATCH):
            with self.lock, self.db:
                self.db.execute(
                    f"INSERT INTO {q(fts)}(rowid, doc) "
                    f"SELECT rid, doc FROM {q(doc_view)} WHERE rid >= ? AND rid < ?",
                    (start, start + BATCH),
                )
            time.sleep(0.02)
        with self.lock, self.db:
            self.db.execute("UPDATE sources SET has_fts=1 WHERE id=?", (source_id,))

    @staticmethod
    def _sniff(head: str) -> str:
        try:
            return csv.Sniffer().sniff(head[:8192], delimiters=",\t;|").delimiter
        except csv.Error:
            first = head.splitlines()[0] if head else ""
            return max(",\t;|", key=first.count)

    @staticmethod
    def _quick_hash(path: str) -> str:
        """Size + head/tail digest. Fast on multi-GB files and stable enough to
        recognise the same evidence file across imports."""
        h = hashlib.sha256()
        size = os.path.getsize(path)
        h.update(str(size).encode())
        with open(path, "rb") as f:
            h.update(f.read(1 << 20))
            if size > (1 << 21):
                f.seek(-(1 << 20), os.SEEK_END)
                h.update(f.read())
        return h.hexdigest()[:32]

    # ----------------------------------------------------------------- sources

    def list_sources(self) -> list[dict]:
        """Bulk per-source annotation (is_open / tagged_row_count /
        note_count) in one locked pass — a per-source query here would be an
        N+1 against row_tags/row_notes/open_tabs on every case with several
        tables open, which is exactly the "8+ tables" case this exists to
        support."""
        with self.lock:
            rows = self.db.execute("SELECT * FROM sources ORDER BY id").fetchall()
            open_ids = {r[0] for r in self.db.execute("SELECT source_id FROM open_tabs")}
            tagged = dict(self.db.execute("SELECT source_id, COUNT(DISTINCT rid) FROM row_tags GROUP BY source_id"))
            notes = dict(self.db.execute("SELECT source_id, COUNT(*) FROM row_notes GROUP BY source_id"))
            derived = self.db.execute(
                "SELECT * FROM derived_columns ORDER BY source_id, id"
            ).fetchall()
        by_src: dict[int, list] = {}
        for r in derived:
            by_src.setdefault(r["source_id"], []).append(r)
        out = []
        for r in rows:
            d = self._src_dict(r)
            for dr in by_src.get(d["id"], []):
                d["columns"].append(self._derived_col_entry(dr))
                d["has_derived"] = True
            d["is_open"] = d["id"] in open_ids
            d["tagged_row_count"] = tagged.get(d["id"], 0)
            d["note_count"] = notes.get(d["id"], 0)
            d["fts_building"] = self._is_fts_building(d["id"])
            out.append(d)
        return out

    def _tab_meta(self, source_id: int) -> dict:
        """is_open for a single signed source_id (real or merge). Not
        tagged_row_count/note_count for merges — row_tags/row_notes are
        never keyed by a negative id, so callers building a merge's dict
        must sum its members' counts instead (see _merge_source_dict)."""
        with self.lock:
            is_open = self.db.execute("SELECT 1 FROM open_tabs WHERE source_id=?", (source_id,)).fetchone() is not None
        return {"is_open": is_open}

    def set_tab_open(self, source_id: int, open_: bool) -> None:
        with self.lock, self.db:
            if open_:
                self.db.execute("INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", (source_id,))
            else:
                self.db.execute("DELETE FROM open_tabs WHERE source_id=?", (source_id,))

    def set_source_nickname(self, source_id: int, nickname: str | None) -> dict:
        """A display name the analyst chooses ("DC01 security log") over the
        imported file's own name ("20240611_EvtxECmd_Output.csv"). `name`
        is deliberately left untouched — it's the file's identity (session
        matching, the record of what was imported), where a nickname is
        presentation. Empty/whitespace clears it. A merge has no file
        behind it, so its name *is* its display name — nicknaming one is
        just a rename of merges.name."""
        nickname = (nickname or "").strip() or None
        if nickname and len(nickname) > 200:
            raise ValueError("That nickname is too long")
        if source_id < 0:
            if not nickname:
                raise ValueError("A merge needs a name — give it one rather than clearing it")
            # Renaming is the second door into merges.name; the uniqueness
            # rule means nothing if it only guards the first.
            if self._merge_name_taken(nickname, exclude_id=-source_id):
                raise ValueError(f'A merge named "{nickname}" already exists')
            with self.lock, self.db:
                cur = self.db.execute("UPDATE merges SET name=? WHERE id=?", (nickname, -source_id))
                if cur.rowcount == 0:
                    raise KeyError(f"No merge {-source_id}")
        else:
            with self.lock, self.db:
                cur = self.db.execute("UPDATE sources SET nickname=? WHERE id=?", (nickname, source_id))
                if cur.rowcount == 0:
                    raise KeyError(f"No source {source_id}")
        return self.get_source(source_id)

    def get_source(self, source_id: int) -> dict:
        """A merge is addressed as a negative source_id (merges.id=m ->
        source_id=-m) so it can flow through every existing source_id: int
        endpoint unchanged. Real sources.id is always >= 1, so this can't
        collide."""
        if source_id < 0:
            return self._merge_source_dict(-source_id)
        with self.lock:
            row = self.db.execute("SELECT * FROM sources WHERE id=?", (source_id,)).fetchone()
            if not row:
                raise KeyError(f"No source {source_id}")
            tagged = self.db.execute("SELECT COUNT(DISTINCT rid) FROM row_tags WHERE source_id=?", (source_id,)).fetchone()[0]
            note_count = self.db.execute("SELECT COUNT(*) FROM row_notes WHERE source_id=?", (source_id,)).fetchone()[0]
            derived = self.db.execute(
                "SELECT * FROM derived_columns WHERE source_id=? ORDER BY id", (source_id,)
            ).fetchall()
        d = self._src_dict(row)
        for dr in derived:
            d["columns"].append(self._derived_col_entry(dr))
            d["has_derived"] = True
        d.update(self._tab_meta(source_id))
        d["tagged_row_count"] = tagged
        d["note_count"] = note_count
        d["fts_building"] = self._is_fts_building(source_id)
        return d

    @staticmethod
    def _src_dict(row: sqlite3.Row) -> dict:
        d = dict(row)
        d["columns"] = json.loads(d["columns"])
        return d

    @staticmethod
    def _derived_col_entry(r: sqlite3.Row) -> dict:
        """A derived_columns row -> the column-list entry merged into
        src["columns"]. Carries everything the UI needs to mark, manage and
        type the column; the `derived` flag is what the handful of
        base-only sites (_base_cols) filter on."""
        op = timeparse.OPERATIONS.get(r["op_id"]) or {}
        return {
            "name": r["name"],
            "type": op.get("value_type", "datetime"),
            "derived": True,
            "derived_id": r["id"],
            "derived_from": r["input_column"],
            "derived_op": r["op_id"],
            "derived_kind": op.get("derived_kind", "datetime"),
            "derived_status": r["status"],
            "parse_failures": r["parse_failures"],
            # The params come along so the UI can show what the column is
            # defined by without a second round trip — for an extracted
            # column that's the field path, which is the whole definition.
            "derived_params": json.loads(r["params"] or "{}"),
        }

    def _source_lite(self, source_id: int) -> dict:
        """Cheap source metadata for internal hot paths: just the sources
        row (columns parsed), or a merge synthesized from its members'
        rows. None of get_source's per-call annotations — tagged_row_count
        is a COUNT(DISTINCT rid) over row_tags (~12ms on a heavily tagged
        source), note_count/is_open are two more locked queries — and
        fetch_rows used to pay all of that at least twice per page, on
        every scroll. Anything user-facing that actually shows those
        counts keeps using get_source; everything that only needs
        table_name/columns/row_count/has_fts to run a query uses this."""
        with self.lock:
            return self._source_lite_on(self.db, source_id)

    def _source_lite_on(self, conn: sqlite3.Connection, source_id: int) -> dict:
        """The query logic behind _source_lite, parameterised on the
        connection: writer-side callers hold self.lock and pass self.db,
        _reader() paths pass their checked-out read-only connection and
        never touch the lock at all. A merge (negative id) synthesizes the
        same dict _merge_source_dict builds minus the per-member
        tag/note/tab counts, with the same live-members contract (a dropped
        member is a KeyError, not a stale merge)."""
        if source_id < 0:
            merge_id = -source_id
            row = conn.execute("SELECT * FROM merges WHERE id=?", (merge_id,)).fetchone()
            if not row:
                raise KeyError(f"No merge {merge_id}")
            member_ids = json.loads(row["source_ids"])
            members = [self._source_lite_on(conn, sid) for sid in member_ids]
            return {
                "id": source_id,
                "name": row["name"],
                "path": None,
                "table_name": None,
                "row_count": sum(m["row_count"] for m in members),
                # canonical — lowest-id (first-created) member, plus every
                # derived column present on ALL members under the same name
                # (each member reads it from its own drv_ sidecar; a column
                # missing on any member would need NULL-padding nothing
                # here does, so it simply isn't exposed until it exists
                # everywhere).
                "columns": ([c for c in members[0]["columns"] if not c.get("derived")]
                            + [c for c in members[0]["columns"] if c.get("derived")
                               and all(any(mc.get("derived") and mc["name"].lower() == c["name"].lower()
                                           for mc in m["columns"]) for m in members[1:])]),
                "file_hash": None,
                "imported_at": row["created_at"],
                "has_fts": 1 if all(m["has_fts"] for m in members) else 0,
                "is_merge": True,
                "member_source_ids": member_ids,
            }
        row = conn.execute("SELECT * FROM sources WHERE id=?", (source_id,)).fetchone()
        if not row:
            raise KeyError(f"No source {source_id}")
        d = self._src_dict(row)
        for dr in conn.execute(
            "SELECT * FROM derived_columns WHERE source_id=? ORDER BY id", (source_id,)
        ):
            d["columns"].append(self._derived_col_entry(dr))
            d["has_derived"] = True
        return d

    def _merge_row(self, merge_id: int) -> sqlite3.Row:
        with self.lock:
            row = self.db.execute("SELECT * FROM merges WHERE id=?", (merge_id,)).fetchone()
        if not row:
            raise KeyError(f"No merge {merge_id}")
        return row

    def _merge_source_dict(self, merge_id: int) -> dict:
        """Synthesizes a get_source()-shaped dict for a merge, computed live
        from its members' CURRENT columns/row_count/has_fts each time it's
        opened — a member dropped after the merge was created surfaces as a
        clear KeyError here rather than a silently stale/broken merge."""
        row = self._merge_row(merge_id)
        member_ids = json.loads(row["source_ids"])
        members = [self.get_source(sid) for sid in member_ids]
        return {
            "id": -merge_id,
            "name": row["name"],
            "path": None,
            "table_name": None,
            "row_count": sum(m["row_count"] for m in members),
            # canonical — lowest-id (first-created) member, plus the derived
            # columns present on every member (same intersection rule as
            # _source_lite_on, which documents why)
            "columns": ([c for c in members[0]["columns"] if not c.get("derived")]
                        + [c for c in members[0]["columns"] if c.get("derived")
                           and all(any(mc.get("derived") and mc["name"].lower() == c["name"].lower()
                                       for mc in m["columns"]) for m in members[1:])]),
            "file_hash": None,
            "imported_at": row["created_at"],
            "has_fts": 1 if all(m["has_fts"] for m in members) else 0,
            "fts_building": any(m["fts_building"] for m in members),
            "is_merge": True,
            "member_source_ids": member_ids,
            "is_open": self._tab_meta(-merge_id)["is_open"],
            "tagged_row_count": sum(m["tagged_row_count"] for m in members),
            "note_count": sum(m["note_count"] for m in members),
        }

    def _resolve_members(self, source_id: int) -> list[dict]:
        """Real source -> single identity-mapped element. Merge -> one
        element per member. SQLite resolves quoted identifiers
        case-insensitively, so canonical column names work directly against
        every member's table without needing per-member name translation —
        the locked-in 'exact column match' rule already guarantees members
        share the same names modulo case."""
        with self.lock:
            return self._resolve_members_on(self.db, source_id)

    def _resolve_members_on(self, conn: sqlite3.Connection, source_id: int) -> list[dict]:
        """_resolve_members parameterised on the connection, same split as
        _source_lite/_source_lite_on."""
        if source_id > 0:
            src = self._source_lite_on(conn, source_id)
            return [{"source_id": source_id, "table_name": src["table_name"]}]
        merge = self._source_lite_on(conn, source_id)
        return [
            {"source_id": sid, "table_name": self._source_lite_on(conn, sid)["table_name"]}
            for sid in merge["member_source_ids"]
        ]

    def _merge_name_taken(self, name: str, exclude_id: int | None = None) -> bool:
        """Case-insensitive: two merges named "EVTX" and "evtx" are one
        typo'd reference apart in every list they'd both appear in, which is
        the confusion a uniqueness rule exists to prevent. NOCASE is ASCII-
        only, matching how header sets are already keyed (str.lower)."""
        with self.lock:
            row = self.db.execute(
                "SELECT id FROM merges WHERE name = ? COLLATE NOCASE AND id != ?",
                (name, exclude_id or 0),
            ).fetchone()
        return row is not None

    def create_merge(self, name: str, source_ids: list[int]) -> dict:
        name = (name or "").strip()
        if not name:
            raise ValueError("A merge needs a name")
        if self._merge_name_taken(name):
            raise ValueError(f'A merge named "{name}" already exists')
        if len(source_ids) < 2:
            raise ValueError("A merge needs at least 2 sources")
        sources = [self.get_source(sid) for sid in source_ids]
        # Base columns only: adding a derived column to one member must not
        # change whether two files of the same format can merge.
        sigs = {column_signature([c["name"] for c in s["columns"] if not c.get("derived")]) for s in sources}
        if len(sigs) > 1:
            raise ValueError("Selected sources don't have matching columns")
        with self.lock, self.db:
            cur = self.db.execute(
                "INSERT INTO merges(name, source_ids, created_at) VALUES (?,?,?)",
                (name, json.dumps(source_ids), time.strftime("%Y-%m-%dT%H:%M:%S")),
            )
            merge_id = cur.lastrowid
            self.db.execute("INSERT OR IGNORE INTO open_tabs(source_id) VALUES (?)", (-merge_id,))
        return self.get_source(-merge_id)

    def list_merges(self) -> list[dict]:
        with self.lock:
            rows = self.db.execute("SELECT * FROM merges ORDER BY id").fetchall()
        out = []
        for r in rows:
            try:
                out.append(self.get_source(-r["id"]))
            except KeyError as e:
                # a member was dropped since the merge was created — surface
                # it as a broken entry the UI can offer to delete, not a
                # silent omission from the list.
                out.append({
                    "id": -r["id"], "name": r["name"], "is_merge": True,
                    "error": str(e), "member_source_ids": json.loads(r["source_ids"]),
                })
        return out

    def delete_merge(self, merge_id: int) -> None:
        with self.lock, self.db:
            self.db.execute("DELETE FROM merges WHERE id=?", (merge_id,))
            self.db.execute("DELETE FROM open_tabs WHERE source_id=?", (-merge_id,))

    def drop_source(self, source_id: int) -> None:
        src = self.get_source(source_id)
        with self.lock, self.db:
            self.db.execute(f"DROP TABLE IF EXISTS {q(src['table_name'])}")
            self.db.execute(f"DROP TABLE IF EXISTS {q('fts_' + str(source_id))}")
            self.db.execute(f"DROP VIEW IF EXISTS {q(src['table_name'] + '_doc')}")
            for t in ("row_tags", "row_notes"):
                self.db.execute(f"DELETE FROM {t} WHERE source_id=?", (source_id,))
            self.db.execute("DELETE FROM layouts WHERE source_id=?", (source_id,))
            self.db.execute("DELETE FROM saved_views WHERE source_id=?", (source_id,))
            self.db.execute("DELETE FROM open_tabs WHERE source_id=?", (source_id,))
            self.db.execute("DELETE FROM sources WHERE id=?", (source_id,))
            self.db.execute(f"DROP TABLE IF EXISTS {q(self._derived_table(source_id))}")
            self.db.execute("DELETE FROM derived_columns WHERE source_id=?", (source_id,))
        self._maxlen_cache.pop(source_id, None)

    # -------------------------------------------------------- derived columns

    @staticmethod
    def _derived_table(source_id: int) -> str:
        return f"drv_{source_id}"

    @staticmethod
    def _base_cols(src: dict) -> list[dict]:
        """The source's own columns, as imported. Used by everything that
        must see the *evidence file's* shape rather than the analyst's
        additions: the FTS doc view and its LIKE-fallback twin (a derived
        value is computed from data that's already searchable), merge
        eligibility, and the session file's column list."""
        return [c for c in src["columns"] if not c.get("derived")]

    def _from_clause(self, src: dict, alias: str | None = None) -> str:
        """FROM-clause text for reading a source's rows: the source table
        alone, or LEFT JOIN'd to its derived-value sidecar.

        `USING(rid)` rather than `ON a.rid = b.rid` so an unqualified `rid`
        stays legal on both sides (every read path selects it that way).
        drv_<id> has rid as its INTEGER PRIMARY KEY, so the join matches at
        most one row and can never change the row count or their order —
        which is what keeps _fetch_virtual_root_rows' `pos = rid - 1`
        exact (invariant #2). A source with no derived columns compiles
        byte-identical SQL to before this feature existed."""
        t = q(src["table_name"])
        if alias:
            t = f"{t} {alias}"
        if not any(c.get("derived") for c in src["columns"]):
            return t
        return f"{t} LEFT JOIN {q(self._derived_table(src['id']))} USING(rid)"

    def _col_ref(self, src: dict, column: str, alias: str | None = None,
                 derived_alias: str | None = None) -> str:
        """SQL reference to one of a source's columns. Base columns take the
        caller's table alias (`s."Time"`); a derived column lives in the
        joined sidecar instead, so it's referenced bare under _from_clause's
        USING(rid) shape, or by `derived_alias` in the chained-join shapes
        (group_summary's view join) that can't use USING. Unambiguous either
        way because add_derived_column refuses a name that collides with a
        base column, case-insensitively — how SQLite resolves quoted
        identifiers."""
        for c in src["columns"]:
            if c.get("derived") and c["name"].lower() == column.lower():
                return f"{derived_alias}.{q(column)}" if derived_alias else q(column)
        return f"{alias}.{q(column)}" if alias else q(column)

    def _is_derived(self, src: dict, column: str) -> bool:
        return any(c.get("derived") and c["name"].lower() == str(column).lower()
                   for c in src["columns"])

    def _member_from(self, src: dict, m: dict, alias: str | None = None) -> str:
        """FROM text for one member table of a (possibly merged) view. A
        plain source picks up its own sidecar join; a merge that exposes
        derived columns (present on every member) reads each member through
        that member's OWN sidecar — same USING(rid) shape, so bare column
        and rid references stay legal."""
        if not src.get("is_merge") and m["source_id"] == src["id"]:
            return self._from_clause(src, alias)
        if src.get("is_merge") and any(c.get("derived") for c in src["columns"]):
            return self._from_clause(self._source_lite(m["source_id"]), alias)
        t = q(m["table_name"])
        return f"{t} {alias}" if alias else t

    def _derived_join(self, src: dict, on_alias: str, alias: str = "d") -> str:
        """`LEFT JOIN drv_<id> d ON d.rid = <on_alias>.rid`, for the query
        shapes that already chain a JOIN ... ON (group_summary's view join)
        and so can't use _from_clause's USING(rid) form — an ON clause after
        a USING join would bind to the wrong join."""
        if src.get("is_merge") or not any(c.get("derived") for c in src["columns"]):
            return ""
        return f" LEFT JOIN {q(self._derived_table(src['id']))} {alias} ON {alias}.rid = {on_alias}.rid"

    def _member_derived_join(self, src: dict, m: dict, on_alias: str) -> str:
        """_derived_join's per-member counterpart for the view-join shapes:
        a merge's rows resolve derived values through each member's own
        sidecar. For a plain source it emits exactly what _derived_join
        would."""
        if not any(c.get("derived") for c in src["columns"]):
            return ""
        mid = m["source_id"]
        msrc = src if (not src.get("is_merge") and mid == src["id"]) else self._source_lite(mid)
        if not any(c.get("derived") for c in msrc["columns"]):
            return ""
        return f" LEFT JOIN {q(self._derived_table(mid))} d ON d.rid = {on_alias}.rid"

    @staticmethod
    def _derived_dict(row: sqlite3.Row) -> dict:
        d = dict(row)
        d["params"] = json.loads(d["params"] or "{}")
        return d

    def list_derived_ops(self) -> list[dict]:
        """The parsing operations the UI offers, straight from the
        timeparse registry (id/label/description/param schema)."""
        return timeparse.list_ops()

    def list_derived_columns(self, source_id: int) -> list[dict]:
        if source_id < 0:
            # A merge's derived columns are the ones its column list exposes
            # (present on every member) — represented by the first member's
            # definitions, since that's the canonical column source.
            src = self._source_lite(source_id)
            exposed = {c["name"].lower() for c in src["columns"] if c.get("derived")}
            first = self._resolve_members(source_id)[0]["source_id"]
            return [d for d in self.list_derived_columns(first) if d["name"].lower() in exposed]
        with self.lock:
            rows = self.db.execute(
                "SELECT * FROM derived_columns WHERE source_id=? ORDER BY id", (source_id,)
            ).fetchall()
        return [self._derived_dict(r) for r in rows]

    def get_derived_column(self, def_id: int) -> dict:
        with self.lock:
            row = self.db.execute("SELECT * FROM derived_columns WHERE id=?", (def_id,)).fetchone()
        if not row:
            raise KeyError(f"No derived column {def_id}")
        return self._derived_dict(row)

    def _find_column(self, src: dict, name: str) -> dict | None:
        for c in src["columns"]:
            if c["name"].lower() == str(name).lower():
                return c
        return None

    def add_derived_column(self, source_id: int, name: str, input_column: str,
                           op_id: str, params: dict | None = None) -> dict:
        """Define a derived column and kick off its backfill job.

        The values are materialised into the drv_<source_id> sidecar rather
        than computed per query: they have to be sortable, filterable,
        groupable and exportable, all of which are server-side SQL over a
        column that either exists or doesn't. The source table itself is
        never touched (invariant #1)."""
        if source_id < 0:
            # A merge has no table of its own — creating "on the merge"
            # means creating the same column on every member, after which
            # _source_lite's intersection rule exposes it on the merge.
            # All-or-nothing: validate against every member before touching
            # any, so a name collision on member 3 doesn't leave members
            # 1-2 with a column the merge never grows.
            members = self._resolve_members(source_id)
            for m in members:
                msrc = self._source_lite(m["source_id"])
                if self._find_column(msrc, (name or "").strip()):
                    raise ValueError(
                        f"Member table {msrc['name']!r} already has a column called {(name or '').strip()!r}")
                if self._find_column(msrc, input_column) is None:
                    raise ValueError(f"Member table {msrc['name']!r} has no column {input_column!r}")
            results = [self.add_derived_column(m["source_id"], name, input_column, op_id, params)
                       for m in members]
            return {"definition": results[0]["definition"],
                    "member_definitions": [r["definition"] for r in results],
                    "job_id": results[0]["job_id"],
                    "job_ids": [r["job_id"] for r in results]}
        op = timeparse.OPERATIONS.get(op_id)
        if op is None:
            raise ValueError(f"Unknown operation: {op_id}")
        params = timeparse.validate_params(op_id, params)
        src = self._source_lite(source_id)

        name = (name or "").strip()
        if not name:
            raise ValueError("The new column needs a name")
        if len(name) > 200:
            raise ValueError("Column name is too long")
        if name.lower() in RESERVED_COLUMN_NAMES:
            raise ValueError(f"{name!r} is a reserved column name")
        if self._find_column(src, name):
            raise ValueError(f"This table already has a column called {name!r}")

        # Inputs: a parse operation reads one of the file's own columns; a
        # two-input operation (duration) can also read an already-built
        # derived datetime column, so deltas can chain off parsed values.
        def _check_input(col: str, label: str) -> dict:
            entry = self._find_column(src, col)
            if entry is None:
                raise ValueError(f"No column called {col!r} to use as the {label}")
            if entry.get("derived"):
                if not op["two_input"]:
                    raise ValueError(f"{col!r} is itself a derived column — parse the original column instead")
                if entry.get("derived_status") != "ready":
                    raise ValueError(f"{col!r} is still building — wait for it to finish first")
            return entry

        _check_input(input_column, "input column")
        if op["two_input"]:
            _check_input(params["other_column"], "second column")

        drv = self._derived_table(source_id)
        with self.lock, self.db:
            self.db.execute(f"CREATE TABLE IF NOT EXISTS {q(drv)} (rid INTEGER PRIMARY KEY)")
            existing = {r[1].lower() for r in self.db.execute(f"PRAGMA table_info({q(drv)})")}
            if name.lower() in existing:
                # An orphan physical column left behind by a remove on a
                # SQLite too old for DROP COLUMN (< 3.35). Definitions drive
                # visibility, so it was invisible; reuse it, blanked.
                self.db.execute(f"UPDATE {q(drv)} SET {q(name)}=NULL")
            else:
                self.db.execute(f"ALTER TABLE {q(drv)} ADD COLUMN {q(name)} TEXT")
            cur = self.db.execute(
                "INSERT INTO derived_columns(source_id, name, input_column, op_id, params, status, created_at) "
                "VALUES (?,?,?,?,?,'building',?)",
                (source_id, name, input_column, op_id, json.dumps(params),
                 time.strftime("%Y-%m-%dT%H:%M:%S")),
            )
            def_id = cur.lastrowid
        job = self.start_ingest_job(
            "derive", "", name=name,
            options={"def_id": def_id, "drop_on_cancel": True, "units_total": src["row_count"]},
        )
        return {"definition": self.get_derived_column(def_id), "job_id": job["job_id"]}

    STRUCT_SAMPLE = 200

    def detect_struct_paths(self, source_id: int, column: str) -> dict:
        """What's inside a column that holds JSON or XML documents.

        Samples the column, decides which of the two it is (or neither),
        and enumerates every field found with how many of the sampled rows
        carried it plus one example value. That coverage number is the
        point: it's what turns "here are 60 paths" into "here are the six
        worth making columns of", and it's why this samples rather than
        scanning — the answer only has to be good enough to tick
        checkboxes against, and a full scan of a 1.2M-row column to
        populate a picker would cost more than building the columns."""
        src = self._source_lite(source_id)
        if not self._find_column(src, column):
            raise KeyError(f"No column called {column!r}")
        values = self._sample_column(src, column, limit=self.STRUCT_SAMPLE)
        kind = structparse.sniff_kind(values)
        if kind is None:
            return {"kind": None, "paths": [], "sampled": len(values)}
        return {
            "kind": kind,
            "sampled": len(values),
            "paths": structparse.discover_paths(values, kind),
        }

    def add_derived_columns(self, source_id: int, specs: list[dict]) -> dict:
        """Define several derived columns over one source and backfill them
        in a single pass.

        This exists for flattening: picking eight fields out of a JSON
        column is one intent, and running it as eight `add_derived_column`
        calls would mean eight full scans of the source table, eight jobs
        in the panel, and eight chances to end up half-done. One job, one
        scan, N columns.

        Definitions are created up front and all-or-nothing — a name
        collision in the fifth spec fails before the first column exists,
        rather than leaving four behind for the analyst to clean up."""
        if not specs:
            raise ValueError("Nothing to add")
        if source_id < 0:
            # Same fan-out contract as add_derived_column (invariant #9):
            # validate every spec against every member first, then create
            # the whole batch on each member — a collision anywhere fails
            # before anything exists anywhere.
            members = self._resolve_members(source_id)
            for m in members:
                msrc = self._source_lite(m["source_id"])
                for spec in specs:
                    nm = (spec.get("name") or "").strip()
                    if self._find_column(msrc, nm):
                        raise ValueError(
                            f"Member table {msrc['name']!r} already has a column called {nm!r}")
                    if self._find_column(msrc, spec.get("input_column")) is None:
                        raise ValueError(
                            f"Member table {msrc['name']!r} has no column {spec.get('input_column')!r}")
            results = [self.add_derived_columns(m["source_id"], specs) for m in members]
            first = results[0]
            return {**first,
                    "member_results": results,
                    "job_ids": [j for r in results for j in (r.get("job_ids") or [r.get("job_id")]) if j]}
        src = self._source_lite(source_id)
        drv = self._derived_table(source_id)

        # Validated against each other as well as against the table: two
        # specs asking for the same name is the realistic mistake here
        # (two paths whose last component is "Name"), and _find_column
        # can't see a column that doesn't exist yet.
        taken = {c["name"].lower() for c in src["columns"]}
        prepared = []
        for spec in specs:
            name = (spec.get("name") or "").strip()
            op_id = spec.get("op_id")
            input_column = spec.get("input_column")
            op = timeparse.OPERATIONS.get(op_id)
            if op is None:
                raise ValueError(f"Unknown operation: {op_id}")
            if op["two_input"]:
                raise ValueError(f"{op['label']!r} takes two columns and can't be added in a batch")
            if not name:
                raise ValueError("Every new column needs a name")
            if len(name) > 200:
                raise ValueError(f"Column name {name!r} is too long")
            if name.lower() in RESERVED_COLUMN_NAMES:
                raise ValueError(f"{name!r} is a reserved column name")
            if name.lower() in taken:
                raise ValueError(f"This table already has a column called {name!r}")
            entry = self._find_column(src, input_column)
            if entry is None:
                raise ValueError(f"No column called {input_column!r} to read")
            if entry.get("derived"):
                raise ValueError(f"{input_column!r} is itself a derived column — read the original instead")
            taken.add(name.lower())
            prepared.append({
                "name": name, "input_column": input_column, "op_id": op_id,
                "params": timeparse.validate_params(op_id, spec.get("params")),
            })

        def_ids = []
        with self.lock, self.db:
            self.db.execute(f"CREATE TABLE IF NOT EXISTS {q(drv)} (rid INTEGER PRIMARY KEY)")
            existing = {r[1].lower() for r in self.db.execute(f"PRAGMA table_info({q(drv)})")}
            for pspec in prepared:
                name = pspec["name"]
                if name.lower() in existing:
                    # Orphan physical column from a remove on SQLite < 3.35
                    # (no DROP COLUMN) — same reuse-it-blanked path
                    # add_derived_column takes.
                    self.db.execute(f"UPDATE {q(drv)} SET {q(name)}=NULL")
                else:
                    self.db.execute(f"ALTER TABLE {q(drv)} ADD COLUMN {q(name)} TEXT")
                cur = self.db.execute(
                    "INSERT INTO derived_columns(source_id, name, input_column, op_id, params, status, created_at) "
                    "VALUES (?,?,?,?,?,'building',?)",
                    (source_id, name, pspec["input_column"], pspec["op_id"],
                     json.dumps(pspec["params"]), time.strftime("%Y-%m-%dT%H:%M:%S")),
                )
                def_ids.append(cur.lastrowid)

        label = prepared[0]["name"] if len(prepared) == 1 else f"{len(prepared)} columns"
        job = self.start_ingest_job(
            "derive", "", name=label,
            options={"def_ids": def_ids, "drop_on_cancel": True, "units_total": src["row_count"]},
        )
        return {
            "definitions": [self.get_derived_column(d) for d in def_ids],
            "job_id": job["job_id"],
        }

    def remove_derived_column(self, def_id: int) -> None:
        d = self.get_derived_column(def_id)
        with self.lock:
            others = self.db.execute(
                "SELECT * FROM derived_columns WHERE source_id=? AND id<>?",
                (d["source_id"], def_id),
            ).fetchall()
        for o in others:
            o_params = json.loads(o["params"] or "{}")
            if (o["input_column"].lower() == d["name"].lower()
                    or str(o_params.get("other_column", "")).lower() == d["name"].lower()):
                raise ValueError(f"{o['name']!r} is computed from this column — remove that one first")
        drv = self._derived_table(d["source_id"])
        with self.lock, self.db:
            self.db.execute("DELETE FROM derived_columns WHERE id=?", (def_id,))
            try:
                self.db.execute(f"ALTER TABLE {q(drv)} DROP COLUMN {q(d['name'])}")
            except sqlite3.OperationalError:
                # SQLite < 3.35 has no DROP COLUMN. The definition is gone,
                # so the column is invisible everywhere; add_derived_column
                # reuses the orphan if the same name comes back.
                pass

    def rederive_column(self, def_id: int, params: dict | None = None) -> dict:
        """Recompute a derived column in place — the path for "I set the
        wrong syslog year" or "these were local time, not UTC"."""
        d = self.get_derived_column(def_id)
        new_params = timeparse.validate_params(d["op_id"], params if params is not None else d["params"])
        with self.lock, self.db:
            self.db.execute(
                "UPDATE derived_columns SET params=?, status='building', parse_failures=NULL WHERE id=?",
                (json.dumps(new_params), def_id),
            )
        src = self._source_lite(d["source_id"])
        job = self.start_ingest_job(
            "derive", "", name=d["name"],
            options={"def_id": def_id, "drop_on_cancel": False, "units_total": src["row_count"]},
        )
        return {"definition": self.get_derived_column(def_id), "job_id": job["job_id"]}

    def backfill_derived_column(self, def_id: int, progress=None, cancel=None,
                                drop_on_cancel: bool = False) -> dict:
        """Compute every row's value for one derived column. The one-column
        case of backfill_derived_columns, kept as its own name because
        that's what the single-column create/re-derive paths mean."""
        out = self.backfill_derived_columns(
            [def_id], progress=progress, cancel=cancel, drop_on_cancel=drop_on_cancel,
        )
        return out["columns"][0] | {"rows": out["rows"]}

    def backfill_derived_columns(self, def_ids: list[int], progress=None, cancel=None,
                                 drop_on_cancel: bool = False) -> dict:
        """Compute every row's values for one or more derived columns of the
        same source, in a single pass.

        Batched exactly like ingest_csv/build_fts: read a BATCH-sized window
        on a pooled reader (committed data — the source table hasn't changed
        since ingest, invariant #1), parse in Python, then take the writer
        lock for just that batch's upsert (invariant #4). Rows are walked in
        rid order because the stateful operations need it — BSD syslog's
        year rollover is decided by comparing each line's month against the
        previous line's, so the order rows are parsed in is part of the
        answer, not an implementation detail. Each definition gets its own
        `state` dict for exactly that reason: two columns derived in the
        same pass must not share a rollover cursor.

        Several columns at once is what flattening a JSON blob into its
        fields needs (see add_derived_columns) — one scan and one progress
        bar instead of N of each. The SELECT reads each distinct input
        column once however many definitions want it."""
        if not def_ids:
            raise ValueError("Nothing to backfill")
        defs = [self.get_derived_column(d) for d in def_ids]
        source_id = defs[0]["source_id"]
        if any(d["source_id"] != source_id for d in defs):
            raise ValueError("All columns in one backfill must belong to the same source")
        src = self._source_lite(source_id)
        drv = self._derived_table(source_id)
        frm = self._from_clause(src)

        # One SELECT slot per distinct input column, shared by every
        # definition reading it.
        slots: dict[str, int] = {}

        def slot(col: str) -> int:
            key = col.lower()
            if key not in slots:
                slots[key] = len(slots)
            return slots[key]

        plans = []
        for d in defs:
            op = timeparse.OPERATIONS[d["op_id"]]
            plan = {"d": d, "op": op, "state": {}, "a": slot(d["input_column"])}
            if op["two_input"]:
                plan["b"] = slot(d["params"]["other_column"])
            plans.append(plan)
        cols = list(slots.keys())
        # slot() keys case-insensitively but _col_ref needs the real name.
        real = {}
        for d in defs:
            real[d["input_column"].lower()] = d["input_column"]
            if timeparse.OPERATIONS[d["op_id"]]["two_input"]:
                other = d["params"]["other_column"]
                real[other.lower()] = other
        sel = ", ".join(self._col_ref(src, real[c]) for c in cols)

        names = [d["name"] for d in defs]
        assigns = ", ".join(f"{q(n)}=excluded.{q(n)}" for n in names)
        upsert = (f"INSERT INTO {q(drv)}(rid, {', '.join(q(n) for n in names)}) "
                  f"VALUES({','.join('?' * (len(names) + 1))}) "
                  f"ON CONFLICT(rid) DO UPDATE SET {assigns}")

        rows_done, last_rid = 0, 0
        total = src["row_count"]
        try:
            while True:
                with self._reader() as ro:
                    rows = ro.execute(
                        f"SELECT rid, {sel} FROM {frm} WHERE rid > ? ORDER BY rid LIMIT {BATCH}",
                        (last_rid,),
                    ).fetchall()
                if not rows:
                    break
                vals = []
                for r in rows:
                    t = tuple(r)
                    out = [t[0]]
                    for plan in plans:
                        op = plan["op"]
                        if op["two_input"]:
                            out.append(op["parse_pair"](t[1 + plan["a"]], t[1 + plan["b"]], plan["d"]["params"]))
                        else:
                            out.append(op["parse"](t[1 + plan["a"]], plan["d"]["params"], plan["state"]))
                    vals.append(tuple(out))
                last_rid = vals[-1][0]
                with self.lock, self.db:
                    if cancel and cancel():
                        raise IngestCancelled("cancelled during derive")
                    self.db.executemany(upsert, vals)
                rows_done += len(vals)
                if progress:
                    progress(rows_done, rows_done, total)
                time.sleep(0.02)  # same anti-starvation yield build_fts uses
        except IngestCancelled:
            if drop_on_cancel:
                # Mirror of cancel-drops-the-partial-source: the analyst
                # asked for these columns not to exist, and a half-filled one
                # looks exactly like a complete one in the grid.
                for d in defs:
                    self.remove_derived_column(d["id"])
            else:
                with self.lock, self.db:
                    self.db.executemany(
                        "UPDATE derived_columns SET status='partial' WHERE id=?",
                        [(d["id"],) for d in defs],
                    )
            raise

        out_cols = []
        for d in defs:
            failures = self._count_parse_failures(src, d)
            with self.lock, self.db:
                self.db.execute(
                    "UPDATE derived_columns SET status='ready', parse_failures=? WHERE id=?",
                    (failures, d["id"]),
                )
            out_cols.append({"def_id": d["id"], "source_id": source_id, "name": d["name"],
                             "parse_failures": failures})
        return {"source_id": source_id, "rows": rows_done, "columns": out_cols}

    def _count_parse_failures(self, src: dict, d: dict) -> int:
        """Rows whose input had something in it but produced no datetime —
        the number the analyst needs to decide whether they picked the wrong
        format. An empty input cell isn't a failure, it's an empty cell."""
        drv = self._derived_table(d["source_id"])
        inp = self._col_ref(src, d["input_column"], "s")
        with self._reader() as ro:
            return ro.execute(
                f"SELECT COUNT(*) FROM {self._from_clause(src, 's')} "
                f"WHERE {inp} IS NOT NULL AND {inp} <> '' AND {q(drv)}.{q(d['name'])} IS NULL"
            ).fetchone()[0]

    def unparsed_where_fragment(self, def_id: int) -> str:
        """An advanced-filter fragment selecting exactly the rows that
        failed to parse, so "12 failures" is one click from "show me
        which 12". Built here rather than in the frontend because it has
        to quote two analyst-named columns into SQL (invariant #5)."""
        d = self.get_derived_column(def_id)
        return f"{q(d['name'])} IS NULL AND {q(d['input_column'])} <> ''"

    DETECT_SAMPLE = 200

    def _sample_column(self, src: dict, column: str, limit: int | None = None) -> list:
        with self._reader() as ro:
            return [r[0] for r in ro.execute(
                f"SELECT {self._col_ref(src, column)} FROM {self._from_clause(src)} "
                f"WHERE {self._col_ref(src, column)} IS NOT NULL "
                f"AND {self._col_ref(src, column)} <> '' LIMIT ?",
                (limit or self.DETECT_SAMPLE,),
            )]

    def detect_timestamp_format(self, source_id: int, column: str) -> list[dict]:
        """Which operations can read this column, best first. Same
        sample-and-round-trip shape preview_sqlite_tables uses to pre-check
        likely WebKit columns: try every parser against real values and
        rank by how many it reads. Ambiguous numeric ranges (Mac absolute
        vs unix seconds) legitimately return several — the caller's preview
        of actual converted values is what settles it."""
        if source_id < 0:  # a merge samples its first (canonical) member
            source_id = self._resolve_members(source_id)[0]["source_id"]
        src = self._source_lite(source_id)
        if self._find_column(src, column) is None:
            raise KeyError(column)
        samples = self._sample_column(src, column)
        out = []
        for r in timeparse.detect(samples):
            out.append({**r, "preview": self._preview_rows(samples[:10], r["op_id"], r["params"])})
        return out

    def detect_source_suggestions(self, source_id: int, min_confidence: float = 0.9) -> list[dict]:
        """Columns that look like timestamps the app can't already read —
        the basis for the post-import "this looks like a Unix epoch, want a
        datetime column?" hint. Columns already typed `datetime` at ingest
        are skipped: those already sort and filter correctly, so converting
        them would just duplicate data."""
        src = self._source_lite(source_id)
        # Nothing to suggest for a column the analyst has already converted.
        converted = {d["input_column"].lower() for d in self.list_derived_columns(source_id)}
        out = []
        for c in self._base_cols(src):
            if c["type"] == "datetime" or c["name"].lower() in converted:
                continue
            samples = self._sample_column(src, c["name"])
            if not samples:
                continue
            ranked = timeparse.detect(samples)
            if ranked and ranked[0]["confidence"] >= min_confidence:
                best = ranked[0]
                out.append({
                    "column": c["name"], "op_id": best["op_id"], "label": best["label"],
                    "confidence": best["confidence"], "params": best["params"],
                    "preview": self._preview_rows(samples[:5], best["op_id"], best["params"]),
                })
        return out

    def _preview_rows(self, values: list, op_id: str, params: dict) -> list[dict]:
        op = timeparse.OPERATIONS[op_id]
        state: dict = {}
        return [{"input": v, "output": op["parse"](v, params, state)} for v in values]

    def preview_derived(self, source_id: int, column: str, op_id: str,
                        params: dict | None = None, limit: int = 10) -> dict:
        """A handful of real input→output pairs plus how many of the
        sampled values this operation can't read, so the analyst sees what
        they're about to create before committing to a full backfill."""
        op = timeparse.OPERATIONS.get(op_id)
        if op is None:
            raise ValueError(f"Unknown operation: {op_id}")
        params = timeparse.validate_params(op_id, params)
        if source_id < 0:
            # Preview against the first member — representative, and the
            # only side with a real table to sample.
            source_id = self._resolve_members(source_id)[0]["source_id"]
        src = self._source_lite(source_id)
        if self._find_column(src, column) is None:
            raise KeyError(column)
        if op["two_input"]:
            other = params["other_column"]
            if self._find_column(src, other) is None:
                raise KeyError(other)
            with self._reader() as ro:
                rows = ro.execute(
                    f"SELECT {self._col_ref(src, column)}, {self._col_ref(src, other)} "
                    f"FROM {self._from_clause(src)} LIMIT ?", (limit,),
                ).fetchall()
            preview = [{"input": r[0], "output": op["parse_pair"](r[0], r[1], params)} for r in rows]
            failures = sum(1 for p in preview if p["output"] is None)
            return {"preview": preview, "sampled": len(preview), "failures": failures}
        samples = self._sample_column(src, column)
        preview = self._preview_rows(samples[:limit], op_id, params)
        state: dict = {}
        failures = sum(1 for v in samples if op["parse"](v, params, state) is None)
        return {"preview": preview, "sampled": len(samples), "failures": failures}

    # ------------------------------------------------------------ case settings

    def get_case_settings(self) -> dict:
        with self.lock:
            return {r["key"]: r["value"] for r in self.db.execute("SELECT key, value FROM case_settings")}

    def set_case_setting(self, key: str, value: str | None) -> None:
        with self.lock, self.db:
            if value is None or value == "":
                self.db.execute("DELETE FROM case_settings WHERE key=?", (key,))
            else:
                self.db.execute(
                    "INSERT INTO case_settings(key, value) VALUES (?,?) "
                    "ON CONFLICT(key) DO UPDATE SET value=excluded.value",
                    (key, str(value)),
                )

    # ------------------------------------------------------------------- views

    def build_view(self, source_id: int, spec: dict) -> dict:
        """Materialise (pos, source_id, rid) for a filter+sort spec. Returns
        view handle. source_id is a constant column here (every row comes
        from the same source) — kept as a real column rather than a
        handle-level constant so fetch_rows/tag_positions/tag_view/export
        share one code path with merged views, whose rows carry their own
        (differing) source_id per row.

        CLAUDE.md invariant #2's carve-out: an unfiltered, unsorted view of
        a single (non-merge) source needs no materialisation at all — its
        "sort" is just rid order, which INTEGER PRIMARY KEY already gives
        for free (verified: `EXPLAIN QUERY PLAN SELECT rid FROM t ORDER BY
        rid ASC` is a bare `SCAN`, no temp b-tree). See
        _build_virtual_root_view."""
        src = self._source_lite(source_id)
        colnames = {c["name"]: c["type"] for c in src["columns"]}
        order = self._compile_order(spec, colnames)
        # Only non-numeric sort columns: _compile_order emits a plain
        # `col COLLATE NOCASE` for those, which a background index can
        # serve; a numeric sort goes through _numeric_expr, a functional
        # expression a plain index can't match (see
        # _ensure_sort_index_building).
        sort_cols = [
            s.get("column") for s in (spec.get("sort") or [])
            if s.get("column") in colnames and colnames[s.get("column")] != "number"
        ]
        # __line__ asc IS the virtual root's order, so it deliberately does
        # NOT count as a sort (keeps the no-materialize fast path); only the
        # descending direction needs a real ORDER BY.
        has_sort = any(
            s.get("column") in colnames
            or (s.get("column") == "__line__" and str(s.get("dir", "asc")).lower() == "desc")
            for s in (spec.get("sort") or [])
        )

        if src.get("is_merge"):
            collist = ", ".join(q(c) for c in colnames)
            branches = []
            params: list[Any] = []
            for m in self._resolve_members(source_id):
                msrc = self._source_lite(m["source_id"])
                where, p = self._compile_where(m["source_id"], msrc, spec, colnames)
                for col in sort_cols:
                    self._ensure_sort_index_building(
                        m["source_id"], col,
                        self._index_table_for(src, col, msrc["table_name"], m["source_id"]))
                branches.append(
                    f"SELECT {int(m['source_id'])} AS source_id, rid, {collist} FROM {self._member_from(src, m)}"
                    + (f" WHERE {where}" if where else "")
                )
                params.extend(p)
            union_sql = " UNION ALL ".join(branches)
            self._view_seq += 1
            vid = f"view_{self._view_seq}"
            sql = (
                f"INSERT INTO v.{q(vid)}(source_id, rid) "
                f"SELECT source_id, rid FROM ({union_sql}) {order}"
            )
        else:
            table = src["table_name"]
            where, params = self._compile_where(source_id, src, spec, colnames)
            derived_names = {c["name"] for c in src["columns"] if c.get("derived")}
            for col in sort_cols:
                # A derived column's values live in the sidecar, so its sort
                # index belongs on that table, not on src_<id>.
                self._ensure_sort_index_building(
                    source_id, col,
                    self._derived_table(source_id) if col in derived_names else table,
                )
            if not where and not has_sort:
                return self._build_virtual_root_view(source_id, src)
            self._view_seq += 1
            vid = f"view_{self._view_seq}"
            sql = (
                f"INSERT INTO v.{q(vid)}(source_id, rid) "
                f"SELECT {int(source_id)}, rid FROM {self._from_clause(src)}"
                + (f" WHERE {where}" if where else "")
                + f" {order}"
            )
        t0 = time.time()
        with self.lock, self._interruptible(spec.get("op_token"), self.db), self.db:
            # pos declared INTEGER PRIMARY KEY makes it *be* the rowid, so
            # this alone gives the same unique/indexed-by-pos lookup the old
            # CTAS + separate CREATE UNIQUE INDEX did, for one less full
            # sort-and-write of every row (see CLAUDE.md's "Performance"
            # section, change 3). pos is deliberately absent from the
            # INSERT's column list: rowids are assigned 1..N in insertion
            # order, and an INSERT..SELECT with ORDER BY inserts in sorted
            # order — measured ~35% faster than numbering the same sort
            # with ROW_NUMBER() OVER, for byte-identical view content.
            # (The ORDER BY always ends in `rid ASC`, so the order — and
            # therefore every pos — is fully determined, not sorter-
            # dependent.) cursor.rowcount then gives the view's row count
            # without re-scanning the table we just wrote.
            self.db.execute(f"CREATE TABLE v.{q(vid)} (pos INTEGER PRIMARY KEY, source_id INTEGER, rid INTEGER)")
            n = self.db.execute(sql, params).rowcount
            # Evicted *after* the new view is built, not before (it used to
            # be first): the whole block is one transaction, so a cancelled
            # (interrupted) build rolls back to a world where the previous
            # view still exists — the frontend keeps its rows instead of
            # every open handle 409ing. The new view's handle isn't in
            # self._views yet, so the eviction loop can't touch it.
            self._evict_root_views(source_id)

        handle = {
            "view_id": vid,
            "source_id": source_id,
            "row_count": n,
            "kind": "root",
            "elapsed_ms": int((time.time() - t0) * 1000),
        }
        self._views[vid] = handle
        return handle

    def _evict_root_views(self, source_id: int) -> None:
        """Evicts every existing root (materialised or virtual) view for
        this source, so opening a new one leaves only the new view live.
        Caller must hold self.lock (and be inside a self.db transaction —
        eviction of a materialised view drops its backing table).

        Only root views for this source are eligible here. Group sub-views
        (kind='group'/'group_virtual') are never evicted by this loop
        directly — they cascade-drop with their parent via
        _evict_view_and_children, so expanding one group never evicts a
        sibling or the outer root view. `old` may already be gone by the
        time we reach it (a previous iteration's cascade can remove a later
        entry in this same snapshot) — .get() and skip, don't [ ]."""
        for old in list(self._views):
            handle = self._views.get(old)
            if handle and handle.get("kind") in ("root", "root_virtual") and handle["source_id"] == source_id:
                self._evict_view_and_children(old)

    def _build_virtual_root_view(self, source_id: int, src: dict) -> dict:
        """No filters, no sort: the view IS the source table in its natural
        rid order, which INTEGER PRIMARY KEY already gives for free — no
        v.view_N to build. row_count is read straight off `src` (already
        the source's own row_count) rather than a fresh COUNT(*).

        fetch_rows/tag_positions/find_position/tag_view/export all handle
        kind='root_virtual' by paging or computing directly against the
        source table instead of a backing view table — every one of them
        can use the exact `pos = rid - 1` shortcut (never a stub) because
        ingest assigns rid contiguously from 1 for every ingest path
        (verified: CSV/SQLite-table/JSON ingest all create `rid INTEGER
        PRIMARY KEY` without ever specifying a value, insert in sequential
        BATCH-sized transactions that either fully commit or fully roll
        back, and no source table is ever mutated after ingest —
        invariant #1) and this view is, by construction, every row of the
        source.

        Deliberately narrower than "sort is absent OR served by an
        existing index" — a virtual view with a sorted (even indexed)
        order is not implemented: tag_positions would need each tagged
        row's rank in that order, which isn't a `pos = rid - 1` shortcut
        and would cost one query per tagged row (up to 20,000) rather than
        one materialise. Change 2 already makes that case's materialise
        cheap (index supplies the order, no temp b-tree); this virtual
        path targets the single biggest and most common case instead:
        opening a table."""
        with self.lock, self.db:
            self._evict_root_views(source_id)
            self._view_seq += 1
            vid = f"view_{self._view_seq}"
            handle = {
                "view_id": vid,
                "source_id": source_id,
                "row_count": src["row_count"],
                "kind": "root_virtual",
                "elapsed_ms": 0,
            }
            self._views[vid] = handle
        return handle

    def _evict_view_and_children(self, vid: str) -> None:
        """Drops a view (and its backing table, unless it's a virtual
        no-table handle) plus every group sub-view whose parent_view_id is
        this one — recursively, so a chain of expansions cleans up fully.
        Caller must hold self.lock (and be inside a self.db transaction if
        dropping tables)."""
        handle = self._views.pop(vid, None)
        if not handle:
            return
        if handle.get("kind") not in ("group_virtual", "root_virtual"):
            self.db.execute(f"DROP TABLE IF EXISTS v.{q(vid)}")
        for child_vid, child in list(self._views.items()):
            if child.get("parent_view_id") == vid:
                self._evict_view_and_children(child_vid)

    def close_view(self, view_id: str) -> None:
        """Explicit free for a group sub-view (called on UI collapse) rather
        than waiting for its parent root view to be rebuilt."""
        with self.lock, self.db:
            self._evict_view_and_children(view_id)

    # -------------------------------------------------------------- timeline

    def build_timeline(self, configs: dict[int, dict] | None = None, tag_ids: list[int] | None = None,
                       op_token: str | None = None) -> dict:
        """Materialises one row per *tagged* row across every real source in
        the case (open or closed — this is "every finding in the case", not
        "every finding in a currently-open tab"), each contributing its own
        timestamp/body/type-label per `configs.get(source_id)`. `configs` is
        resolved by the caller (server.py's _resolve_timeline_configs,
        matching each source's header set against workspace.timeline_
        templates) rather than looked up in here — store.py can't import
        workspace.py (workspace.py already imports from store.py; see
        pop_legacy_presets for the same constraint) — a source missing from
        `configs`, or with an override that doesn't actually apply to it
        (a timestamp_column that isn't one of its own columns), falls back
        to its first datetime column, every column, and its own name.

        `tag_ids` narrows to rows carrying at least one of those tags;
        None/empty means every tagged row regardless of which tag(s).

        Same materialize-into-v.-then-page-by-pos pattern as build_view
        (CLAUDE.md invariant #2) — a heavily-tagged case can still be
        thousands of rows, so this isn't exempt from the no-LIMIT/OFFSET
        rule just because it's a cross-table view rather than a single-
        source one. Only one timeline view is ever alive at a time (an
        analyst has one Timeline tab, not several) — building a new one
        evicts whatever timeline view existed before, same as build_view
        evicts the previous root view for a source."""
        configs = configs or {}
        tag_clause = f" AND tag_id IN ({','.join('?' * len(tag_ids))})" if tag_ids else ""

        branches = []
        params: list[Any] = []
        for src in self.list_sources():
            if not src["tagged_row_count"]:
                continue
            source_id = src["id"]
            col_names = {c["name"] for c in src["columns"]}
            dt_cols = [c["name"] for c in src["columns"] if c["type"] == "datetime"]
            cfg = configs.get(source_id, {})

            ts_col = cfg.get("timestamp_column")
            if ts_col not in col_names:
                ts_col = dt_cols[0] if dt_cols else None
            # A derived column is already canonical "YYYY-MM-DD HH:MM:SS",
            # so it's used as-is: TS_NORMALIZE would only truncate the
            # sub-second part that makes two same-second events sort in the
            # order they happened. Everything else still normalises, which
            # is what lets sources in different formats interleave.
            ts_expr = "NULL"
            if ts_col:
                ref = self._col_ref(src, ts_col, "s")
                ts_expr = ref if self._is_derived(src, ts_col) else f"TS_NORMALIZE({ref})"

            body_cols = [c for c in (cfg.get("body_columns") or []) if c in col_names]
            if not body_cols:
                body_cols = [c["name"] for c in src["columns"]]
            body_expr = " || ' | ' || ".join(
                f"COALESCE({self._col_ref(src, c, 's')}, '')" for c in body_cols
            )

            type_label = cfg.get("type_label") or src["name"]
            branches.append(
                f"SELECT {int(source_id)} AS source_id, s.rid AS rid, {ts_expr} AS ts, ({body_expr}) AS body, "
                f"? AS type_label, ? AS source_name, "
                f"(SELECT GROUP_CONCAT(tag_id) FROM row_tags WHERE source_id={int(source_id)} AND rid=s.rid) AS tag_ids "
                f"FROM {self._from_clause(src, 's')} "
                f"WHERE s.rid IN (SELECT rid FROM row_tags WHERE source_id={int(source_id)}{tag_clause})"
            )
            params.append(type_label)
            params.append(src["name"])
            if tag_ids:
                params.extend(tag_ids)

        self._view_seq += 1
        vid = f"view_{self._view_seq}"
        with self.lock, self._interruptible(op_token, self.db), self.db:
            self.db.execute(
                f"CREATE TABLE v.{q(vid)} (pos INTEGER PRIMARY KEY, source_id INTEGER, rid INTEGER, "
                f"ts TEXT, body TEXT, type_label TEXT, source_name TEXT, tag_ids TEXT)"
            )
            n = 0
            if branches:
                union_sql = " UNION ALL ".join(branches)
                # ORDER BY-fed insert, same as build_view: pos auto-assigns
                # in sorted order ((source_id, rid) makes the order fully
                # determined), rowcount replaces a count(*) re-scan.
                n = self.db.execute(
                    f"INSERT INTO v.{q(vid)}(source_id, rid, ts, body, type_label, source_name, tag_ids) "
                    f"SELECT source_id, rid, ts, body, type_label, source_name, tag_ids FROM ({union_sql}) "
                    f"ORDER BY (ts IS NULL) ASC, ts ASC, source_id, rid",
                    params,
                ).rowcount
            # Same evict-after-build ordering (and reasoning) as build_view.
            for old in list(self._views):
                h = self._views.get(old)
                if h and h.get("kind") == "timeline":
                    self._evict_view_and_children(old)

        self._views[vid] = {"view_id": vid, "kind": "timeline", "row_count": n}
        return {"view_id": vid, "row_count": n}

    def fetch_timeline_rows(self, view_id: str, start: int, count: int) -> dict:
        handle = self._views.get(view_id)
        if not handle or handle.get("kind") != "timeline":
            raise KeyError("View expired — rebuild it")
        with self._reader() as ro, self._dropped_view_is_expired():
            rows = ro.execute(
                f"SELECT pos, source_id, rid, ts, body, type_label, source_name, tag_ids "
                f"FROM v.{q(view_id)} WHERE pos >= ? AND pos < ? ORDER BY pos",
                (start + 1, start + 1 + count),
            ).fetchall()
        out = [
            {
                "pos": r["pos"] - 1,
                "source_id": r["source_id"],
                "rid": r["rid"],
                "ts": r["ts"],
                "body": r["body"],
                "type_label": r["type_label"],
                "source_name": r["source_name"],
                "tags": [int(x) for x in r["tag_ids"].split(",")] if r["tag_ids"] else [],
            }
            for r in rows
        ]
        return {"start": start, "rows": out}

    # ---------------------------------------------------------------- group-by

    GROUP_MATERIALIZE_THRESHOLD = 5000  # rows; above this (or for a merge) a group gets a real indexed sub-view

    @staticmethod
    def _eq_condition(col_ident: str, value: Any, is_datetime: bool = False) -> tuple[str, list]:
        """`is_datetime` must match how group_summary grouped this column
        (DAY_BUCKET'd or not) — otherwise a group's day-bucket value like
        "2026-01-01" would get compared against the raw full-timestamp
        column and never match any row, breaking expand/tag/export for
        every datetime group."""
        expr = f"DAY_BUCKET({col_ident})" if is_datetime else col_ident
        if value is None:
            return f"{expr} IS ?", [None]
        return f"{expr} = ?", [value]

    @staticmethod
    def _tag_condition(alias: str, source_id: int, value: Any) -> tuple[str, list]:
        """"Rows carrying tag `value`", or "rows carrying no tag at all"
        when `value` is None — the tag pseudo-column's answer to
        _eq_condition, and the reason grouping by tag needs its own
        condition builder rather than a clever column expression: the thing
        being compared lives in the row_tags sidecar, keyed by
        (source_id, rid), not in the source table at all.

        `value` is a tag *id*, never a name: tag_defs has no unique
        constraint on name, so grouping by name would silently merge two
        tags an analyst deliberately kept apart. The frontend renders the
        name it already has in S.tags."""
        prefix = f"{alias}." if alias else ""
        if value is None:
            return f"{prefix}rid NOT IN (SELECT rid FROM row_tags WHERE source_id = ?)", [source_id]
        return (
            f"{prefix}rid IN (SELECT rid FROM row_tags WHERE source_id = ? AND tag_id = ?)",
            [source_id, int(value)],
        )

    def _path_where(self, path: list[dict] | None, colnames: dict, alias: str = "",
                    src: dict | None = None, derived_alias: str | None = None,
                    member_sid: int | None = None) -> tuple[list[str], list]:
        """Compile a group-by path (the outer levels' already-fixed values,
        for nested multi-column grouping) into per-column condition
        fragments. `alias` (e.g. "s"), when given, is baked into the
        identifier _eq_condition builds on — NOT prefixed onto the
        returned fragment by the caller afterward, since a datetime
        column's fragment starts with DAY_BUCKET(...) and "s.DAY_BUCKET(...)"
        isn't valid SQL the way "s.\"col\"" is.

        `src` routes each column through _col_ref, so a level grouped on a
        derived column resolves against the sidecar rather than taking the
        source table's alias (which would be "no such column")."""
        prefix = f"{alias}." if alias else ""
        clauses: list[str] = []
        params: list[Any] = []
        for p in path or []:
            col = p.get("column")
            if col == TAG_GROUP_COLUMN:
                # A tag level's predicate is per *member* (row_tags is keyed
                # by source_id), which is why member_sid is threaded down
                # here rather than the whole path being compiled once and
                # reused across a merge's members.
                if member_sid is None:
                    raise ValueError("grouping by tag needs a member source id")
                c, v = self._tag_condition(alias, member_sid, p.get("value"))
                clauses.append(c)
                params.extend(v)
                continue
            if col not in colnames:
                raise KeyError(col)
            ident = (self._col_ref(src, col, alias or None, derived_alias) if src is not None
                     else f"{prefix}{q(col)}")
            c, v = self._eq_condition(ident, p.get("value"), colnames[col] == "datetime")
            clauses.append(c)
            params.extend(v)
        return clauses, params

    @staticmethod
    def _grouping_covers_whole_source(handle: dict, src: dict) -> bool:
        """True when this view provably contains every row of every member
        table, so an aggregate over the view and one over the tables
        themselves are the same question.

        Row counts are enough to prove it, not just a heuristic: a view's
        rows are always distinct (source_id, rid) pairs drawn from its
        members (build_view materialises one row per matching source row),
        so a view holding as many rows as the source holds can only be all
        of them. A filter that happens to match every row is
        indistinguishable from no filter here — and that's fine, because the
        answer is genuinely identical either way. `src` covers a merge
        without special-casing: _merge_source_dict's row_count is already
        the sum of its members'.

        Takes the source dict the caller already has rather than looking it
        up: get_source runs a COUNT(DISTINCT rid) over row_tags, which on a
        heavily tagged source is ~12ms — enough that adding a lookup here
        would have eaten the entire saving this check exists to unlock.

        Restricted to root views: a 'group'/'group_virtual' sub-view is a
        subset by construction, and its row_count comparison would be
        meaningless anyway. A root_virtual view is always this case by
        construction (build_view only goes virtual when unfiltered — see
        _build_virtual_root_view) without needing the row_count comparison,
        which matters because it has no v.view_N to have been counted from
        in the first place."""
        if handle.get("kind") == "root_virtual":
            return True
        if handle.get("kind", "root") != "root":
            return False
        return bool(src["row_count"]) and handle["row_count"] == src["row_count"]

    def group_summary(self, view_id: str, column: str, order: str = "count", direction: str | None = None,
                       limit: int = 1000, path: list[dict] | None = None,
                       op_token: str | None = None, bucket_datetime: bool = True) -> dict:
        """One aggregate pass over the already-filtered view — SELECT val,
        count(*) per member, unioned and re-summed. Not a paging operation,
        so no O(window) concern here; capped at `limit` distinct groups.

        `path` narrows to one nested group's rows before aggregating the
        next column — [{"column": ..., "value": ...}, ...] for every outer
        level already chosen, e.g. computing the User breakdown *within*
        the Process="svchost.exe" group of a Process-then-User grouping.

        `column` may be TAG_GROUP_COLUMN, which groups by the tags on the
        rows rather than by anything in the file: one group per tag (its
        `value` is the tag *id* — tag names aren't unique) plus one for the
        untagged remainder (`value` None). It's the one grouping whose
        counts can sum to more than the view holds, because a row carrying
        two tags belongs to both groups; that's what makes it answer "how
        much of this have I marked, and as what", which a partition into
        tag-set combinations would not. It nests in either direction like
        any other level. See _tag_group_branches for the query shapes and
        why their join order is pinned.

        A 'datetime' column groups by calendar day (DAY_BUCKET) rather than
        exact timestamp — grouping by the raw value would put nearly every
        row in a group of its own at second/millisecond precision, which
        isn't useful for anything. `expand_group`/`_virtual_group_where`
        match this same DAY_BUCKET'ing (via _eq_condition's is_datetime
        flag) so a click on a day group's row actually finds its rows.

        `bucket_datetime=False` turns that off, and exists for exactly one
        caller: the header value-picker dropdown, which lists a column's
        distinct values so the analyst can tick the ones to filter to. What
        it writes is an ordinary `=`/`in` filter on the *stored* value, so
        every value it shows has to be one such a filter can match — a
        DAY_BUCKET'ed one never is. Nothing about the grouping contract
        above is relaxed: this path returns values, never groups anything
        gets expanded against, so there is no `_eq_condition` on the other
        side to keep in step. Raw datetime values also make the column
        index worth building (a DAY_BUCKET is a functional expression a
        plain index can't serve), so the whole_source gate below admits
        them.

        `direction` defaults per `order` if not given explicitly (matches
        the behavior before either was ever configurable): "desc" for
        count (most-common-first), "asc" for value (alphabetical / chrono
        ascending) — both are swappable independent of that default.

        An unfiltered root view skips the join and aggregates the member
        tables directly — see _grouping_covers_whole_source. That's not a
        micro-optimisation: the joined shape can't use a column index at
        all (verified with EXPLAIN QUERY PLAN — adding one leaves the plan
        byte-identical: SCAN vv, SEARCH s USING INTEGER PRIMARY KEY, USE
        TEMP B-TREE FOR GROUP BY, because the plan is driven from the view
        table and reaches the source by rid), whereas the direct shape gets
        the same covering-index scan column_values does. Group-by-a-column
        on a table you just opened is the common case, and on a big merge
        it's the one that hurts."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        by_tag = column == TAG_GROUP_COLUMN
        with self._reader() as ro:
            src = self._source_lite_on(ro, handle["source_id"])
            members = self._resolve_members_on(ro, handle["source_id"])
            # Only grouping by tag needs these, and only for its cheapest
            # shape (below): "how many rows carry no tag" over a whole
            # member table is that table's row count minus the number of
            # distinct tagged rids, both of which come out of metadata and
            # an index rather than a scan.
            member_rows = (
                {m["source_id"]: self._source_lite_on(ro, m["source_id"])["row_count"] for m in members}
                if by_tag else {}
            )
        colnames = {c["name"]: c["type"] for c in src["columns"]}
        if not by_tag and column not in colnames:
            raise KeyError(column)
        is_datetime = not by_tag and colnames[column] == "datetime" and bucket_datetime

        # The two branches below reach the derived sidecar differently — the
        # direct shape joins it with USING(rid) (bare column refs), the view
        # shape chains a second ON join (refs go through the `d` alias) — so
        # the value expression and the path conditions are built per branch.
        def _val_expr(derived_alias: str | None) -> str:
            ref = self._col_ref(src, column, "s", derived_alias)
            return f"DAY_BUCKET({ref})" if is_datetime else ref

        def _path_for(m: dict, derived_alias: str | None) -> tuple[str, list]:
            clauses, p = self._path_where(path, colnames, "s", src, derived_alias, m["source_id"])
            return "".join(f" AND {c}" for c in clauses), p

        has_path = bool(path)
        whole_source = not has_path and self._grouping_covers_whole_source(handle, src)
        # A root_virtual view has no v.view_N to join at all (it was never
        # materialised — see _build_virtual_root_view), path or no path, so
        # it always takes the direct-against-member-tables shape; a real
        # 'root' view only gets to skip the join when it's provably
        # unfiltered (whole_source, no path).
        direct = whole_source or handle.get("kind") == "root_virtual"
        if direct and not by_tag:
            # A datetime column groups by DAY_BUCKET(col), a functional
            # expression a plain index on the raw column can't serve — same
            # exclusion _ensure_column_index_building already documents for
            # _numeric_expr'd numeric comparisons. Only triggered for the
            # true whole_source (unfiltered, no path) case, same scope as
            # before this had to also cover root_virtual.
            if whole_source and not is_datetime:
                for m in members:
                    self._ensure_column_index_building(
                        m["source_id"], column, self._index_table_for(src, column, m["table_name"], m["source_id"])
                    )

        branches: list[str] = []
        params: list[Any] = []
        if by_tag:
            branches, params = self._tag_group_branches(view_id, src, members, direct, member_rows, _path_for)
        else:
            for m in members:
                derived_alias = None if direct else "d"
                extra_where, path_params = _path_for(m, derived_alias)
                if direct:
                    scope = f"FROM {self._member_from(src, m, 's')} WHERE 1=1{extra_where}"
                else:
                    scope = (
                        f"FROM v.{q(view_id)} vv "
                        f"JOIN {q(m['table_name'])} s ON s.rid = vv.rid AND vv.source_id = {int(m['source_id'])}"
                        f"{self._member_derived_join(src, m, 's')} "
                        f"WHERE 1=1{extra_where}"
                    )
                branches.append(f"SELECT {_val_expr(derived_alias)} AS val, count(*) AS n {scope} GROUP BY 1")
                params.extend(path_params)
        union_sql = " UNION ALL ".join(branches)
        if direction is None:
            direction = "asc" if order == "value" else "desc"
        dir_sql = "DESC" if direction == "desc" else "ASC"
        if order == "value":
            if by_tag:
                # "By value" for tags means by tag name — the label the
                # analyst actually sees — not by the id the group carries.
                order_sql = (
                    f"(SELECT name FROM tag_defs WHERE id = val) COLLATE NOCASE {dir_sql}"
                )
            elif colnames[column] == "number":
                order_sql = f"{_numeric_expr('val')} {dir_sql}"
            else:
                order_sql = f"val COLLATE NOCASE {dir_sql}"
        else:
            order_sql = f"n {dir_sql}"
        # _interruptible innermost: an interrupt becomes OpCancelled before
        # _dropped_view_is_expired can see the OperationalError, and _reader
        # then closes (not repools) the interrupted connection on the way out.
        with self._reader() as ro, self._dropped_view_is_expired(), self._interruptible(op_token, ro):
            # HAVING is a no-op for a column grouping (every branch row is
            # a real group of at least one row) and load-bearing for a tag
            # one: the untagged branch is a bare aggregate, so it emits a
            # zero on a source where every row is tagged.
            rows = ro.execute(
                f"SELECT val, SUM(n) AS n FROM ({union_sql}) GROUP BY val HAVING SUM(n) > 0 "
                f"ORDER BY {order_sql} LIMIT ?",
                (*params, limit + 1),
            ).fetchall()
        truncated = len(rows) > limit
        rows = rows[:limit]
        return {"groups": [{"value": r["val"], "count": r["n"]} for r in rows], "truncated": truncated}

    def _tag_group_branches(self, view_id: str, src: dict, members: list[dict], direct: bool,
                            member_rows: dict[int, int], path_for) -> tuple[list[str], list]:
        """The UNION ALL branches behind grouping by tag: one group per tag,
        plus one for the untagged remainder. A row carrying two tags is
        counted under both — that's what makes "group by tag" answer the
        question analysts actually ask, and it's why these counts can sum to
        more than the view holds.

        Every shape here is written to pin the join order, because the
        obvious spellings do not survive the query planner. `v.view_N` is
        indexed on `pos` (its INTEGER PRIMARY KEY) and on nothing else, so
        reaching a view row by `rid` is a full scan of it. Given a
        `WHERE vv.source_id = ?` to work with, SQLite happily drives from
        row_tags' covering index and scans the entire view once per tagged
        row — measured at 150k x 300k row visits, minutes, on a case where
        the correct plan takes 200ms. So:

        - The view-scoped branches use `CROSS JOIN`, which SQLite documents
          as suppressing join reordering, to force the view to be the outer
          loop and row_tags to be probed by its (source_id, rid, tag_id)
          primary key.
        - They also drop the per-member `vv.source_id = ?` restriction
          entirely when there's no nested path: the ON clause already
          matches each view row against its own source's tags, so one
          view-wide pair of branches is both correct for a merge and
          cheaper than one pair per member.
        - The whole-table case (an unfiltered view, no path) never touches
          the source table at all: per-tag counts come from row_tags'
          own aggregate, and the untagged remainder is the member's row
          count minus its distinct tagged rids.
        - The source table, when a path needs it, is reached through a
          self-contained EXISTS rather than another join — nothing to
          reorder, and it works against _from_clause's USING(rid) shape
          without the alias juggling _derived_join exists for.
        """
        branches: list[str] = []
        params: list[Any] = []
        # Path conditions are compiled against _from_clause's USING(rid)
        # shape (derived_alias None) since the source table only ever
        # appears inside a standalone EXISTS here.
        per_member = [(m, *path_for(m, None)) for m in members]
        any_path = any(clause for _, clause, _ in per_member)

        if not direct and not any_path:
            branches.append(
                f"SELECT rt.tag_id AS val, count(*) AS n FROM v.{q(view_id)} vv "
                f"CROSS JOIN row_tags rt ON rt.source_id = vv.source_id AND rt.rid = vv.rid GROUP BY 1"
            )
            branches.append(
                f"SELECT NULL AS val, count(*) AS n FROM v.{q(view_id)} vv "
                f"WHERE NOT EXISTS (SELECT 1 FROM row_tags rt2 "
                f"WHERE rt2.source_id = vv.source_id AND rt2.rid = vv.rid)"
            )
            return branches, params

        for m, path_clause, path_params in per_member:
            sid = int(m["source_id"])
            member_from = self._member_from(src, m, "s")
            if direct and not path_clause:
                branches.append(
                    f"SELECT tag_id AS val, count(*) AS n FROM row_tags WHERE source_id = {sid} GROUP BY 1"
                )
                branches.append(
                    f"SELECT NULL AS val, {int(member_rows.get(sid, 0))} - "
                    f"(SELECT count(DISTINCT rid) FROM row_tags WHERE source_id = {sid}) AS n"
                )
            elif direct:
                branches.append(
                    f"SELECT rt.tag_id AS val, count(*) AS n FROM row_tags rt WHERE rt.source_id = {sid} "
                    f"AND EXISTS (SELECT 1 FROM {member_from} WHERE s.rid = rt.rid{path_clause}) GROUP BY 1"
                )
                params.extend(path_params)
                branches.append(
                    f"SELECT NULL AS val, count(*) AS n FROM {member_from} WHERE 1=1{path_clause} "
                    f"AND NOT EXISTS (SELECT 1 FROM row_tags rt2 "
                    f"WHERE rt2.source_id = {sid} AND rt2.rid = s.rid)"
                )
                params.extend(path_params)
            else:
                in_scope = (
                    f"EXISTS (SELECT 1 FROM {member_from} "
                    f"WHERE s.rid = vv.rid AND vv.source_id = {sid}{path_clause})"
                )
                branches.append(
                    f"SELECT rt.tag_id AS val, count(*) AS n FROM v.{q(view_id)} vv "
                    f"CROSS JOIN row_tags rt ON rt.source_id = vv.source_id AND rt.rid = vv.rid "
                    f"WHERE {in_scope} GROUP BY 1"
                )
                params.extend(path_params)
                branches.append(
                    f"SELECT NULL AS val, count(*) AS n FROM v.{q(view_id)} vv WHERE {in_scope} "
                    f"AND NOT EXISTS (SELECT 1 FROM row_tags rt2 "
                    f"WHERE rt2.source_id = vv.source_id AND rt2.rid = vv.rid)"
                )
                params.extend(path_params)
        return branches, params

    def expand_group(self, view_id: str, column: str, value: Any, path: list[dict] | None = None) -> dict:
        """Measures the group's real size within the current filtered root
        view (never trusts a client-cached count — the outer filter may have
        changed), then picks a strategy: small single-source groups of an
        *unfiltered* parent page directly (no v.view_N churn); large groups,
        groups of a filtered parent, and any group on a merge get a real
        indexed sub-view so paging stays O(window) even for a single
        dominant value across hundreds of thousands of rows.

        A root_virtual parent (see _build_virtual_root_view) has no
        v.view_N to join — it's always a single, unfiltered source, so
        every query here goes straight at the member table instead, and a
        materialised sub-view's root_pos is just the member table's own
        rid (root_virtual's outer order is rid order by construction)."""
        root = self._views.get(view_id)
        if not root:
            raise KeyError("View expired — rebuild it")
        src = self._source_lite(root["source_id"])
        colnames = {c["name"]: c["type"] for c in src["columns"]}
        by_tag = column == TAG_GROUP_COLUMN
        if not by_tag and column not in colnames:
            raise KeyError(column)

        members = self._resolve_members(root["source_id"])
        is_root_virtual = root.get("kind") == "root_virtual"
        # Same split group_summary makes: the direct-against-the-table shape
        # picks the derived sidecar up through _from_clause's USING(rid)
        # (bare refs), the view-join shape chains a second ON join and
        # refers to it by alias.
        d_alias = None if is_root_virtual else "d"

        def _conds(m: dict) -> tuple[str, list]:
            """This group's predicate for one member. Per member because
            both halves can be: a tag test is keyed by source_id (row_tags
            is), and so is any tag level in the nested path."""
            if by_tag:
                c, v = self._tag_condition("s", m["source_id"], value)
            else:
                c, v = self._eq_condition(
                    self._col_ref(src, column, "s", d_alias), value, colnames[column] == "datetime"
                )
            pc, pp = self._path_where(path, colnames, "s", src, d_alias, m["source_id"])
            return f"{c}" + "".join(f" AND {x}" for x in pc), [*v, *pp]

        with self.lock:
            total = 0
            for m in members:
                where_sql, where_params = _conds(m)
                if is_root_virtual:
                    n = self.db.execute(
                        f"SELECT count(*) FROM {self._member_from(src, m, 's')} WHERE {where_sql}",
                        where_params,
                    ).fetchone()[0]
                else:
                    n = self.db.execute(
                        f"SELECT count(*) FROM v.{q(view_id)} vv JOIN {q(m['table_name'])} s "
                        f"ON s.rid = vv.rid AND vv.source_id = ?{self._member_derived_join(src, m, 's')} WHERE {where_sql}",
                        [m["source_id"], *where_params],
                    ).fetchone()[0]
                total += n

        # The virtual fast path reads straight off the member table with
        # nothing but column=value (+ the nested path) — _virtual_group_where
        # has no view to join and so no way to express the parent's own
        # filters/search/timeframe. It is therefore only correct when the
        # parent provably holds every row of the source; a filtered parent
        # has to materialise. The failure it prevents is quiet, because the
        # counts come from the other side: group_summary and `total` above
        # both join the view and stay right, so the grid asks for row_count
        # rows and gets the first row_count of a *longer*, unfiltered list —
        # rows the filter excluded, rendered under a correct count. Tag and
        # export on the group read the same way, which makes it an
        # over-tagging bug too, not just a display one.
        covers_source = self._grouping_covers_whole_source(root, src)
        is_merge = bool(src.get("is_merge"))
        if is_merge or not covers_source or total > self.GROUP_MATERIALIZE_THRESHOLD:
            self._view_seq += 1
            vid = f"view_{self._view_seq}"
            branches = []
            params: list[Any] = []
            for m in members:
                where_sql, where_params = _conds(m)
                if is_root_virtual:
                    branches.append(
                        f"SELECT s.rid AS root_pos, {int(m['source_id'])} AS source_id, s.rid AS rid "
                        f"FROM {self._member_from(src, m, 's')} WHERE {where_sql}"
                    )
                    params.extend(where_params)
                else:
                    branches.append(
                        f"SELECT vv.pos AS root_pos, vv.source_id, vv.rid FROM v.{q(view_id)} vv "
                        f"JOIN {q(m['table_name'])} s ON s.rid = vv.rid AND vv.source_id = ?{self._member_derived_join(src, m, 's')} "
                        f"WHERE {where_sql}"
                    )
                    params.extend([m["source_id"], *where_params])
            union_sql = " UNION ALL ".join(branches)
            with self.lock, self.db:
                # Same ORDER BY-fed insert as build_view: pos auto-assigns
                # 1..N in insertion (= sorted) order, rowcount replaces a
                # count(*) re-scan. root_pos is unique across the union
                # (vv.pos from one view, or s.rid from one member table),
                # so the order is fully determined.
                self.db.execute(f"CREATE TABLE v.{q(vid)} (pos INTEGER PRIMARY KEY, source_id INTEGER, rid INTEGER)")
                n = self.db.execute(
                    f"INSERT INTO v.{q(vid)}(source_id, rid) "
                    f"SELECT source_id, rid FROM ({union_sql}) ORDER BY root_pos",
                    params,
                ).rowcount
            self._views[vid] = {
                "view_id": vid, "source_id": root["source_id"], "row_count": n,
                "kind": "group", "parent_view_id": view_id,
            }
            return {"view_id": vid, "row_count": n, "elapsed_ms": 0}

        self._view_seq += 1
        vid = f"view_{self._view_seq}"
        self._views[vid] = {
            "view_id": vid, "source_id": root["source_id"], "row_count": total,
            "kind": "group_virtual", "parent_view_id": view_id,
            "column": column, "value": value, "members": members, "path": path,
        }
        return {"view_id": vid, "row_count": total, "elapsed_ms": 0}

    def _virtual_group_where(self, handle: dict, conn: sqlite3.Connection | None = None) -> tuple[str, list]:
        """WHERE fragment (unqualified column names, against the member
        table directly) for a group_virtual handle's own column=value plus
        every outer path level already fixed by nested grouping — shared by
        every place that reads straight off the backing member table for a
        small, ungathered group. `conn` lets a _reader() caller resolve the
        member's column types on its own connection; writer-side callers
        (tag/export on a group) omit it and go through the lock as before."""
        column, value = handle["column"], handle["value"]
        member_sid = handle["members"][0]["source_id"]
        src = self._source_lite_on(conn, member_sid) if conn is not None else self._source_lite(member_sid)
        colnames = {c["name"]: c["type"] for c in src["columns"]}
        # Unqualified names throughout: every caller runs this against the
        # member table joined to its sidecar with USING(rid), where a base
        # and a derived column are both addressable bare.
        if column == TAG_GROUP_COLUMN:
            cond_sql, cond_val = self._tag_condition("", member_sid, value)
        else:
            cond_sql, cond_val = self._eq_condition(q(column), value, colnames[column] == "datetime")
        path_clauses, path_params = self._path_where(handle.get("path"), colnames, "", src, None, member_sid)
        extra = "".join(f" AND {c}" for c in path_clauses)
        return f"{cond_sql}{extra}", [*cond_val, *path_params]

    @staticmethod
    def _compile_condition(col: str, op: str, val: Any, colnames: dict) -> tuple[str, list]:
        """Compile one column/op/value condition into a parameterized clause.
        Shared by the flat quick-filter list and the guided filter-tree's
        'cond' leaves. Returns ("", []) for an unknown column/op or an empty
        value on ops that require one."""
        if col not in colnames:
            return "", []
        c = q(col)
        numeric = colnames[col] == "number"
        if op == "contains":
            if val == "":
                return "", []
            return f"{c} LIKE ? ESCAPE '\\'", [f"%{_esc_like(val)}%"]
        if op == "not_contains":
            if val == "":
                return "", []
            return f"({c} NOT LIKE ? ESCAPE '\\' OR {c} IS NULL)", [f"%{_esc_like(val)}%"]
        if op == "equals":
            if val == "":
                return "", []
            return f"{c} = ?", [val]
        if op == "not_equals":
            if val == "":
                return "", []
            return f"({c} <> ? OR {c} IS NULL)", [val]
        if op == "starts":
            if val == "":
                return "", []
            return f"{c} LIKE ? ESCAPE '\\'", [f"{_esc_like(val)}%"]
        if op == "regex":
            if val == "":
                return "", []
            return f"{c} REGEXP ?", [val]
        if op in (">", ">=", "<", "<="):
            if val == "":
                return "", []
            lhs = _numeric_expr(c) if numeric else c
            return f"{lhs} {op} ?", [float(val) if numeric else val]
        if op == "empty":
            return f"({c} IS NULL OR {c} = '')", []
        if op == "not_empty":
            return f"({c} IS NOT NULL AND {c} <> '')", []
        if op == "in":
            items = [s for s in (val if isinstance(val, list) else str(val).split("\n")) if s != ""]
            if not items:
                return "", []
            return f"{c} IN ({','.join('?' * len(items))})", items
        return "", []

    @staticmethod
    def _tree_has_raw(node: dict | None) -> bool:
        """True if any node in the filter tree is a raw-SQL fragment. Raw
        nodes assume a single table_name for their EXPLAIN QUERY PLAN
        validation, so they're rejected for merges rather than extended to
        dry-run per member."""
        if not node:
            return False
        if node.get("type") == "raw":
            return True
        if node.get("type") == "group":
            return any(Store._tree_has_raw(c) for c in node.get("children", []))
        return False

    def _index_table_for(self, src: dict, column: str, table_name: str | None = None,
                         member_id: int | None = None) -> str:
        """Which physical table an auto-created index for this column
        belongs on — a derived column's values live in the drv_<id>
        sidecar, not in src_<id>. `table_name` is the caller's already
        resolved member table (a merge's src has none of its own), and
        `member_id` names the member whose sidecar carries a merge's
        derived values (drv_<-3> doesn't exist)."""
        for c in src["columns"]:
            if c.get("derived") and c["name"].lower() == str(column).lower():
                return self._derived_table(member_id if member_id is not None else src["id"])
        return table_name or src["table_name"]

    def _compile_tree(self, node: dict | None, colnames: dict, source_id: int, src: dict) -> tuple[str, list]:
        """Compile a guided filter-builder tree (group/cond/raw nodes) into a
        parameterized WHERE fragment."""
        if not node:
            return "", []
        kind = node.get("type")
        if kind == "group":
            parts: list[str] = []
            params: list[Any] = []
            for child in node.get("children", []):
                c, p = self._compile_tree(child, colnames, source_id, src)
                if c:
                    parts.append(c)
                    params.extend(p)
            if not parts:
                return "", []
            joiner = " OR " if node.get("op") == "OR" else " AND "
            return "(" + joiner.join(parts) + ")", params
        if kind == "raw":
            sql = node.get("sql", "")
            # Re-validated on every compile, not just at authoring time — a
            # stored preset may be replayed later against a different schema.
            self.validate_where_fragment(source_id, sql)
            return f"({sql})", []
        if kind == "cond":
            col, op = node.get("column"), node.get("op")
            c, p = self._compile_condition(col, op, node.get("value", ""), colnames)
            if c and op in SARGABLE_OPS and col in colnames:
                self._ensure_column_index_building(source_id, col, self._index_table_for(src, col))
            return c, p
        return "", []

    FORBIDDEN_SQL_RE = re.compile(
        r"\b(select|insert|update|delete|drop|alter|create|replace|reindex|"
        r"union|attach|detach|pragma|vacuum|exec)\b",
        re.IGNORECASE,
    )
    ALLOWED_SQL_WORDS = {
        "AND", "OR", "NOT", "IS", "NULL", "IN", "BETWEEN", "LIKE", "GLOB", "REGEXP", "ESCAPE",
        "CAST", "REAL", "TEXT", "INTEGER", "COALESCE", "LENGTH", "LOWER", "UPPER", "SUBSTR",
        "TRIM", "TRUE", "FALSE",
    }

    def validate_where_fragment(self, source_id: int, fragment: str) -> None:
        """Raise ValueError with a human-readable reason for anything that is
        not a safe, side-effect-free boolean expression over this source's
        own columns.

        The actual safety boundary is that this fragment is only ever spliced
        into `WHERE (<fragment>)` against a FROM clause the caller cannot
        influence (always this source's own table) — user text never
        controls FROM/JOIN. Blacklisting SELECT closes the one way that
        boundary could be defeated (a subquery reading row_tags,
        sqlite_master, or another source's table). The identifier allowlist
        and the EXPLAIN QUERY PLAN dry-run below are good defense-in-depth
        and produce better error messages, but are not by themselves a
        sufficient boundary — regex-based SQL sanitization has known edge
        cases, so raw mode's blast radius is kept small by construction
        (no SELECT at all, ever) rather than by trying to perfect this list.
        """
        frag = (fragment or "").strip()
        if not frag:
            raise ValueError("Empty expression")
        if frag.count("(") != frag.count(")"):
            raise ValueError("Unbalanced parentheses")
        for qc in ("'", '"'):
            if (frag.count(qc) - 2 * frag.count(qc * 2)) % 2 != 0:
                raise ValueError("Unbalanced quotes")

        # Blank out quoted-literal contents before keyword/identifier scanning
        # so a forensic value like CommandLine = 'SELECT * FROM users' isn't
        # mistaken for a SELECT statement, and a value like 'Sysmon' isn't
        # mistaken for a bare identifier.
        structural = _blank_string_literals(frag)
        if ";" in structural or "--" in structural or "/*" in structural or "*/" in structural:
            raise ValueError("Statement separators and comments are not allowed")
        if self.FORBIDDEN_SQL_RE.search(structural):
            raise ValueError("Only a boolean filter expression is allowed — no SELECT/PRAGMA/ATTACH/etc.")

        src = self.get_source(source_id)
        colnames = {c["name"] for c in src["columns"]}
        for ident in re.findall(r"\b[A-Za-z_][A-Za-z0-9_]*\b", structural):
            if ident.upper() in self.ALLOWED_SQL_WORDS or ident in colnames or NUM_RE.match(ident):
                continue
            raise ValueError(f"Unknown identifier: {ident}")

        # A merge has no table of its own: the fragment is later spliced
        # into each member's branch, so the dry-run below must pass against
        # every member (invariant #9 — its identifier check above already
        # ran against the merge's exposed column set).
        froms = ([self._from_clause(self._source_lite(m["source_id"]))
                  for m in self._resolve_members(source_id)]
                 if source_id < 0 else [self._from_clause(src)])
        ro = sqlite3.connect(f"file:{self.path}?mode=ro", uri=True, check_same_thread=False)
        ro.row_factory = sqlite3.Row
        ro.create_function("REGEXP", 2, _regexp, deterministic=True)
        try:
            for f_clause in froms:
                ro.execute(f"EXPLAIN QUERY PLAN SELECT 1 FROM {f_clause} WHERE ({frag}) LIMIT 0")
        except sqlite3.Error as e:
            raise ValueError(f"Not a valid filter expression: {e}")
        finally:
            ro.close()

    def _compile_where(self, source_id, src, spec, colnames) -> tuple[str, list]:
        clauses: list[str] = []
        params: list[Any] = []

        for f in spec.get("filters", []):
            col, op = f.get("column"), f.get("op", "contains")
            c, p = self._compile_condition(col, op, f.get("value", ""), colnames)
            if c:
                clauses.append(c)
                params.extend(p)
                if op in SARGABLE_OPS and col in colnames:
                    self._ensure_column_index_building(source_id, col, self._index_table_for(src, col))

        tree = spec.get("filter_tree")
        if tree:
            c, p = self._compile_tree(tree, colnames, source_id, src)
            if c:
                clauses.append(c)
                params.extend(p)

        search_mode = spec.get("search_mode", "contains")
        search = (spec.get("search") or "").strip()
        if search_mode == "regex" and search:
            # Base columns only, matching build_fts's doc view exactly — the
            # indexed and fallback paths have to search the same text, and
            # derived values are computed from data that's already in it.
            blob = _blob_expr([c["name"] for c in self._base_cols(src)])
            clauses.append(f"({blob}) REGEXP ?")
            params.append(search)
        elif search_mode == "advanced":
            terms = spec.get("search_terms") or []
            if terms:
                cols = [c["name"] for c in self._base_cols(src)]
                if not src["has_fts"]:
                    self._ensure_fts_building(source_id)
                if src["has_fts"]:
                    clause, p = _advanced_fts_clause(q("fts_" + str(source_id)), _blob_expr(cols), terms)
                else:
                    clause, p = _advanced_like_clause(cols, terms)
                if clause:
                    clauses.append(clause)
                    params.extend(p)
        elif search:
            # A true substring match either way — the trigram index (a bare
            # `doc LIKE ?` against the fts table, pushed down onto the
            # trigram tokenizer; see build_fts/_fts_like_pattern for why
            # not MATCH) makes it fast once built. Falls back to a blob
            # LIKE scan — same results, just slower — when the index isn't
            # ready yet (kicked off here in the background) or the term
            # can't route through it (under 3 chars, or contains a LIKE
            # wildcard the unescaped pushdown form would misinterpret).
            if not src["has_fts"]:
                self._ensure_fts_building(source_id)
            pattern = _fts_like_pattern(search)
            if src["has_fts"] and pattern is not None:
                fts_ident = q("fts_" + str(source_id))
                clauses.append(f"rid IN (SELECT rowid FROM {fts_ident} WHERE doc LIKE ?)")
                params.append(pattern)
            else:
                blob = _blob_expr([c["name"] for c in self._base_cols(src)])
                clauses.append(f"({blob}) LIKE ? ESCAPE '\\'")
                params.append(f"%{_esc_like(search)}%")

        tag_filter = spec.get("tags") or []
        if tag_filter:
            if tag_filter == ["__any__"]:
                clauses.append("rid IN (SELECT rid FROM row_tags WHERE source_id=?)")
                params.append(source_id)
            elif tag_filter == ["__none__"]:
                clauses.append("rid NOT IN (SELECT rid FROM row_tags WHERE source_id=?)")
                params.append(source_id)
            else:
                ids = [int(t) for t in tag_filter if str(t).isdigit()]
                if ids:
                    clauses.append(
                        f"rid IN (SELECT rid FROM row_tags WHERE source_id=? AND tag_id IN ({','.join('?' * len(ids))}))"
                    )
                    params.append(source_id)
                    params.extend(ids)

        time_range = spec.get("time_range")
        if time_range and time_range.get("enabled") and (time_range.get("start") or time_range.get("end")):
            start_norm = _ts_normalize(time_range["start"]) if time_range.get("start") else None
            end_norm = _ts_normalize(time_range["end"]) if time_range.get("end") else None
            col = time_range.get("column")
            # An explicit column only applies here if this member actually
            # has it as a datetime column — otherwise (blank/"all columns"
            # mode, or a merge member that doesn't share this column) fall
            # back to every datetime column on this table, OR'd together:
            # a row counts as in-range if *any* of its timestamps is —
            # the "timestomped creation date, but a real modified date"
            # case the timeframe filter exists for in the first place.
            if col and colnames.get(col) == "datetime":
                ts_cols = [col]
            else:
                ts_cols = [c for c, t in colnames.items() if t == "datetime"]
            if ts_cols and (start_norm or end_norm):
                parts = []
                for c in ts_cols:
                    expr = f"TS_NORMALIZE({q(c)})"
                    if start_norm is not None and end_norm is not None:
                        parts.append(f"{expr} BETWEEN ? AND ?")
                        params.extend([start_norm, end_norm])
                    elif start_norm is not None:
                        parts.append(f"{expr} >= ?")
                        params.append(start_norm)
                    else:
                        parts.append(f"{expr} <= ?")
                        params.append(end_norm)
                clauses.append("(" + " OR ".join(parts) + ")")

        return " AND ".join(clauses), params

    @staticmethod
    def _compile_order(spec, colnames) -> str:
        sort = spec.get("sort") or []
        parts = []
        for s in sort:
            col = s.get("column")
            direction = "DESC" if str(s.get("dir", "asc")).lower() == "desc" else "ASC"
            if col == "__line__":
                # The gutter's Line header: original file order, both ways.
                # rid IS the line order, so ascending is the no-sort default
                # — this exists for DESC, and so the client can state line
                # order explicitly.
                parts.append(f"rid {direction}")
                continue
            if col not in colnames:
                continue
            if colnames[col] == "number":
                parts.append(f"{_numeric_expr(q(col))} {direction}")
            else:
                parts.append(f"{q(col)} COLLATE NOCASE {direction}")
        parts.append("rid ASC")
        return "ORDER BY " + ", ".join(parts)

    @staticmethod
    def _inline_sql_params(sql: str, params: list) -> str:
        """Substitute ? placeholders with safely-quoted literals, producing
        SQL fit to hand the analyst in the SQL pane. Skips any ? that sits
        inside a string literal — a raw filter fragment the analyst typed
        can legitimately contain one — by walking the text with the same
        ''-doubling rule SQLite itself uses."""
        out: list[str] = []
        it = iter(params)
        in_str = False
        i = 0
        while i < len(sql):
            ch = sql[i]
            if in_str:
                if ch == "'":
                    if i + 1 < len(sql) and sql[i + 1] == "'":
                        out.append("''")
                        i += 2
                        continue
                    in_str = False
                out.append(ch)
            elif ch == "'":
                in_str = True
                out.append(ch)
            elif ch == "?":
                try:
                    v = next(it)
                except StopIteration:
                    raise ValueError("Filter could not be rendered as SQL (placeholder mismatch)")
                if v is None:
                    out.append("NULL")
                elif isinstance(v, (int, float)) and not isinstance(v, bool):
                    out.append(repr(v))
                else:
                    out.append("'" + str(v).replace("'", "''") + "'")
            else:
                out.append(ch)
            i += 1
        return "".join(out)

    def spec_sql(self, source_id: int, spec: dict) -> str:
        """The current filter/sort/search spec rendered as a runnable
        SELECT — the "open this filter in the SQL pane" action. Compiled by
        the exact same _compile_where/_compile_order the view build uses,
        so what the analyst gets is the query their grid is actually
        showing, not a hand-maintained approximation; the only difference
        is that bound parameters are inlined as literals so the text stands
        alone. run_sql's own connection registers REGEXP/TS_NORMALIZE/
        DAY_BUCKET, so every compiled shape runs there unchanged."""
        src = self._source_lite(source_id)
        colnames = {c["name"]: c["type"] for c in src["columns"]}
        order = self._compile_order(spec, colnames)
        if src.get("is_merge"):
            collist = ", ".join(q(c) for c in colnames)
            branches = []
            params: list[Any] = []
            for m in self._resolve_members(source_id):
                msrc = self._source_lite(m["source_id"])
                where, p = self._compile_where(m["source_id"], msrc, spec, colnames)
                branches.append(
                    f"SELECT {int(m['source_id'])} AS source_id, rid, {collist}\n"
                    f"FROM {self._member_from(src, m)}"
                    + (f"\nWHERE {where}" if where else "")
                )
                params.extend(p)
            sql = "\nUNION ALL\n".join(branches) + f"\n{order}"
        else:
            where, params = self._compile_where(source_id, src, spec, colnames)
            collist = ", ".join(q(c) for c in colnames)
            sql = f"SELECT rid, {collist}\nFROM {self._from_clause(src)}"
            if where:
                sql += f"\nWHERE {where}"
            sql += f"\n{order}"
        return self._inline_sql_params(sql, params)

    def fetch_rows(self, view_id: str, start: int, count: int) -> dict:
        """Pages v.view_N by pos, then resolves each row's cells/tags/notes
        against ITS OWN source_id — not a single handle-level constant. For
        an ordinary view every row shares one source_id (one query, same
        cost as before); for a merged view rows are grouped by their real
        distinct member source_id (bounded by window size × member count)
        and resolved per member. Either way this stays O(window)."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        if handle.get("kind") == "group_virtual":
            return self._fetch_virtual_group_rows(handle, start, count)
        if handle.get("kind") == "root_virtual":
            return self._fetch_virtual_root_rows(handle, start, count)

        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, handle["source_id"])
            cols = [c["name"] for c in src["columns"]]

            vrows = ro.execute(
                f"SELECT pos, source_id, rid FROM v.{q(view_id)} "
                f"WHERE pos >= ? AND pos < ? ORDER BY pos",
                (start + 1, start + 1 + count),
            ).fetchall()

            by_source: dict[int, list[int]] = {}
            for r in vrows:
                by_source.setdefault(r["source_id"], []).append(r["rid"])

            cellmap: dict[tuple[int, int], tuple] = {}
            tags: dict[tuple[int, int], list[int]] = {}
            notes: dict[tuple[int, int], str] = {}
            sel = ", ".join(q(c) for c in cols)
            for sid, rids in by_source.items():
                member_src = self._source_lite_on(ro, sid)
                ph = ",".join("?" * len(rids))
                for row in ro.execute(
                    f"SELECT rid, {sel} FROM {self._from_clause(member_src)} WHERE rid IN ({ph})", rids
                ):
                    # positional: the SELECT puts rid first, then cols in
                    # order — tuple(row) is one C-level copy, vs a
                    # per-cell name lookup for row[c]
                    t = tuple(row)
                    cellmap[(sid, t[0])] = t[1:]
                for rid, tid in ro.execute(
                    f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})",
                    [sid, *rids],
                ):
                    tags.setdefault((sid, rid), []).append(tid)
                for rid, note in ro.execute(
                    f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})",
                    [sid, *rids],
                ):
                    notes[(sid, rid)] = note

        out = []
        for r in vrows:
            key = (r["source_id"], r["rid"])
            out.append({
                "pos": r["pos"] - 1,
                "source_id": r["source_id"],
                "rid": r["rid"],
                "cells": list(cellmap.get(key, ("",) * len(cols))),
                "tags": tags.get(key, []),
                "note": notes.get(key),
            })
        return {"start": start, "rows": out}

    def _fetch_virtual_group_rows(self, handle: dict, start: int, count: int) -> dict:
        """Pages a small, ungathered group directly with LIMIT/OFFSET —
        expand_group only chooses this path for a single-source (non-merge)
        group under GROUP_MATERIALIZE_THRESHOLD, so `handle['members']`
        always has exactly one entry here. Ordered by rid (stable insertion
        order), not the outer view's actual sort — an accepted simplification
        for this fast/small-group path; a group above the threshold (or on a
        merge) always gets the materialized path instead, which does
        preserve the outer sort via root_pos."""
        member = handle["members"][0]
        sid = member["source_id"]
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, sid)
            cols = [c["name"] for c in src["columns"]]
            where_sql, where_params = self._virtual_group_where(handle, ro)
            sel = ", ".join(q(c) for c in cols)

            rows = ro.execute(
                f"SELECT rid, {sel} FROM {self._from_clause(src)} WHERE {where_sql} "
                f"ORDER BY rid LIMIT ? OFFSET ?",
                [*where_params, count, start],
            ).fetchall()
            rids = [r["rid"] for r in rows]
            tags: dict[int, list[int]] = {}
            notes: dict[int, str] = {}
            if rids:
                ph = ",".join("?" * len(rids))
                for rid, tid in ro.execute(
                    f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    tags.setdefault(rid, []).append(tid)
                for rid, note in ro.execute(
                    f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    notes[rid] = note

        out = []
        for i, r in enumerate(rows):
            t = tuple(r)
            out.append({
                "pos": start + i,
                "source_id": sid,
                "rid": t[0],
                "cells": list(t[1:]),
                "tags": tags.get(t[0], []),
                "note": notes.get(t[0]),
            })
        return {"start": start, "rows": out}

    def _fetch_virtual_root_rows(self, handle: dict, start: int, count: int) -> dict:
        """Pages the source table directly by rid *range* — no v.view_N to
        build, since an unfiltered/unsorted root_virtual view is, by
        construction, every row of the source in its natural (already
        free) order, and rid is contiguous from 1 (see
        _build_virtual_root_view), so the window at `start` is exactly
        rids start+1 .. start+count. A rid-range seek is O(log n) at any
        depth, where the LIMIT/OFFSET shape _fetch_virtual_group_rows
        uses (fine for its ≤GROUP_MATERIALIZE_THRESHOLD rows) walks and
        discards `start` rows first — measured 26ms vs 0.1ms at the far
        end of 2M rows, and that gap grows linearly with depth."""
        sid = handle["source_id"]
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, sid)
            cols = [c["name"] for c in src["columns"]]
            sel = ", ".join(q(c) for c in cols)

            # _from_clause adds the derived-value sidecar join only when the
            # source has derived columns, and that join is on drv's INTEGER
            # PRIMARY KEY — at most one match per rid, so the rid window
            # still yields exactly its rows in rid order and `pos = rid - 1`
            # below stays exact (invariant #2).
            rows = ro.execute(
                f"SELECT rid, {sel} FROM {self._from_clause(src)} WHERE rid >= ? AND rid < ? ORDER BY rid",
                (start + 1, start + 1 + count),
            ).fetchall()
            rids = [r["rid"] for r in rows]
            tags: dict[int, list[int]] = {}
            notes: dict[int, str] = {}
            if rids:
                ph = ",".join("?" * len(rids))
                for rid, tid in ro.execute(
                    f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    tags.setdefault(rid, []).append(tid)
                for rid, note in ro.execute(
                    f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    notes[rid] = note

        out = []
        for i, r in enumerate(rows):
            t = tuple(r)
            out.append({
                "pos": t[0] - 1,  # rid - 1: contiguous rids make this exact
                "source_id": sid,
                "rid": t[0],
                "cells": list(t[1:]),
                "tags": tags.get(t[0], []),
                "note": notes.get(t[0]),
            })
        return {"start": start, "rows": out}

    def tag_positions(self, view_id: str, limit: int = 20_000) -> list[list[int]]:
        """Positions of tagged rows inside a view, for the scrollbar rail."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        if handle.get("kind") == "group_virtual":
            return []  # no pos-ordered backing table for a small ungathered group — documented limitation
        with self._reader() as ro, self._dropped_view_is_expired():
            if handle.get("kind") == "root_virtual":
                # pos = rid - 1 directly: an unfiltered root_virtual view is
                # every row of the source in rid order, so a tagged row's rank
                # IS its rid - no join against a backing view table needed, and
                # (unlike group_virtual) never a stub — see _build_virtual_root_view.
                rows = ro.execute(
                    "SELECT rid - 1 AS p, tag_id FROM row_tags WHERE source_id=? ORDER BY rid LIMIT ?",
                    (handle["source_id"], limit),
                ).fetchall()
                return [[r["p"], r["tag_id"]] for r in rows]
            # The join below has no better plan than scanning the whole view
            # probing row_tags per row (the view has no index on rid — measured
            # ~115ms per 2M view rows even with row_tags empty). An untagged
            # source is the common case and this runs after every view build,
            # so check for any tag at all first — one indexed probe per member.
            if "source_id" in handle:
                any_tags = any(
                    ro.execute(
                        "SELECT 1 FROM row_tags WHERE source_id=? LIMIT 1", (m["source_id"],)
                    ).fetchone()
                    for m in self._resolve_members_on(ro, handle["source_id"])
                )
                if not any_tags:
                    return []
            rows = ro.execute(
                f"SELECT vv.pos - 1 AS p, rt.tag_id FROM v.{q(view_id)} vv "
                f"JOIN row_tags rt ON rt.rid = vv.rid AND rt.source_id = vv.source_id "
                f"ORDER BY vv.pos LIMIT ?",
                (limit,),
            ).fetchall()
        return [[r["p"], r["tag_id"]] for r in rows]

    def find_position(self, view_id: str, source_id: int, rid: int) -> int | None:
        """0-indexed position of one (source_id, rid) row inside a view, or
        None if it's not in this view. Positions are view-specific and get
        wiped on every rebuild (see CLAUDE.md); this is how the frontend
        re-centers the grid on a previously-selected row after its view is
        rebuilt out from under it, e.g. clearing filters."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        if handle.get("kind") == "group_virtual":
            return None  # no pos-ordered backing table for a small ungathered group
        with self._reader() as ro, self._dropped_view_is_expired():
            if handle.get("kind") == "root_virtual":
                if source_id != handle["source_id"]:
                    return None
                src = self._source_lite_on(ro, source_id)
                row = ro.execute(
                    f"SELECT 1 FROM {q(src['table_name'])} WHERE rid=?", (rid,)
                ).fetchone()
                return rid - 1 if row else None
            row = ro.execute(
                f"SELECT pos FROM v.{q(view_id)} WHERE source_id=? AND rid=?",
                (source_id, rid),
            ).fetchone()
        return row["pos"] - 1 if row else None

    def find_nearest_timestamp(self, view_id: str, value: str, column: str | None = None) -> dict | None:
        """Position (0-indexed, within this view) of the row whose timestamp
        is closest in time to `value` — "jump to timestamp". `column` picks
        which datetime column measures closeness; None means the nearest
        across every datetime column, same spirit as the timeframe filter's
        all-columns mode.

        Closeness is |julianday(a) - julianday(b)| over TS_NORMALIZE'd
        values — string comparison can order timestamps but can't measure
        between them. That makes this a scan of the view (TS_NORMALIZE is a
        registered Python function), which is the same cost shape as
        group_summary's whole-source aggregate: fine for a user-initiated,
        occasional action, and it runs on a pooled reader so it never
        blocks anything. Rows whose timestamp doesn't parse simply can't
        win. Returns None when no row has a usable timestamp."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        norm = _ts_normalize(value)
        if norm is None:
            raise ValueError("Not a recognized timestamp — try YYYY-MM-DD HH:MM:SS")
        best: tuple | None = None  # (diff, source_id, rid, pos_or_None, ts)
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, handle["source_id"])
            colnames = {c["name"]: c["type"] for c in src["columns"]}
            if column:
                if colnames.get(column) != "datetime":
                    raise ValueError(f"{column!r} is not a datetime column")
                ts_cols = [column]
            else:
                ts_cols = [c for c, t in colnames.items() if t == "datetime"]
            if not ts_cols:
                raise ValueError("This table has no datetime columns")
            kind = handle.get("kind")
            members = self._resolve_members_on(ro, handle["source_id"])
            for m in members:
                msrc = self._source_lite_on(ro, m["source_id"])
                for c in ts_cols:
                    if not any(cc["name"] == c for cc in msrc["columns"]):
                        continue  # merge member without this column
                    if kind == "root_virtual":
                        ts_expr = f"TS_NORMALIZE({q(c)})"
                        row = ro.execute(
                            f"SELECT rid, NULL AS pos, {ts_expr} AS ts, "
                            f"ABS(julianday({ts_expr}) - julianday(?)) AS d "
                            f"FROM {self._from_clause(msrc)} "
                            f"WHERE {ts_expr} IS NOT NULL ORDER BY d LIMIT 1",
                            (norm,),
                        ).fetchone()
                    elif kind == "group_virtual":
                        where_sql, where_params = self._virtual_group_where(handle, ro)
                        ts_expr = f"TS_NORMALIZE({q(c)})"
                        row = ro.execute(
                            f"SELECT rid, NULL AS pos, {ts_expr} AS ts, "
                            f"ABS(julianday({ts_expr}) - julianday(?)) AS d "
                            f"FROM {self._member_from(msrc, m)} WHERE ({where_sql}) "
                            f"AND {ts_expr} IS NOT NULL ORDER BY d LIMIT 1",
                            (norm, *where_params),
                        ).fetchone()
                    else:
                        ts_expr = f"TS_NORMALIZE({self._col_ref(msrc, c, 's', 'd')})"
                        row = ro.execute(
                            f"SELECT vv.rid AS rid, vv.pos AS pos, {ts_expr} AS ts, "
                            f"ABS(julianday({ts_expr}) - julianday(?)) AS d "
                            f"FROM v.{q(view_id)} vv "
                            f"JOIN {q(m['table_name'])} s ON s.rid = vv.rid AND vv.source_id = ?"
                            f"{self._derived_join(msrc, 's')} "
                            f"WHERE {ts_expr} IS NOT NULL ORDER BY d LIMIT 1",
                            (norm, m["source_id"]),
                        ).fetchone()
                    if row and row["d"] is not None and (best is None or row["d"] < best[0]):
                        best = (row["d"], m["source_id"], row["rid"], row["pos"], row["ts"])
            if best is None:
                return None
            d, sid, rid, pos, ts = best
            if kind == "root_virtual":
                pos = rid - 1  # exact — see _fetch_virtual_root_rows / invariant #2
            elif kind == "group_virtual":
                # pos within a virtual group = how many of its rows page in
                # before this rid (they page in rid order, see
                # _fetch_virtual_group_rows).
                where_sql, where_params = self._virtual_group_where(handle, ro)
                member = handle["members"][0]
                msrc = self._source_lite_on(ro, member["source_id"])
                pos = ro.execute(
                    f"SELECT COUNT(*) FROM {self._member_from(msrc, member)} "
                    f"WHERE ({where_sql}) AND rid < ?",
                    (*where_params, rid),
                ).fetchone()[0]
            else:
                pos = pos - 1  # v.view_N pos is 1-based
        return {"pos": pos, "source_id": sid, "rid": rid, "ts": ts}

    def column_values(self, source_id: int, column: str, limit: int = 200) -> list[dict]:
        """Distinct values + counts for one column — the filter box's
        value-picker dropdown.

        Kicks off the same lazy B-tree index _compile_where's sargable
        filters do (see _ensure_column_index_building), because this exact
        shape is what a plain single-column index is *best* at: verified
        with EXPLAIN QUERY PLAN, `SELECT col, count(*) FROM src GROUP BY 1`
        goes from `SCAN src` + `USE TEMP B-TREE FOR GROUP BY` to `SCAN src
        USING COVERING INDEX` — never touching the table pages, and with the
        grouping sort gone entirely. Fire-and-forget as everywhere else: this
        call still runs today's scan, the next open of the dropdown gets the
        index. It's also the same index an equals/in filter on this column
        wants, which is the overwhelmingly likely next action after picking
        a value out of this list."""
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, source_id)
            names = {c["name"] for c in src["columns"]}
            if column not in names:
                raise KeyError(column)
            members = self._resolve_members_on(ro, source_id)
            for m in members:
                self._ensure_column_index_building(
                    m["source_id"], column, self._index_table_for(src, column, m["table_name"], m["source_id"])
                )
            union_sql = " UNION ALL ".join(
                f"SELECT {q(column)} AS val FROM {self._member_from(src, m)}" for m in members
            )
            rows = ro.execute(
                f"SELECT val, count(*) AS n FROM ({union_sql}) GROUP BY 1 ORDER BY n DESC LIMIT ?",
                (limit,),
            ).fetchall()
        return [{"value": r["val"], "count": r["n"]} for r in rows]

    def column_max_lengths(self, source_id: int) -> dict[str, int]:
        """MAX(LENGTH(col)) per column, across all member tables — backs the
        UI's autofit-column-width shortcut. One sequential scan per member
        table, every column computed in a single pass.

        Cached per source, invalidated by row_count changing — a full scan
        over every column of a 10M-row source isn't cheap, and autofit is a
        button a user can click repeatedly (once per column double-click,
        plus a "fit all" action) without the underlying data changing at
        all between clicks."""
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, source_id)
            cached = self._maxlen_cache.get(source_id)
            if cached and cached[0] == src["row_count"]:
                return cached[1]
            cols = [c["name"] for c in src["columns"]]
            members = self._resolve_members_on(ro, source_id)
            sel = ", ".join(f"MAX(LENGTH({q(c)}))" for c in cols)
            maxes = [0] * len(cols)
            for m in members:
                row = ro.execute(f"SELECT {sel} FROM {self._member_from(src, m)}").fetchone()
                for i, v in enumerate(row):
                    if v and v > maxes[i]:
                        maxes[i] = v
        result = {c: maxes[i] for i, c in enumerate(cols)}
        self._maxlen_cache[source_id] = (src["row_count"], result)
        return result

    # -------------------------------------------------------------------- tags

    def tag_time_bounds(self, source_id: int, tag_ids: list[int] | None = None,
                        column: str | None = None) -> dict:
        """Earliest and latest timestamp among this source's tagged rows —
        what "set the timeframe filter from my findings" needs. `tag_ids`
        empty/None means any tag; `column` narrows which datetime column
        counts, else every datetime column does (same all-columns semantics
        as the timeframe filter itself, so the range this fills is
        guaranteed to cover the rows it was derived from). Values are
        TS_NORMALIZE'd, i.e. exactly the shape the timeframe filter
        compares through."""
        src = self._source_lite(source_id)
        colnames = {c["name"]: c["type"] for c in src["columns"]}
        if column and colnames.get(column) == "datetime":
            ts_cols = [column]
        else:
            ts_cols = [c for c, t in colnames.items() if t == "datetime"]
        if not ts_cols:
            return {"start": None, "end": None}
        ids = [int(t) for t in (tag_ids or [])]
        tag_clause = f" AND tag_id IN ({','.join('?' * len(ids))})" if ids else ""
        lo = hi = None
        with self._reader() as ro:
            for m in self._resolve_members_on(ro, source_id):
                msrc = self._source_lite_on(ro, m["source_id"])
                member_names = {c["name"] for c in msrc["columns"]}
                for c in ts_cols:
                    if c not in member_names:
                        continue
                    row = ro.execute(
                        f"SELECT MIN(TS_NORMALIZE({q(c)})), MAX(TS_NORMALIZE({q(c)})) "
                        f"FROM {self._from_clause(msrc)} "
                        f"WHERE rid IN (SELECT rid FROM row_tags WHERE source_id=?{tag_clause})",
                        [m["source_id"], *ids],
                    ).fetchone()
                    if row[0] is not None and (lo is None or row[0] < lo):
                        lo = row[0]
                    if row[1] is not None and (hi is None or row[1] > hi):
                        hi = row[1]
        return {"start": lo, "end": hi}

    def list_tags(self) -> list[dict]:
        with self.lock:
            return [dict(r) for r in self.db.execute("SELECT * FROM tag_defs ORDER BY id")]

    def upsert_tag(self, tag_id: int | None, name: str, color: str, hotkey: str | None) -> dict:
        with self.lock, self.db:
            if tag_id:
                self.db.execute(
                    "UPDATE tag_defs SET name=?, color=?, hotkey=? WHERE id=?",
                    (name, color, hotkey, tag_id),
                )
            else:
                cur = self.db.execute(
                    "INSERT INTO tag_defs(name, color, hotkey) VALUES (?,?,?)",
                    (name, color, hotkey),
                )
                tag_id = cur.lastrowid
            row = self.db.execute("SELECT * FROM tag_defs WHERE id=?", (tag_id,)).fetchone()
        return dict(row)

    def delete_tag(self, tag_id: int) -> None:
        with self.lock, self.db:
            self._drop_undo_for_tag(tag_id)
            self.db.execute("DELETE FROM row_tags WHERE tag_id=?", (tag_id,))
            self.db.execute("DELETE FROM tag_defs WHERE id=?", (tag_id,))

    # ------------------------------------------------------------------ undo

    # How many tag changes deep Ctrl+Z goes, and the ceiling on how many
    # delta rows the whole history is allowed to retain. The row budget is
    # the one that actually bites: "tag every row in this 1.2M-row view"
    # records 1.2M pairs, and a 25-deep stack of those would put half a
    # gigabyte of scratch in /dev/shm. Whichever limit trips first evicts
    # from the oldest end.
    UNDO_LIMIT = 25
    UNDO_ROW_BUDGET = 5_000_000

    def _apply_tag_change(self, *, tag_id: int, on: bool, source_id: int, scope: str,
                          target_sql: str | None = None, target_params: Sequence = (),
                          pairs: Sequence[Sequence[int]] | None = None,
                          count_ids: Sequence[int] | None = None) -> int:
        """The single write path behind every tag apply/remove.

        The target — the rows the analyst asked to change — arrives either
        as a SELECT yielding (source_id, rid) (`target_sql`, the whole-view
        paths) or as an explicit pair list (`pairs`, the selection paths).
        What actually *changes* is a subset either way: tagging skips rows
        that already carry the tag, untagging skips rows that don't.

        Recording that delta rather than the target is the whole reason
        undo can be correct. Undoing "tag these 200 rows" by deleting the
        tag from all 200 would strip it off the rows that already had it
        beforehand — a silent, unnoticeable corruption of the analyst's own
        triage state, which is the worst failure mode this tool has (the
        same trap `tag_view`'s `exclude` argument exists to avoid, one
        level up).

        The delta lands in a `v.undo_<n>` table and the change is then
        applied *from* that table, so the rows recorded and the rows
        written are the same set by construction rather than by two queries
        agreeing. Returns how many rows actually changed.

        CLAUDE.md invariant #7 — every tag write goes through here; nothing
        INSERTs or DELETEs row_tags directly.

        Caller must hold self.lock and be inside a self.db transaction."""
        self._undo_seq += 1
        table = f"undo_{self._undo_seq}"
        if pairs is not None:
            # Explicit pairs go in through executemany and are then filtered
            # down in place, rather than being inlined as a VALUES list: a
            # select-all-then-tag can carry tens of thousands of rows, and
            # two bound parameters apiece blows past SQLITE_MAX_VARIABLE_NUMBER
            # long before it blows past anything else.
            self.db.execute(
                f"CREATE TABLE v.{q(table)} (source_id INTEGER NOT NULL, rid INTEGER NOT NULL, "
                f"PRIMARY KEY (source_id, rid)) WITHOUT ROWID"
            )
            self.db.executemany(
                f"INSERT OR IGNORE INTO v.{q(table)}(source_id, rid) VALUES (?,?)",
                [(int(sid), int(rid)) for sid, rid in pairs],
            )
            # Delete the rows this change *wouldn't* move: already tagged on
            # the way in, not tagged on the way out.
            self.db.execute(
                f"DELETE FROM v.{q(table)} WHERE {'EXISTS' if on else 'NOT EXISTS'} "
                f"(SELECT 1 FROM row_tags rt WHERE rt.source_id=v.{q(table)}.source_id "
                f"AND rt.rid=v.{q(table)}.rid AND rt.tag_id=?)",
                (tag_id,),
            )
        else:
            # Keep the rows it *would* move — the mirror of the clause above.
            self.db.execute(
                f"CREATE TABLE v.{q(table)} AS "
                f"SELECT DISTINCT t.source_id AS source_id, t.rid AS rid FROM ({target_sql}) t "
                f"WHERE {'NOT EXISTS' if on else 'EXISTS'} "
                f"(SELECT 1 FROM row_tags rt WHERE rt.source_id=t.source_id "
                f"AND rt.rid=t.rid AND rt.tag_id=?)",
                (*target_params, tag_id),
            )
        changed = self.db.execute(f"SELECT count(*) FROM v.{q(table)}").fetchone()[0]
        self._write_tag_delta(table, tag_id, on)
        if changed:
            self._undo.append({
                "undo_id": self._undo_seq, "table": table, "tag_id": tag_id,
                "on": on, "count": changed, "source_id": source_id, "scope": scope,
                "count_ids": list(count_ids) if count_ids is not None else None,
                "at": time.time(),
            })
            self._trim_undo()
        else:
            # Nothing moved (re-tagging already-tagged rows). An entry here
            # would make Ctrl+Z a no-op that still consumed a press, which
            # reads as "undo is broken" rather than "there was nothing to
            # undo" — so drop it rather than record it.
            self.db.execute(f"DROP TABLE IF EXISTS v.{q(table)}")
        return changed

    def _write_tag_delta(self, table: str, tag_id: int, on: bool) -> None:
        """Apply (or re-apply) one delta table's rows in the given
        direction. Both the original write and its undo go through here —
        undo is the same operation with `on` flipped, not a second
        implementation of it."""
        if on:
            self.db.execute(
                f"INSERT OR IGNORE INTO row_tags(source_id, rid, tag_id) "
                f"SELECT source_id, rid, ? FROM v.{q(table)}",
                (tag_id,),
            )
        else:
            self.db.execute(
                f"DELETE FROM row_tags WHERE tag_id=? AND (source_id, rid) IN "
                f"(SELECT source_id, rid FROM v.{q(table)})",
                (tag_id,),
            )

    def _trim_undo(self) -> None:
        """Evict from the oldest end until both limits hold. Caller must
        hold self.lock and be inside a self.db transaction."""
        total = sum(e["count"] for e in self._undo)
        while self._undo and (len(self._undo) > self.UNDO_LIMIT or total > self.UNDO_ROW_BUDGET):
            dead = self._undo.pop(0)
            total -= dead["count"]
            self.db.execute(f"DROP TABLE IF EXISTS v.{q(dead['table'])}")

    def _drop_undo_for_tag(self, tag_id: int) -> None:
        """Deleting a tag definition takes its assignments with it, so every
        history entry naming that tag is now describing a change to
        something that no longer exists. Undoing one would resurrect rows
        pointing at a dead tag_id. Caller holds self.lock, inside a
        transaction."""
        keep = []
        for e in self._undo:
            if e["tag_id"] == tag_id:
                self.db.execute(f"DROP TABLE IF EXISTS v.{q(e['table'])}")
            else:
                keep.append(e)
        self._undo = keep

    def _undo_entry_label(self, entry: dict) -> str:
        row = self.db.execute("SELECT name FROM tag_defs WHERE id=?", (entry["tag_id"],)).fetchone()
        name = row["name"] if row else "deleted tag"
        n = entry["count"]
        verb = "Tag" if entry["on"] else "Untag"
        return f"{verb} {n:,} row{'' if n == 1 else 's'} · {name}"

    def undo_peek(self) -> dict:
        """What Ctrl+Z would undo, for the menu label and enabled state."""
        with self.lock:
            if not self._undo:
                return {"available": False, "depth": 0}
            e = self._undo[-1]
            return {
                "available": True, "depth": len(self._undo),
                "label": self._undo_entry_label(e), "tag_id": e["tag_id"],
                "on": e["on"], "count": e["count"], "source_id": e["source_id"],
            }

    def undo_last_tag_change(self) -> dict:
        """Reverse the most recent tag change, exactly and only over the
        rows it changed. Returns the source it touched (so the frontend
        knows whether the visible table even needs repainting) alongside
        the usual fresh counts."""
        with self.lock, self.db:
            if not self._undo:
                raise ValueError("Nothing to undo")
            e = self._undo[-1]
            label = self._undo_entry_label(e)
            self._write_tag_delta(e["table"], e["tag_id"], not e["on"])
            self._undo.pop()
            self.db.execute(f"DROP TABLE IF EXISTS v.{q(e['table'])}")
            ids = e["count_ids"]
            if ids is None:
                ids = [m["source_id"] for m in self._resolve_members_on(self.db, e["source_id"])]
            out = self._tag_counts_for_ids(ids)
        out.update({
            "undone": label, "affected": e["count"], "tag_id": e["tag_id"],
            "source_id": e["source_id"], "was_on": e["on"],
        })
        out.update({"next": self.undo_peek()})
        return out

    def set_tags(self, source_id: int, rids: list[int], tag_id: int, on: bool) -> dict:
        if not rids:
            return self.tag_counts(source_id)
        with self.lock, self.db:
            self._apply_tag_change(
                tag_id=tag_id, on=on, source_id=source_id, scope="rows",
                pairs=[(source_id, r) for r in rids],
            )
        return self.tag_counts(source_id)

    def set_tags_pairs(self, pairs: list[list[int]], tag_id: int, on: bool) -> dict:
        """Same as set_tags, but for a merged view where selected rows can
        belong to different real source_ids — each pair is that row's own
        (source_id, rid), never the merge's synthetic negative id."""
        if not pairs:
            return {"counts": {}}
        with self.lock, self.db:
            self._apply_tag_change(
                tag_id=tag_id, on=on, source_id=int(pairs[0][0]), scope="rows",
                pairs=pairs, count_ids=sorted({int(sid) for sid, _ in pairs}),
            )
        member_ids = sorted({sid for sid, _ in pairs})
        ph = ",".join("?" * len(member_ids))
        with self.lock:
            rows = self.db.execute(
                f"SELECT tag_id, count(*) n FROM row_tags WHERE source_id IN ({ph}) GROUP BY 1",
                member_ids,
            ).fetchall()
        return {"counts": {str(r["tag_id"]): r["n"] for r in rows}}

    def tag_view(self, view_id: str, tag_id: int, on: bool, exclude: list[Iterable[int]] | None = None) -> dict:
        """Tag every row in a materialised view — the 'filter to it, then mark
        the lot' move that filtering exists for in the first place. Uses
        each row's own source_id from the view now, so this works unchanged
        for merged views (a single INSERT/DELETE, no per-member looping).

        `exclude` is a list of (source_id, rid) pairs to leave alone. It's
        what backs "select all, then uncheck a few, then tag": the frontend
        models a select-all as a flag plus an exclusion set rather than a
        materialised Set of every position, and without this it would have
        to fall back to sending millions of individual rids (or, worse, tag
        everything and then untag the exclusions — which would strip the tag
        off an excluded row that legitimately already had it). Excluded
        rows are by construction rows the analyst had on screen to uncheck,
        so this list is small; `affected` accounts for them without a
        second count, since they came out of this same view."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        if handle.get("kind") == "group_virtual":
            return self._tag_virtual_group(handle, tag_id, on, exclude)
        if handle.get("kind") == "root_virtual":
            return self._tag_virtual_root(handle, tag_id, on, exclude)
        skip_sql, skip_params = self._exclude_clause(exclude, "source_id", "rid")
        with self.lock, self.db:
            changed = self._apply_tag_change(
                tag_id=tag_id, on=on, source_id=handle["source_id"], scope="view",
                target_sql=f"SELECT source_id, rid FROM v.{q(view_id)}{skip_sql}",
                target_params=skip_params,
                count_ids=[m["source_id"] for m in self._resolve_members_on(self.db, handle["source_id"])],
            )
        out = self.tag_counts(handle["source_id"])
        out["affected"] = max(0, handle["row_count"] - len(exclude or []))
        out["changed"] = changed
        return out

    @staticmethod
    def _exclude_clause(exclude: list[Iterable[int]] | None, sid_col: str, rid_col: str) -> tuple[str, list[int]]:
        """` WHERE (sid, rid) NOT IN ((?,?), ...)` for a small exclusion list,
        or ('', []) when there's nothing to exclude."""
        pairs = [(int(sid), int(rid)) for sid, rid in (exclude or [])]
        if not pairs:
            return "", []
        tuples = ",".join(["(?,?)"] * len(pairs))
        params: list[int] = []
        for sid, rid in pairs:
            params.extend((sid, rid))
        return f" WHERE ({sid_col}, {rid_col}) NOT IN ({tuples})", params

    def _tag_virtual_group(self, handle: dict, tag_id: int, on: bool,
                           exclude: list[Iterable[int]] | None = None) -> dict:
        member = handle["members"][0]
        sid = member["source_id"]
        src = self._source_lite(sid)
        member_from = self._member_from(src, member)
        where_sql, where_params = self._virtual_group_where(handle)
        # A virtual group is always one real source, so only that source's
        # rids in the exclusion list can possibly be in it.
        skip_rids = [int(rid) for s, rid in (exclude or []) if int(s) == sid]
        if skip_rids:
            where_sql += f" AND rid NOT IN ({','.join('?' * len(skip_rids))})"
            where_params = [*where_params, *skip_rids]
        with self.lock, self.db:
            changed = self._apply_tag_change(
                tag_id=tag_id, on=on, source_id=sid, scope="view",
                target_sql=f"SELECT ? AS source_id, rid FROM {member_from} WHERE {where_sql}",
                target_params=[sid, *where_params],
                count_ids=[m["source_id"] for m in self._resolve_members_on(self.db, sid)],
            )
        out = self.tag_counts(sid)
        out["affected"] = max(0, handle["row_count"] - len(skip_rids))
        out["changed"] = changed
        return out

    def _tag_virtual_root(self, handle: dict, tag_id: int, on: bool,
                          exclude: list[Iterable[int]] | None = None) -> dict:
        """Tagging the whole view is tagging the whole (unfiltered) source
        table directly — same shape as _tag_virtual_group, but with no
        column=value condition at all since a root_virtual view is every
        row. No materialise needed regardless of view size: ordering never
        enters into "tag every row"."""
        sid = handle["source_id"]
        table = self._source_lite(sid)["table_name"]
        skip_rids = [int(rid) for s, rid in (exclude or []) if int(s) == sid]
        skip_sql = f" WHERE rid NOT IN ({','.join('?' * len(skip_rids))})" if skip_rids else ""
        with self.lock, self.db:
            changed = self._apply_tag_change(
                tag_id=tag_id, on=on, source_id=sid, scope="view",
                target_sql=f"SELECT ? AS source_id, rid FROM {q(table)}{skip_sql}",
                target_params=[sid, *skip_rids],
                count_ids=[m["source_id"] for m in self._resolve_members_on(self.db, sid)],
            )
        out = self.tag_counts(sid)
        out["affected"] = max(0, handle["row_count"] - len(skip_rids))
        out["changed"] = changed
        return out

    def tag_counts(self, source_id: int) -> dict:
        member_ids = [m["source_id"] for m in self._resolve_members(source_id)]
        with self.lock:
            return self._tag_counts_for_ids(member_ids)

    def _tag_counts_for_ids(self, member_ids: Sequence[int]) -> dict:
        """Counts over an explicit member list. Split out of tag_counts
        because undo has to recount from inside its own open transaction
        (self.lock is not re-entrant) and over the exact member set the
        original change reported against — a merged view's ribbon counts
        every member, not just the one the changed rows happened to
        belong to."""
        ids = [int(i) for i in member_ids]
        if not ids:
            return {"counts": {}}
        ph = ",".join("?" * len(ids))
        rows = self.db.execute(
            f"SELECT tag_id, count(*) n FROM row_tags WHERE source_id IN ({ph}) GROUP BY 1", ids,
        ).fetchall()
        return {"counts": {str(r["tag_id"]): r["n"] for r in rows}}

    def tag_counts_in_view(self, view_id: str) -> dict:
        """The same shape tag_counts returns, but counting only the rows
        inside one view — so the tag ribbon can answer "how many of these
        are tagged X" rather than "how many rows in the whole table are
        tagged X", which is a different and usually less interesting
        question once a filter or a search is on.

        Scope is the view exactly as built, tag filter included: a ribbon
        that quietly dropped one of the filters in play would be reporting a
        count for a view nobody is looking at. The frontend keeps the
        whole-table counts alongside these and shows both, rather than
        picking one and hoping the analyst infers which it meant.

        Cost is the same join tag_positions makes, with the same
        untagged-source short-circuit in front of it — one indexed probe per
        member — because this runs after every view build and an untagged
        source is the common case. An unfiltered root_virtual view is every
        row of the source by construction, so it skips straight to the
        plain per-source counts with no join at all."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        if handle.get("kind") == "root_virtual":
            return self.tag_counts(handle["source_id"])
        with self._reader() as ro, self._dropped_view_is_expired():
            if handle.get("kind") == "group_virtual":
                member = handle["members"][0]
                sid = member["source_id"]
                msrc = self._source_lite_on(ro, sid)
                where_sql, where_params = self._virtual_group_where(handle, ro)
                # `rid IN (subquery)` rather than a join: _virtual_group_where
                # builds unqualified column references (it has to — a derived
                # column reaches its sidecar through USING(rid)), and row_tags
                # has a `rid` of its own that would make them ambiguous.
                rows = ro.execute(
                    "SELECT tag_id, count(*) n FROM row_tags WHERE source_id=? AND rid IN "
                    f"(SELECT rid FROM {self._member_from(msrc, member)} WHERE {where_sql}) GROUP BY 1",
                    [sid, *where_params],
                ).fetchall()
                return {"counts": {str(r["tag_id"]): r["n"] for r in rows}}
            any_tags = any(
                ro.execute("SELECT 1 FROM row_tags WHERE source_id=? LIMIT 1", (m["source_id"],)).fetchone()
                for m in self._resolve_members_on(ro, handle["source_id"])
            )
            if not any_tags:
                return {"counts": {}}
            rows = ro.execute(
                f"SELECT rt.tag_id, count(*) n FROM v.{q(view_id)} vv "
                f"JOIN row_tags rt ON rt.rid = vv.rid AND rt.source_id = vv.source_id GROUP BY 1"
            ).fetchall()
        return {"counts": {str(r["tag_id"]): r["n"] for r in rows}}

    def set_note(self, source_id: int, rid: int, note: str) -> None:
        with self.lock, self.db:
            if note.strip():
                self.db.execute(
                    "INSERT INTO row_notes(source_id, rid, note) VALUES (?,?,?) "
                    "ON CONFLICT(source_id, rid) DO UPDATE SET note=excluded.note",
                    (source_id, rid, note),
                )
            else:
                self.db.execute(
                    "DELETE FROM row_notes WHERE source_id=? AND rid=?", (source_id, rid)
                )

    # ------------------------------------------------------------ layout/views

    def save_layout(self, source_id: int, payload: dict) -> None:
        with self.lock, self.db:
            self.db.execute(
                "INSERT INTO layouts(source_id, payload) VALUES (?,?) "
                "ON CONFLICT(source_id) DO UPDATE SET payload=excluded.payload",
                (source_id, json.dumps(payload)),
            )

    def get_layout(self, source_id: int) -> dict | None:
        with self.lock:
            row = self.db.execute(
                "SELECT payload FROM layouts WHERE source_id=?", (source_id,)
            ).fetchone()
        return json.loads(row["payload"]) if row else None

    def save_view(self, source_id: int, name: str, payload: dict) -> dict:
        with self.lock, self.db:
            cur = self.db.execute(
                "INSERT INTO saved_views(source_id, name, payload, saved_at) VALUES (?,?,?,?)",
                (source_id, name, json.dumps(payload), time.strftime("%Y-%m-%dT%H:%M:%S")),
            )
        return {"id": cur.lastrowid, "name": name, "payload": payload}

    def list_saved_views(self, source_id: int) -> list[dict]:
        with self.lock:
            rows = self.db.execute(
                "SELECT * FROM saved_views WHERE source_id=? ORDER BY id DESC", (source_id,)
            ).fetchall()
        return [{"id": r["id"], "name": r["name"], "payload": json.loads(r["payload"])} for r in rows]

    def delete_saved_view(self, view_id: int) -> None:
        with self.lock, self.db:
            self.db.execute("DELETE FROM saved_views WHERE id=?", (view_id,))

    # ------------------------------------------------------- sql pane sub-tabs

    def list_sql_tabs(self) -> list[dict]:
        with self.lock:
            rows = self.db.execute(
                "SELECT id, name, sql, pos FROM sql_tabs ORDER BY pos, id"
            ).fetchall()
        return [dict(r) for r in rows]

    def create_sql_tab(self, name: str, sql: str = "") -> dict:
        """New tab lands at the right end of the strip (max(pos)+1)."""
        with self.lock, self.db:
            pos = (self.db.execute("SELECT COALESCE(MAX(pos), -1) + 1 FROM sql_tabs").fetchone()[0])
            cur = self.db.execute(
                "INSERT INTO sql_tabs(name, sql, pos) VALUES (?,?,?)", (name, sql, pos),
            )
        return {"id": cur.lastrowid, "name": name, "sql": sql, "pos": pos}

    def update_sql_tab(self, tab_id: int, *, name: str | None = None, sql: str | None = None) -> dict:
        """Partial, same None-means-leave-alone convention as
        workspace.SavedFilters.update — a rename sends name, the editor's
        debounced autosave sends sql."""
        with self.lock, self.db:
            row = self.db.execute("SELECT id, name, sql, pos FROM sql_tabs WHERE id=?", (tab_id,)).fetchone()
            if row is None:
                raise KeyError(f"No SQL tab {tab_id}")
            rec = dict(row)
            if name is not None:
                rec["name"] = name
            if sql is not None:
                rec["sql"] = sql
            self.db.execute("UPDATE sql_tabs SET name=?, sql=? WHERE id=?", (rec["name"], rec["sql"], tab_id))
        return rec

    def delete_sql_tab(self, tab_id: int) -> None:
        with self.lock, self.db:
            self.db.execute("DELETE FROM sql_tabs WHERE id=?", (tab_id,))

    def reorder_sql_tabs(self, ordered_ids: list[int]) -> list[dict]:
        """Renumbers pos to match the given id order. Unlike
        workspace.SavedFilters.reorder (which has to preserve the relative
        order of filters *outside* the group being moved), there's only one
        flat strip here, so any id not listed just sorts after the listed
        ones, keeping its own relative order."""
        with self.lock, self.db:
            known = {r["id"] for r in self.db.execute("SELECT id FROM sql_tabs").fetchall()}
            seq = [i for i in ordered_ids if i in known]
            self.db.executemany(
                "UPDATE sql_tabs SET pos=? WHERE id=?", [(p, i) for p, i in enumerate(seq)],
            )
            # Anything unlisted keeps its order but sits after the listed run.
            rest = [i for i in sorted(known - set(seq))]
            self.db.executemany(
                "UPDATE sql_tabs SET pos=? WHERE id=?", [(len(seq) + p, i) for p, i in enumerate(rest)],
            )
        return self.list_sql_tabs()

    # ------------------------------------------ legacy filter-preset migration

    def pop_legacy_presets(self) -> list[dict]:
        """filter_presets used to be this case's own SQLite-backed table of
        saved filters, scoped to just this case file. Presets are saved
        filters now — one workspace-level, cross-case mechanism (see
        CLAUDE.md and workspace.SavedFilters) instead of two overlapping
        ones. Reads out whatever's left in a case file created before this
        change and clears the table, so this is a one-time migration on
        first open rather than an ongoing read on every open (the table's
        empty for good afterward — including for a case file that's always
        been on the new scheme, where this is just a cheap empty SELECT).
        The caller (server.py, on case open) folds the result into
        workspace.filters."""
        with self.lock, self.db:
            rows = self.db.execute("SELECT * FROM filter_presets").fetchall()
            if rows:
                self.db.execute("DELETE FROM filter_presets")
        return [
            {
                "name": r["name"],
                "col_names": json.loads(r["col_names"]),
                "payload": json.loads(r["payload"]),
            }
            for r in rows
        ]

    # ---------------------------------------------------------------- sessions

    def export_session(self, source_id: int) -> dict:
        src = self.get_source(source_id)
        with self.lock:
            tags = [dict(r) for r in self.db.execute("SELECT * FROM tag_defs ORDER BY id")]
            rt = [
                {"rid": r["rid"], "tag_id": r["tag_id"]}
                for r in self.db.execute(
                    "SELECT rid, tag_id FROM row_tags WHERE source_id=? ORDER BY rid", (source_id,)
                )
            ]
            notes = [
                {"rid": r["rid"], "note": r["note"]}
                for r in self.db.execute(
                    "SELECT rid, note FROM row_notes WHERE source_id=? ORDER BY rid", (source_id,)
                )
            ]
        return {
            "format": "winnow-session/1",
            "exported_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "source": {
                "name": src["name"],
                "path": src["path"],
                "file_hash": src["file_hash"],
                "row_count": src["row_count"],
                # The imported file's own columns — this is what
                # import_session matches against, and a derived column the
                # exporting analyst happened to add isn't part of the
                # evidence's shape.
                "columns": self._base_cols(src),
            },
            "tag_defs": tags,
            "row_tags": rt,
            "row_notes": notes,
            "layout": self.get_layout(source_id),
            "saved_views": self.list_saved_views(source_id),
            # Definitions, not values: a derived column is a deterministic
            # function of data the receiving case already has, so importing
            # re-runs the backfill rather than shipping a second copy of
            # every timestamp.
            "derived_columns": [
                {k: d[k] for k in ("name", "input_column", "op_id", "params")}
                for d in self.list_derived_columns(source_id)
            ],
        }

    def import_session(self, source_id: int, session: dict, merge: bool = True) -> dict:
        src = self.get_source(source_id)
        warnings = []
        s_src = session.get("source", {})
        if s_src.get("file_hash") and s_src["file_hash"] != src["file_hash"]:
            warnings.append(
                "Session was recorded against a different file. Row numbers may not line up."
            )
        # Map session tag ids onto local tag ids by name, creating what's missing.
        local = {t["name"].lower(): t["id"] for t in self.list_tags()}
        idmap = {}
        for t in session.get("tag_defs", []):
            key = t["name"].lower()
            if key not in local:
                new = self.upsert_tag(None, t["name"], t.get("color", "#888888"), t.get("hotkey"))
                local[key] = new["id"]
            idmap[t["id"]] = local[key]

        with self.lock, self.db:
            if not merge:
                self.db.execute("DELETE FROM row_tags WHERE source_id=?", (source_id,))
                self.db.execute("DELETE FROM row_notes WHERE source_id=?", (source_id,))
            self.db.executemany(
                "INSERT OR IGNORE INTO row_tags(source_id, rid, tag_id) VALUES (?,?,?)",
                [
                    (source_id, r["rid"], idmap.get(r["tag_id"], r["tag_id"]))
                    for r in session.get("row_tags", [])
                ],
            )
            self.db.executemany(
                "INSERT INTO row_notes(source_id, rid, note) VALUES (?,?,?) "
                "ON CONFLICT(source_id, rid) DO UPDATE SET note=excluded.note",
                [(source_id, r["rid"], r["note"]) for r in session.get("row_notes", [])],
            )
        if session.get("layout"):
            self.save_layout(source_id, session["layout"])
        for sv in session.get("saved_views", []):
            self.save_view(source_id, sv["name"], sv["payload"])

        # Derived columns arrive as definitions and are recomputed here —
        # each starts its own backfill job. A name that's already taken
        # (the analyst built the same column locally, or a base column has
        # that name) is a warning, never a failure: the tags and notes are
        # the part of a session that can't be reconstructed.
        derived_added = 0
        for d in session.get("derived_columns", []):
            try:
                self.add_derived_column(source_id, d["name"], d["input_column"],
                                        d["op_id"], d.get("params"))
                derived_added += 1
            except (ValueError, KeyError) as e:
                warnings.append(f"Derived column {d.get('name')!r} not recreated: {e}")
        return {"warnings": warnings, "tags_applied": len(session.get("row_tags", [])),
                "derived_columns_added": derived_added}

    # ---------------------------------------------------------- case sessions

    def export_case_session(self) -> dict:
        """Whole-case session: every open source's tags/notes/layout, in one
        file. Builds directly on export_session — no new per-source shape."""
        return {
            "format": "winnow-case-session/1",
            "exported_at": time.strftime("%Y-%m-%dT%H:%M:%S"),
            "sources": [self.export_session(s["id"]) for s in self.list_sources()],
        }

    def import_case_session(self, session: dict, merge: bool = True) -> dict:
        """Reattaches each source session by file_hash to whatever's already
        open; if a source isn't open but its original file still exists at
        the recorded path, re-imports it first. Local single-user tool with
        full filesystem access already (same trust level as --open), so this
        is the one action that can fully reconstruct a case from a session."""
        warnings: list[str] = []
        tags_applied = 0
        sources_restored = 0
        sources_reimported = 0
        by_hash = {s["file_hash"]: s["id"] for s in self.list_sources() if s.get("file_hash")}

        for src_session in session.get("sources", []):
            s_meta = src_session.get("source", {})
            source_id = by_hash.get(s_meta.get("file_hash"))
            if source_id is None:
                path = s_meta.get("path")
                if path and os.path.isfile(path):
                    try:
                        rec = self.ingest_csv(path, name=s_meta.get("name"))
                        source_id = rec["id"]
                        sources_reimported += 1
                    except Exception as e:
                        warnings.append(f"Could not re-import {s_meta.get('name', path)}: {e}")
                        continue
                else:
                    warnings.append(
                        f"'{s_meta.get('name', 'unknown')}' isn't open and its original file "
                        f"wasn't found at {path or 'an unknown path'} — skipped"
                    )
                    continue
            res = self.import_session(source_id, src_session, merge=merge)
            warnings.extend(res["warnings"])
            tags_applied += res["tags_applied"]
            sources_restored += 1

        return {
            "warnings": warnings,
            "tags_applied": tags_applied,
            "sources_restored": sources_restored,
            "sources_reimported": sources_reimported,
        }

    _SESSION_NAME_RE = re.compile(r"[^A-Za-z0-9_ -]")

    def _sessions_dir(self) -> str:
        d = os.path.join(os.path.dirname(os.path.abspath(self.path)), "sessions")
        os.makedirs(d, exist_ok=True)
        return d

    def _session_path(self, name: str) -> str:
        safe = self._SESSION_NAME_RE.sub("_", name).strip() or "session"
        return os.path.join(self._sessions_dir(), f"{safe}.winnow_case.json")

    def save_named_session(self, name: str) -> dict:
        data = self.export_case_session()
        with open(self._session_path(name), "w", encoding="utf-8") as f:
            json.dump(data, f)
        return {"name": name, "source_count": len(data["sources"]), "saved_at": data["exported_at"]}

    def list_named_sessions(self) -> list[dict]:
        out = []
        d = self._sessions_dir()
        for fn in os.listdir(d):
            if not fn.endswith(".winnow_case.json"):
                continue
            try:
                with open(os.path.join(d, fn), "r", encoding="utf-8") as f:
                    data = json.load(f)
                out.append({
                    "name": fn[: -len(".winnow_case.json")],
                    "saved_at": data.get("exported_at"),
                    "source_count": len(data.get("sources", [])),
                })
            except (OSError, json.JSONDecodeError):
                continue
        out.sort(key=lambda s: s.get("saved_at") or "", reverse=True)
        return out

    def load_named_session(self, name: str, merge: bool = True) -> dict:
        path = self._session_path(name)
        if not os.path.isfile(path):
            raise KeyError(f"No saved session named {name!r}")
        with open(path, "r", encoding="utf-8") as f:
            data = json.load(f)
        return self.import_case_session(data, merge=merge)

    def delete_named_session(self, name: str) -> None:
        path = self._session_path(name)
        if os.path.isfile(path):
            os.remove(path)

    # --------------------------------------------------------------- exporting

    def export_view_csv(self, view_id: str, tagged_only: bool = False):
        """Validates the view eagerly, then returns a generator for the CSV
        body. Splitting this out matters: a generator function's body
        (including a KeyError check) doesn't run until first iterated, which
        would happen inside StreamingResponse — too late for the route's
        try/except to turn it into a 409."""
        handle = self._views.get(view_id)
        if not handle:
            raise KeyError("View expired — rebuild it")
        if handle.get("kind") == "group_virtual":
            return self._export_virtual_group_csv_rows(handle, tagged_only)
        if handle.get("kind") == "root_virtual":
            return self._export_virtual_root_csv_rows(handle, tagged_only)
        return self._export_view_csv_rows(view_id, handle, tagged_only)

    def _export_virtual_group_csv_rows(self, handle: dict, tagged_only: bool):
        member = handle["members"][0]
        sid = member["source_id"]
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, sid)
            cols = [c["name"] for c in src["columns"]]
            where_sql, where_params = self._virtual_group_where(handle, ro)
            sel = ", ".join(q(c) for c in cols)

            tagnames = {r["id"]: r["name"] for r in ro.execute("SELECT id, name FROM tag_defs")}
            buf = io.StringIO()
            w = csv.writer(buf, lineterminator="\n")
            w.writerow(["Line", "Tags", "Note", *cols])
            yield buf.getvalue()
            buf.seek(0), buf.truncate(0)

            rows = ro.execute(
                f"SELECT rid, {sel} FROM {self._member_from(src, member)} WHERE {where_sql} ORDER BY rid",
                where_params,
            ).fetchall()
            rids = [r["rid"] for r in rows]
            tmap: dict[int, list[str]] = {}
            nmap: dict[int, str] = {}
            if rids:
                ph = ",".join("?" * len(rids))
                for rid, tid in ro.execute(
                    f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    tmap.setdefault(rid, []).append(tagnames.get(tid, str(tid)))
                for rid, note in ro.execute(
                    f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    nmap[rid] = note

        for r in rows:
            if tagged_only and r["rid"] not in tmap:
                continue
            w.writerow([r["rid"], "; ".join(tmap.get(r["rid"], [])), _csv_safe(nmap.get(r["rid"], "")), *[_csv_safe(r[c]) for c in cols]])
            yield buf.getvalue()
            buf.seek(0), buf.truncate(0)

    def _export_virtual_root_csv_rows(self, handle: dict, tagged_only: bool):
        """Streams the whole source table in rid order, chunked the same
        keyset way _export_view_csv_rows pages v.view_N (`rid > ?` here
        instead of `pos > ?`) — a root_virtual view can be the entire
        source, potentially millions of rows, unlike the small bounded
        group _export_virtual_group_csv_rows fetches in one shot. Runs
        entirely on one checked-out _reader() connection, held across
        yields (safe — see _reader's docstring; an abandoned generator's
        GeneratorExit closes it), so a multi-minute export never touches
        self.lock at all."""
        sid = handle["source_id"]
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, sid)
            table = src["table_name"]
            cols = [c["name"] for c in src["columns"]]
            sel = ", ".join(q(c) for c in cols)

            tagnames = {r["id"]: r["name"] for r in ro.execute("SELECT id, name FROM tag_defs")}
            buf = io.StringIO()
            w = csv.writer(buf, lineterminator="\n")
            w.writerow(["Line", "Tags", "Note", *cols])
            yield buf.getvalue()
            buf.seek(0), buf.truncate(0)

            last_rid = 0
            while True:
                chunk = ro.execute(
                    f"SELECT rid, {sel} FROM {self._from_clause(src)} WHERE rid > ? ORDER BY rid LIMIT 5000",
                    (last_rid,),
                ).fetchall()
                if not chunk:
                    break
                last_rid = chunk[-1]["rid"]
                rids = [r["rid"] for r in chunk]
                ph = ",".join("?" * len(rids))
                tmap: dict[int, list[str]] = {}
                nmap: dict[int, str] = {}
                if tagged_only:
                    tagged_set = {
                        row[0] for row in ro.execute(
                            f"SELECT DISTINCT rid FROM row_tags WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                        )
                    }
                else:
                    tagged_set = None
                for rid, tid in ro.execute(
                    f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    tmap.setdefault(rid, []).append(tagnames.get(tid, str(tid)))
                for rid, note in ro.execute(
                    f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                ):
                    nmap[rid] = note

                for r in chunk:
                    if tagged_only and r["rid"] not in tagged_set:
                        continue
                    w.writerow([r["rid"], "; ".join(tmap.get(r["rid"], [])), _csv_safe(nmap.get(r["rid"], "")),
                               *[_csv_safe(r[c]) for c in cols]])
                yield buf.getvalue()
                buf.seek(0), buf.truncate(0)

    def _export_view_csv_rows(self, view_id: str, handle: dict, tagged_only: bool):
        """Streams chunks of v.view_N grouped by each row's own source_id —
        same per-member resolution pattern as fetch_rows, so this works for
        merged views without a single cross-table JOIN.

        Runs entirely on one checked-out _reader() connection, held across
        yields — safe, unlike the RLock this used to have to release before
        every yield (a generator can be resumed or abandoned from a
        different thread than the one that acquired the lock; a connection
        is just an object, and GeneratorExit on early client disconnect
        closes it via _reader's exception path). Keyset pagination
        (`pos > ?`) over v.view_N's unique index avoids the
        OFFSET-rescans-from-zero cost that would otherwise reintroduce.
        """
        with self._reader() as ro, self._dropped_view_is_expired():
            src = self._source_lite_on(ro, handle["source_id"])
            cols = [c["name"] for c in src["columns"]]
            sel = ", ".join(q(c) for c in cols)

            tagnames = {r["id"]: r["name"] for r in ro.execute("SELECT id, name FROM tag_defs")}
            buf = io.StringIO()
            w = csv.writer(buf, lineterminator="\n")
            w.writerow(["Line", "Tags", "Note", *cols])
            yield buf.getvalue()
            buf.seek(0), buf.truncate(0)

            last_pos = 0
            while True:
                chunk = ro.execute(
                    f"SELECT pos, source_id, rid FROM v.{q(view_id)} WHERE pos > ? ORDER BY pos LIMIT 5000",
                    (last_pos,),
                ).fetchall()
                if not chunk:
                    break
                last_pos = chunk[-1]["pos"]

                by_source: dict[int, list[int]] = {}
                for r in chunk:
                    by_source.setdefault(r["source_id"], []).append(r["rid"])

                cellmap: dict[tuple[int, int], tuple] = {}
                tmap: dict[tuple[int, int], list[str]] = {}
                nmap: dict[tuple[int, int], str] = {}
                for sid, rids in by_source.items():
                    if tagged_only:
                        ph0 = ",".join("?" * len(rids))
                        tagged_set = {
                            row[0] for row in ro.execute(
                                f"SELECT DISTINCT rid FROM row_tags WHERE source_id=? AND rid IN ({ph0})",
                                [sid, *rids],
                            )
                        }
                        rids = [r for r in rids if r in tagged_set]
                        if not rids:
                            continue
                    member_src = self._source_lite_on(ro, sid)
                    ph = ",".join("?" * len(rids))
                    for row in ro.execute(
                        f"SELECT rid, {sel} FROM {self._from_clause(member_src)} WHERE rid IN ({ph})", rids
                    ):
                        cellmap[(sid, row["rid"])] = tuple(row[c] for c in cols)
                    for rid, tid in ro.execute(
                        f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                    ):
                        tmap.setdefault((sid, rid), []).append(tagnames.get(tid, str(tid)))
                    for rid, note in ro.execute(
                        f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})", [sid, *rids]
                    ):
                        nmap[(sid, rid)] = note

                for r in chunk:
                    key = (r["source_id"], r["rid"])
                    if key not in cellmap:
                        continue  # tagged_only filtered this row out
                    w.writerow([
                        r["rid"],
                        "; ".join(tmap.get(key, [])),
                        _csv_safe(nmap.get(key, "")),
                        *[_csv_safe(v) for v in cellmap[key]],
                    ])
                yield buf.getvalue()
                buf.seek(0), buf.truncate(0)

    def export_tagged_xlsx(self) -> io.BytesIO:
        """One worksheet per real source that has at least one tagged row —
        merges are skipped since their rows already belong to a real source
        sheet, and sources with zero tagged rows are skipped so the workbook
        doesn't fill up with empty tabs. No view/filter/sort involved (only
        "tagged or not"), so unlike _export_view_csv_rows this doesn't need
        v.view_N at all — just row_tags per source_id. Reads run on a
        _reader() connection checked out per source, so a big multi-sheet
        export never touches self.lock; openpyxl's in-memory writes below
        don't touch the db at all."""
        tagnames = {t["id"]: t["name"] for t in self.list_tags()}
        wb = Workbook()
        wb.remove(wb.active)
        used_names: set[str] = set()

        for src in self.list_sources():
            if not src.get("tagged_row_count"):
                continue
            source_id = src["id"]
            cols = [c["name"] for c in src["columns"]]
            sel = ", ".join(q(c) for c in cols)
            with self._reader() as ro, self._dropped_view_is_expired():
                rids = [r[0] for r in ro.execute(
                    "SELECT DISTINCT rid FROM row_tags WHERE source_id=? ORDER BY rid", (source_id,)
                )]
                cellmap: dict[int, tuple] = {}
                tmap: dict[int, list[str]] = {}
                nmap: dict[int, str] = {}
                if rids:
                    ph = ",".join("?" * len(rids))
                    for row in ro.execute(
                        f"SELECT rid, {sel} FROM {self._from_clause(src)} WHERE rid IN ({ph})", rids
                    ):
                        cellmap[row["rid"]] = tuple(row[c] for c in cols)
                    for rid, tid in ro.execute(
                        f"SELECT rid, tag_id FROM row_tags WHERE source_id=? AND rid IN ({ph})", [source_id, *rids]
                    ):
                        tmap.setdefault(rid, []).append(tagnames.get(tid, str(tid)))
                    for rid, note in ro.execute(
                        f"SELECT rid, note FROM row_notes WHERE source_id=? AND rid IN ({ph})", [source_id, *rids]
                    ):
                        nmap[rid] = note

            ws = wb.create_sheet(_xlsx_sheet_name(src["name"], used_names))
            ws.append(["Line", "Tags", "Note", *cols])
            for rid in rids:
                if rid not in cellmap:
                    continue
                ws.append([
                    rid,
                    "; ".join(tmap.get(rid, [])),
                    _csv_safe(nmap.get(rid, "")),
                    *[_csv_safe(v) for v in cellmap[rid]],
                ])

        if not wb.sheetnames:
            wb.create_sheet("No tagged rows").append(["No rows are tagged in this case yet."])

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)
        return buf

    def search_all_sources(self, query: str = "", terms: list[dict] | None = None) -> list[dict]:
        """Every source's match count, sorted heaviest-first — the whole
        sweep, run to completion. See _iter_search_all_sources for the
        actual scan and its lock discipline; this is the collect-it-all
        wrapper, used by the synchronous endpoint and the tests.

        start_search_all_job is the same sweep run on a background thread
        with incremental results, which is what the UI uses."""
        out = [hit for _, _, hit in self._iter_search_all_sources(query, terms) if hit]
        out.sort(key=lambda d: -d["match_count"])
        return out

    def _iter_search_all_sources(
        self, query: str = "", terms: list[dict] | None = None
    ) -> Iterator[tuple[int, int, dict | None]]:
        """Per-source match counts for a search across every real source in
        the case (merges excluded — their rows already belong to a real
        source), no filter/sort/view materialisation involved.

        Yields `(scanned, total, hit_or_None)` after each source so a caller
        can report progress and surface partial results while the sweep is
        still running (start_search_all_job) — a source with no matches
        yields a None hit rather than being skipped silently, so `scanned`
        always advances. Unsorted, in source order: sorting is the
        collecting caller's job, since a running sweep has nothing stable
        to sort yet.

        Same indexed-when-ready-else-LIKE reasoning as _compile_where's
        Contains/Advanced branches for both `query` and `terms` — including
        triggering a background index build for any source that doesn't
        have one yet, and falling back to a blob LIKE for a term
        _fts_like_pattern can't route through the trigram index.

        Two things keep this from monopolising the case for its duration:

        - The lock is taken per source's count, not once around the whole
          loop. This is a read-only sweep over every table in the case, and
          on a case whose indexes aren't built yet that's N full LIKE scans
          back to back — minutes on a 42 GB merge. Holding self.lock across
          all of them freezes every other request (paging, tagging, view
          builds) for the entire sweep, which is exactly what ingest_csv and
          build_fts already go out of their way to avoid: hold the lock for
          one unit of work, not for a whole loop.
        - Each count stops at SEARCH_ALL_COUNT_CAP rows rather than counting
          every match, bounding the worst case per source (a term matching
          most of a huge unindexed table). A capped result reports
          match_count == SEARCH_ALL_COUNT_CAP with capped=True so the UI can
          say "1,000+" rather than imply a precise number it didn't compute.
        """
        query = (query or "").strip()
        terms = [t for t in (terms or []) if (t.get("term") or "").strip()]
        if not query and not terms:
            return
        # A pasted list of IOCs is a plain OR of positive terms, and there
        # the analyst wants an entry per term per table, not one row per
        # table with everything summed into it. Anything else (mixed
        # AND/NOT from the Advanced builder, a bare `query`) keeps the
        # single union count — see _or_of_positive_terms.
        breakdown = _or_of_positive_terms(terms) if terms else None
        if breakdown and len(breakdown) > SEARCH_ALL_TERM_BREAKDOWN_MAX:
            breakdown = None
        # Snapshot the source list up front (list_sources takes the lock
        # itself); the per-source work below re-acquires it one count at a
        # time. A source removed mid-sweep just yields a SQL error we skip.
        sources = [s for s in self.list_sources() if not s.get("error")]
        total = len(sources)
        scanned = 0
        for src in sources:
            source_id = src["id"]
            table = src["table_name"]
            # Base columns only — the search-all sweep has to agree with the
            # per-source search paths, which scan build_fts's doc view.
            cols = [c["name"] for c in self._base_cols(src)]
            if not src.get("has_fts"):
                self._ensure_fts_building(source_id)
            inner, params = self._search_all_count_sql(src, table, cols, query, terms)
            n = 0
            per_term: list[dict] = []
            if inner is not None:
                # One _reader() checkout per source's count — the sweep
                # never touches self.lock at all now, so even its worst
                # case (N full LIKE scans on an unindexed 42 GB merge)
                # can't stall a single paging/tagging/view-build request.
                # Checked out per count rather than once for the sweep so
                # the connection goes back in the pool between sources.
                try:
                    with self._reader() as ro:
                        n = ro.execute(f"SELECT COUNT(*) FROM ({inner})", params).fetchone()[0]
                        # The per-term breakdown, and why it's here rather
                        # than folded into the query above: counting each
                        # term separately is the only way to say which of a
                        # pasted IOC list actually hit, and a single query
                        # with one SUM(CASE ...) per term would have to
                        # evaluate *every* term against every candidate row
                        # — losing the OR's short-circuit — on every source,
                        # including the ones that don't match at all. Gated
                        # on n instead, the sources that miss cost exactly
                        # what they cost before this existed, and only a
                        # source that already matched pays per term.
                        if n and breakdown:
                            per_term = self._search_all_term_counts(ro, src, table, cols, breakdown)
                except sqlite3.Error:
                    n = 0  # source dropped (or its index swapped) mid-sweep
                    per_term = []
            hit = None
            if n:
                hit = {
                    "source_id": source_id,
                    "name": src["name"],
                    "match_count": min(n, SEARCH_ALL_COUNT_CAP),
                    "capped": n > SEARCH_ALL_COUNT_CAP,
                    # One entry per term that actually matched this source,
                    # heaviest first. Empty when the query isn't a plain OR
                    # of terms (see _or_of_positive_terms) — the modal falls
                    # back to the single row per table it always showed.
                    # Terms can overlap on a row, so these don't have to sum
                    # to match_count.
                    "terms": per_term,
                }
            scanned += 1
            # Every source yields, matches or not, so `scanned` tracks real
            # progress through the sweep rather than only counting hits.
            yield scanned, total, hit
            # No lock to starve anyone on anymore — this is now just a
            # scheduling courtesy so a long sweep's back-to-back scans
            # don't pin a core against interactive requests, and the gap
            # cancellation checks land in.
            time.sleep(0.02)

    # ------------------------------------------------ search-all as a job

    def _search_all_count_sql(
        self, src: dict, table: str, cols: list[str], query: str, terms: list[dict] | None
    ) -> tuple[str | None, tuple]:
        """The capped `SELECT 1 FROM ... WHERE <match> LIMIT cap+1` a
        search-all count wraps in COUNT(*), for either a term list or a bare
        substring query. Returns (None, ()) when there's nothing to match on.

        Extracted so the per-term breakdown counts a term through exactly
        the same shapes the whole-query count uses — indexed doc-LIKE when
        the source has a usable trigram index, escaped blob LIKE otherwise
        (see _fts_like_pattern). A single positive term is the same question
        as a bare Contains query, which is why the breakdown reaches this
        with `terms=None` and one string rather than a one-element chip
        list."""
        source_id = src["id"]
        if terms:
            if src.get("has_fts"):
                clause, params = _advanced_fts_clause(q("fts_" + str(source_id)), _blob_expr(cols), terms)
            else:
                clause, params = _advanced_like_clause(cols, terms)
            if not clause:
                return None, ()
            return f"SELECT 1 FROM {q(table)} WHERE {clause} LIMIT {SEARCH_ALL_COUNT_CAP + 1}", tuple(params)
        if not query:
            return None, ()
        pattern = _fts_like_pattern(query)
        if src.get("has_fts") and pattern is not None:
            fts_ident = q("fts_" + str(source_id))
            return (
                f"SELECT 1 FROM {q(table)} WHERE rid IN "
                f"(SELECT rowid FROM {fts_ident} WHERE doc LIKE ?) "
                f"LIMIT {SEARCH_ALL_COUNT_CAP + 1}",
                (pattern,),
            )
        blob = _blob_expr(cols)
        return (
            f"SELECT 1 FROM {q(table)} WHERE ({blob}) LIKE ? ESCAPE '\\' "
            f"LIMIT {SEARCH_ALL_COUNT_CAP + 1}",
            (f"%{_esc_like(query)}%",),
        )

    def _search_all_term_counts(
        self, ro: sqlite3.Connection, src: dict, table: str, cols: list[str], terms: list[str]
    ) -> list[dict]:
        """One capped count per term against one source, heaviest first,
        omitting the terms that didn't match it at all — the entry-per-term
        half of a search-all hit.

        Runs on the caller's already-checked-out reader so the whole source
        (union count + breakdown) is one pool checkout, not N+1. Each count
        is capped independently, so a term's own `capped` says "1,000+ of
        *this* term", not "the source was busy" — and because a row can
        match several terms, these deliberately don't have to sum to the
        source's match_count."""
        out = []
        for term in terms:
            inner, params = self._search_all_count_sql(src, table, cols, term, None)
            if inner is None:
                continue
            n = ro.execute(f"SELECT COUNT(*) FROM ({inner})", params).fetchone()[0]
            if n:
                out.append({
                    "term": term,
                    "match_count": min(n, SEARCH_ALL_COUNT_CAP),
                    "capped": n > SEARCH_ALL_COUNT_CAP,
                })
        out.sort(key=lambda d: -d["match_count"])
        return out

    def start_search_all_job(self, query: str = "", terms: list[dict] | None = None) -> dict:
        """Runs a search_all sweep on a background daemon thread and returns
        its job record immediately, so the HTTP request that started it
        doesn't sit open for the length of the sweep.

        The sweep was already careful not to hold self.lock across the whole
        loop, so other requests were never actually blocked at the *server* —
        but the old single POST still took as long as the sweep did, which
        meant the analyst had one open modal, no results until the very end,
        and nothing to come back to if they closed it. Progress
        (scanned/total) and hits accumulate into the record as each source
        finishes, so a poller sees partial results while it's still running.

        Only one job runs at a time: starting a new one marks any live job
        cancelled (the worker checks between sources) and replaces it. There
        is no queue — a second concurrent sweep would just contend for the
        same lock with the first and make both slower.

        Same fire-and-forget daemon-thread pattern as _ensure_fts_building.
        Cancellation is cooperative and only checked between sources, so a
        cancel during one huge table's count still waits out that count."""
        with self._search_job_lock:
            if self._search_job is not None:
                self._search_job["cancelled"] = True
            self._search_job_seq += 1
            job = {
                "job_id": self._search_job_seq,
                "query": query,
                "terms": terms or [],
                "scanned": 0,
                "total": 0,
                "hits": [],
                "done": False,
                "error": None,
                "cancelled": False,
                "started_at": time.time(),
            }
            self._search_job = job
        t = threading.Thread(target=self._search_all_worker, args=(job,), daemon=True)
        with self._search_job_lock:
            self._search_job_thread = t
        t.start()
        return self.get_search_all_job(job["job_id"])

    def _search_all_worker(self, job: dict) -> None:
        try:
            for scanned, total, hit in self._iter_search_all_sources(job["query"], job["terms"]):
                with self._search_job_lock:
                    if job["cancelled"]:
                        break
                    job["scanned"] = scanned
                    job["total"] = total
                    if hit:
                        job["hits"].append(hit)
        except Exception as e:  # noqa: BLE001 — surfaced to the UI as job.error
            with self._search_job_lock:
                job["error"] = str(e)
        finally:
            with self._search_job_lock:
                job["done"] = True

    def get_search_all_job(self, job_id: int | None = None) -> dict | None:
        """Snapshot of the current (or specifically requested) job. Hits come
        back sorted heaviest-first the same way search_all_sources returns
        them, re-sorted on each poll since the running sweep appends in
        source order. Returns None when there's no such job — a poller whose
        job got superseded should stop rather than keep asking."""
        with self._search_job_lock:
            job = self._search_job
            if job is None or (job_id is not None and job["job_id"] != job_id):
                return None
            return {
                "job_id": job["job_id"],
                "scanned": job["scanned"],
                "total": job["total"],
                "hits": sorted(job["hits"], key=lambda d: -d["match_count"]),
                "done": job["done"],
                "error": job["error"],
                "cancelled": job["cancelled"],
            }

    def cancel_search_all_job(self, job_id: int | None = None) -> bool:
        with self._search_job_lock:
            job = self._search_job
            if job is None or (job_id is not None and job["job_id"] != job_id):
                return False
            job["cancelled"] = True
            return True

    def wait_for_search_all_job(self, timeout: float | None = None) -> dict | None:
        """Blocks until the running sweep finishes. Used by tests; nothing in
        the request path waits on a job."""
        with self._search_job_lock:
            t = self._search_job_thread
        if t:
            t.join(timeout)
        return self.get_search_all_job()

    # -------------------------------------------------------------- sql window

    SQL_PANE_FORBIDDEN_RE = re.compile(r"\b(attach|detach|pragma|vacuum)\b", re.IGNORECASE)

    def _merge_union_sql(self, conn: sqlite3.Connection, merge_id: int) -> str:
        """SELECT text behind the pane's merge_<id> TEMP VIEW: a UNION ALL
        of the members carrying source_id, rid and every exposed column
        (derived included, each member reading its own sidecar). Every
        table reference is main.-qualified: the pane also creates TEMP
        views SHADOWING src_<id> (see _source_view_sql), and an unqualified
        src_N inside this view would bind to that shadow — double-joining
        the sidecar and making every derived name ambiguous."""
        msrc = self._source_lite_on(conn, -merge_id)
        cols = [c["name"] for c in msrc["columns"]]
        branches = []
        for mid in msrc["member_source_ids"]:
            m = self._source_lite_on(conn, mid)
            has_drv = any(c.get("derived") for c in m["columns"])
            frm = f"main.{q(m['table_name'])} s"
            if has_drv:
                frm += f" LEFT JOIN main.{q(self._derived_table(mid))} d ON d.rid = s.rid"
            sel = ", ".join(f"{self._col_ref(m, c, 's', 'd')} AS {q(c)}" for c in cols)
            branches.append(
                f"SELECT {int(mid)} AS source_id, s.rid AS rid, {sel} FROM {frm}")
        return " UNION ALL ".join(branches)

    def _source_view_sql(self, conn: sqlite3.Connection, source_id: int) -> str:
        """SELECT text behind the pane's TEMP VIEW shadowing src_<id> for a
        source with derived columns: the table as the analyst sees it in
        the grid — rid, base columns, then derived, the sidecar joined on
        its PRIMARY KEY so SQLite drops the join entirely for queries that
        never touch a derived column. main.src_<id> stays reachable as the
        byte-faithful raw import."""
        src = self._source_lite_on(conn, source_id)
        sel = ", ".join(f"{self._col_ref(src, c['name'], 's', 'd')} AS {q(c['name'])}"
                        for c in src["columns"])
        return (f"SELECT s.rid AS rid, {sel} FROM main.{q(src['table_name'])} s "
                f"LEFT JOIN main.{q(self._derived_table(source_id))} d ON d.rid = s.rid")

    def run_sql(self, sql: str, limit: int = 5000) -> dict:
        """Read-only ad-hoc query against the case file.

        This intentionally allows arbitrary SELECT/EXPLAIN — that's the
        point of the pane — unlike validate_where_fragment's tight allowlist
        for filter fragments. But ATTACH/DETACH/PRAGMA/VACUUM serve no
        purpose for a read-only query pane and are worth blocking anyway as
        defense-in-depth: the `mode=ro` connection only restricts writes to
        *this* database file, and ATTACH opens a second, independent
        connection to whatever path it's given (modern SQLite inherits the
        read-only restriction onto attached databases too, but that's a
        SQLite-version-dependent guarantee this code shouldn't have to lean
        on). Stacked statements aren't a separate concern here: Python's
        sqlite3 already refuses to execute more than one statement per call.
        """
        structural = _blank_string_literals(sql)
        if self.SQL_PANE_FORBIDDEN_RE.search(structural):
            raise ValueError("ATTACH, DETACH, PRAGMA and VACUUM aren't allowed in the SQL pane")
        ro = sqlite3.connect(f"file:{self.path}?mode=ro", uri=True, check_same_thread=False)
        ro.row_factory = sqlite3.Row
        ro.create_function("REGEXP", 2, _regexp, deterministic=True)
        # The full trio, matching the writer/reader-pool connections: a
        # filter opened in the SQL pane (spec_sql) can legitimately contain
        # TS_NORMALIZE/DAY_BUCKET, and they're useful in hand-written pane
        # queries anyway.
        ro.create_function("DAY_BUCKET", 1, _day_bucket, deterministic=True)
        ro.create_function("TS_NORMALIZE", 1, _ts_normalize, deterministic=True)
        # Merges have no src_N of their own — expose each as a TEMP VIEW
        # merge_<id> (source_id, rid, every exposed column) so the pane can
        # query one by name (invariant #9). TEMP lives on this connection
        # only: the case file is never written, and mode=ro stays honest.
        try:
            for row in ro.execute("SELECT id FROM merges ORDER BY id"):
                try:
                    ro.execute(f"CREATE TEMP VIEW {q('merge_' + str(row['id']))} AS "
                               + self._merge_union_sql(ro, row["id"]))
                except (sqlite3.Error, KeyError):
                    continue  # one broken merge shouldn't take the pane down
        except sqlite3.Error:
            pass  # a case predating merges has no merges table to read
        # A source WITH derived columns gets a TEMP VIEW shadowing its own
        # src_<id> name (temp schema wins resolution), so `SELECT Host FROM
        # src_5` reads the derived value instead of falling into SQLite's
        # double-quoted-string trap and returning the literal 'Host'.
        # Sources without derived columns get nothing — byte-identical to
        # before. main.src_<id> is the explicit raw-table escape hatch.
        try:
            for row in ro.execute("SELECT DISTINCT source_id FROM derived_columns ORDER BY 1"):
                try:
                    ro.execute(f"CREATE TEMP VIEW {q('src_' + str(row[0]))} AS "
                               + self._source_view_sql(ro, row[0]))
                except (sqlite3.Error, KeyError):
                    continue
        except sqlite3.Error:
            pass
        try:
            t0 = time.time()
            cur = ro.execute(sql)
            rows = cur.fetchmany(limit)
            cols = [d[0] for d in cur.description] if cur.description else []
            return {
                "columns": cols,
                "rows": [[r[i] for i in range(len(cols))] for r in rows],
                "truncated": len(rows) == limit,
                "elapsed_ms": int((time.time() - t0) * 1000),
            }
        finally:
            ro.close()


_CSV_FORMULA_LEAD = ("=", "+", "-", "@", "\t", "\r")


def _csv_safe(v):
    """OWASP CSV/formula-injection mitigation: a cell that opens with
    =, +, -, @, tab or CR can execute as a formula when the export is later
    opened in Excel/Sheets. Forensic values (command lines, filenames,
    subject lines, notes) are attacker- or user-influenced text by
    definition, and this export is explicitly meant to be handed to other
    analysts/reports — exactly the CSV-injection threat model. Only the
    exported *copy* is affected; the stored case-file value is never
    touched, so evidence fidelity in the case itself is unchanged."""
    if isinstance(v, str) and v and v[0] in _CSV_FORMULA_LEAD:
        return "'" + v
    return v


def _esc_like(s: str) -> str:
    return s.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")


_XLSX_SHEET_INVALID = re.compile(r"[\\/?*\[\]:]")


def _xlsx_sheet_name(name: str, used: set[str]) -> str:
    """Excel worksheet names: no \\/?*[]:, max 31 chars, can't be blank or
    start/end with an apostrophe, and must be unique within the workbook.
    Dedup mirrors sanitize_columns' approach (store.py:137) — append a
    numeric suffix, truncating the base to make room for it."""
    clean = _XLSX_SHEET_INVALID.sub("_", (name or "Sheet").strip()).strip("'") or "Sheet"
    clean = clean[:31]
    base = clean
    n = 1
    while clean.lower() in used:
        suffix = f"_{n}"
        clean = base[: 31 - len(suffix)] + suffix
        n += 1
    used.add(clean.lower())
    return clean


def _blank_string_literals(frag: str) -> str:
    """Replace the contents of single/double-quoted literals with spaces,
    preserving length and quote delimiters, so keyword/identifier scanning
    over the result only sees SQL structure, never literal payload text."""
    out: list[str] = []
    i, n = 0, len(frag)
    while i < n:
        ch = frag[i]
        if ch in ("'", '"'):
            out.append(ch)
            i += 1
            while i < n:
                if frag[i] == ch:
                    if i + 1 < n and frag[i + 1] == ch:  # escaped '' or "" inside the literal
                        out.append("  ")
                        i += 2
                        continue
                    out.append(ch)
                    i += 1
                    break
                out.append(" ")
                i += 1
        else:
            out.append(ch)
            i += 1
    return "".join(out)


def _advanced_fts_clause(fts_ident: str, blob_expr: str, terms: list[dict]) -> tuple[str, list]:
    """Compile Advanced-mode term chips into a WHERE fragment against the
    trigram FTS5 index — one `rid IN/NOT IN (SELECT rowid FROM fts WHERE
    doc LIKE ?)` per term, combined with AND/OR at the SQL level.

    The doc-LIKE form (not MATCH) is required by the index's detail=none
    shape — a multi-trigram MATCH is a phrase query, which detail=none
    refuses outright; see build_fts. Applying NOT as SQL-level `NOT IN`
    (rather than any in-query negation) keeps NOT working in any position,
    including as the first term — an earlier MATCH-string version silently
    dropped a leading term's exclude flag because FTS5's NOT is a binary
    "and not" with no unary form.

    Per-term, not just per-query: a term _fts_like_pattern can't route
    through the index (under 3 characters, or containing a LIKE wildcard
    the unescaped pushdown form would misinterpret) gets its own escaped
    `(blob_expr) LIKE ?` atom instead, joined into the exact same AND/OR
    chain as every indexed term — such a term never silently drops out of
    the query, and never changes meaning."""
    fts_in = f"rid IN (SELECT rowid FROM {fts_ident} WHERE doc LIKE ?)"
    fts_not_in = f"rid NOT IN (SELECT rowid FROM {fts_ident} WHERE doc LIKE ?)"
    like_atom = f"({blob_expr}) LIKE ? ESCAPE '\\'"
    parts: list[str] = []
    params: list[Any] = []
    for t in terms:
        term = (t.get("term") or "").strip()
        if not term:
            continue
        exclude = t.get("exclude")
        pattern = _fts_like_pattern(term)
        if pattern is not None:
            atom = fts_not_in if exclude else fts_in
            params.append(pattern)
        else:
            atom = f"NOT {like_atom}" if exclude else like_atom
            params.append(f"%{_esc_like(term)}%")
        if not parts:
            parts.append(atom)
        else:
            connector = "OR" if str(t.get("connector", "AND")).upper() == "OR" else "AND"
            parts.append(f"{connector} {atom}")
    if not parts:
        return "", []
    return "(" + " ".join(parts) + ")", params


def _or_of_positive_terms(terms: list[dict]) -> list[str] | None:
    """The term strings, when `terms` is a plain OR of positive terms —
    otherwise None.

    That shape is exactly what the Search-all modal's "Paste a list" mode
    produces (one term per line, OR'd, nothing excluded), and it's the only
    shape where a per-term count means anything on its own: the terms are
    independent alternatives, so "which of my IOCs hit this table, and how
    hard" is a real question with a real answer. Under mixed AND/NOT the
    terms constrain each other — a count for one of them in isolation
    describes a query the analyst never asked — so those keep the single
    union count they've always had.

    Two terms or more: a one-term list's breakdown is its own total."""
    strings = [(t.get("term") or "").strip() for t in terms]
    if len(strings) < 2 or not all(strings):
        return None
    for i, t in enumerate(terms):
        if t.get("exclude"):
            return None
        if i and str(t.get("connector", "AND")).upper() != "OR":
            return None
    return strings


def _advanced_like_clause(cols: list[str], terms: list[dict]) -> tuple[str, list]:
    """LIKE-chain fallback for sources without an FTS index. Each term is its
    own '(blob) LIKE ?' atom, optionally NOT-prefixed, joined by AND/OR per
    its stated connector — standard SQL boolean precedence applies (NOT >
    AND > OR), same as a hand-typed FTS AND/OR/NOT query would behave."""
    blob = _blob_expr(cols)
    parts: list[str] = []
    params: list[str] = []
    for t in terms:
        term = (t.get("term") or "").strip()
        if not term:
            continue
        atom = f"({blob}) LIKE ? ESCAPE '\\'"
        if t.get("exclude"):
            atom = f"NOT {atom}"
        params.append(f"%{_esc_like(term)}%")
        if not parts:
            parts.append(atom)
        else:
            connector = "OR" if str(t.get("connector", "AND")).upper() == "OR" else "AND"
            parts.append(f"{connector} {atom}")
    if not parts:
        return "", []
    return "(" + " ".join(parts) + ")", params
