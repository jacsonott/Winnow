"""Reading .xlsx/.xlsm workbooks for import — a thin text-normalising
layer over openpyxl's read-only mode.

openpyxl is already a runtime dependency (the XLSX *export* writes
through it — see export_view_xlsx), so reading through it adds nothing
to the airgapped install, and its streaming read-only mode has had a
decade of real-world workbooks thrown at it: shared strings, rich-text
runs, both date epochs (1900 with its leap-year bug, and 1904-mode Mac
workbooks), style-based date detection, sparse rows padded by cell
reference. Hand-rolling that against ECMA-376 would re-earn each of
those behaviors one bug at a time. Legacy .xls (the OLE2 binary format)
is a different beast entirely and is deliberately out of scope.

What this module owns is the conversion from openpyxl's typed cells to
the TEXT the grid stores, and the choices behind it are analyst-facing:

- **Date-styled cells arrive as ISO text** ("2024-05-01 13:07:00") —
  openpyxl surfaces them as datetime objects precisely when Excel would
  have *displayed* a date, which is the only signal there is. Seconds
  are rounded (day-serials are floats; 13:07:00 stored imprecisely
  reads back as 13:06:59.999999 otherwise); exact-midnight values and
  date-typed cells render as a bare "YYYY-MM-DD"; time-of-day cells as
  a bare "HH:MM:SS". All three shapes are ones TS_NORMALIZE accepts.
- **Numbers render without float noise** — a cell holding 42 arrives as
  "42", not "42.0", since the string is what the analyst greps.
- **Booleans render TRUE/FALSE**, matching what Excel shows.
- **Formulas yield their last cached result** (data_only=True), never a
  recalculation — and an empty string when the file carries no cache
  (i.e. was written by a tool and never opened in Excel).
- **Gaps are real**: skipped cells come back as "" in place rather than
  shifting values left under the wrong headers.

The workbook is opened strictly read-only, so the evidence-file
guarantee holds by construction.
"""

from __future__ import annotations

import datetime
from typing import Iterator

import openpyxl


def cell_text(v) -> str:
    """One cell value as the TEXT the source table stores."""
    if v is None:
        return ""
    if isinstance(v, str):
        return v
    if v is True or v is False:  # before any numeric check — bool is an int
        return "TRUE" if v else "FALSE"
    if isinstance(v, datetime.datetime):
        if v.microsecond:
            v = (v + datetime.timedelta(microseconds=500_000)).replace(microsecond=0)
        if v.hour == 0 and v.minute == 0 and v.second == 0:
            return v.strftime("%Y-%m-%d")
        return v.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(v, datetime.date):
        return v.strftime("%Y-%m-%d")
    if isinstance(v, datetime.time):
        return v.strftime("%H:%M:%S")
    if isinstance(v, datetime.timedelta):  # elapsed [h]:mm cells
        total = round(v.total_seconds())
        return f"{total // 3600}:{total % 3600 // 60:02d}:{total % 60:02d}"
    if isinstance(v, float):
        if v.is_integer() and abs(v) < 1e16:
            return str(int(v))
        return repr(v)
    return str(v)


def workbook_sheets(path: str) -> list[dict]:
    """Every data sheet, for the import picker: [{name, row_count, columns}]
    in workbook order (chartsheets never appear). row_count is *data* rows
    (the header row the ingest will consume is subtracted), taken from the
    sheet's declared dimension where present — a picker-grade estimate,
    not a promise. columns is the first non-empty row."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        out = []
        for ws in wb.worksheets:
            rows_hint = ws.max_row
            if rows_hint is None:  # no <dimension> — count the rows
                rows_hint = sum(1 for _ in ws.iter_rows(values_only=True))
            header: list[str] = []
            for row in ws.iter_rows(values_only=True):
                texts = [cell_text(v).strip() for v in row]
                if any(texts):
                    header = texts
                    break
            out.append({
                "name": ws.title,
                "row_count": max(rows_hint - 1, 0) if header else 0,
                "columns": header,
            })
        return out
    finally:
        wb.close()


def sheet_reader(path: str, sheet_name: str):
    """(row_count_hint, width_hint, rows) for one sheet — the ingest entry
    point. The hints come from the sheet's declared dimension (0 when
    absent); `rows` is a generator of per-row string lists that holds the
    workbook open until exhausted or closed, so callers that stop early
    must close() it."""
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    try:
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"No such sheet in this workbook: {sheet_name}")
        ws = wb[sheet_name]
        if not hasattr(ws, "iter_rows"):
            raise ValueError(f"{sheet_name!r} is a chart sheet, not a data sheet")
        rows_hint = ws.max_row or 0
        width_hint = ws.max_column or 0
    except BaseException:
        wb.close()
        raise

    def rows() -> Iterator[list[str]]:
        try:
            for row in ws.iter_rows(values_only=True):
                yield [cell_text(v) for v in row]
        finally:
            wb.close()

    return rows_hint, width_hint, rows()


def iter_sheet_rows(path: str, sheet_name: str) -> Iterator[list[str]]:
    return sheet_reader(path, sheet_name)[2]
