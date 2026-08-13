"""store.py export: CSV formula-injection prefixing (without mutating the
stored value), tagged_only filtering, and the tagged-rows xlsx workbook."""

from __future__ import annotations

import csv
import io

from openpyxl import load_workbook


def _export_csv_text(store, view_id, tagged_only=False):
    return "".join(store.export_view_csv(view_id, tagged_only=tagged_only))


def test_csv_export_prefixes_formula_leads_without_mutating_stored_value(store, write_csv):
    path = write_csv([
        ["Note"],
        ["=cmd|' /C calc'!A1"],
        ["+1+1"],
        ["-1+1"],
        ["@SUM(1,1)"],
        ["plain text"],
    ])
    rec = store.ingest_csv(path, name="formulas.csv", build_fts=False)
    view = store.build_view(rec["id"], {"source_id": rec["id"], "filters": [], "sort": []})

    text = _export_csv_text(store, view["view_id"])
    rows = list(csv.reader(io.StringIO(text)))
    body = rows[1:]
    values = [r[-1] for r in body]  # last column is "Note"
    prefixed = {v for v in values if v.startswith("'")}
    assert "'=cmd|' /C calc'!A1" in prefixed
    assert "'+1+1" in prefixed
    assert "'-1+1" in prefixed
    assert "'@SUM(1,1)" in prefixed
    assert "plain text" in values  # untouched

    # The exported copy is prefixed, but the stored value in the case file
    # itself must be byte-for-byte what was ingested — export never mutates
    # source data.
    stored = [r["cells"][0] for r in store.fetch_rows(view["view_id"], 0, 10)["rows"]]
    assert "=cmd|' /C calc'!A1" in stored
    assert "+1+1" in stored


def test_csv_export_tagged_only(ingested):
    store, source_id = ingested
    tag = store.upsert_tag(None, "Keep", "#ff0000", None)
    store.set_tags(source_id, [1], tag["id"], True)
    view = store.build_view(source_id, {"source_id": source_id, "filters": [], "sort": []})

    all_rows = list(csv.reader(io.StringIO(_export_csv_text(store, view["view_id"]))))
    tagged_rows = list(csv.reader(io.StringIO(_export_csv_text(store, view["view_id"], tagged_only=True))))
    assert len(all_rows) - 1 == 4  # header + 4 fixture rows
    assert len(tagged_rows) - 1 == 1
    assert tagged_rows[1][0] == "1"  # Line column is the rid
    assert tagged_rows[1][1] == "Keep"


def test_xlsx_export_one_sheet_per_tagged_source(store, write_csv):
    p1 = write_csv([["A"], ["1"], ["2"]], name="x1.csv")
    p2 = write_csv([["A"], ["3"]], name="x2.csv")  # never tagged -> no sheet
    rec1 = store.ingest_csv(p1, name="x1.csv", build_fts=False)
    store.ingest_csv(p2, name="x2.csv", build_fts=False)
    tag = store.upsert_tag(None, "T", "#000000", None)
    store.set_tags(rec1["id"], [1], tag["id"], True)

    buf = store.export_tagged_xlsx()
    wb = load_workbook(buf)
    assert wb.sheetnames == ["x1.csv"]  # only the tagged source gets a sheet
    ws = wb["x1.csv"]
    header = [c.value for c in ws[1]]
    assert header == ["Line", "Tags", "Note", "A"]
    data_rows = [[c.value for c in row] for row in ws.iter_rows(min_row=2)]
    assert data_rows == [[1, "T", None, "1"]]


def test_xlsx_export_empty_case_still_produces_a_workbook(store):
    buf = store.export_tagged_xlsx()
    wb = load_workbook(buf)
    assert wb.sheetnames == ["No tagged rows"]
